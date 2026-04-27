from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify
import requests
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from datetime import datetime
import pdfkit

# Script global de máscara de CNPJ
SCRIPT_MASCARA_CNPJ = """
<script>
document.addEventListener('DOMContentLoaded', function() {
    const inputs = document.querySelectorAll('.mascara-cnpj');
    inputs.forEach(input => {
        input.addEventListener('input', function(e) {
            let value = e.target.value.replace(/\D/g, "");
            value = value.replace(/^(\\d{2})(\\d)/, "$1.$2");
            value = value.replace(/^(\\d{2})\\.(\\d{3})(\\d)/, "$1.$2.$3");
            value = value.replace(/\\.(\\d{3})(\\d)/, ".$1/$2");
            value = value.replace(/(\\d{4})(\\d)/, "$1-$2");
            e.target.value = value;
        });
    });
});
</script>
"""

# ========================
# 🧩 MENU FLUTUANTE
# ========================
MENU_FLUTUANTE = '''
<div style="margin-bottom: 20px;">
    <div style="position: relative; display: inline-block;">
        <button id="btnMenu" style="background: #2c3e50; color: white; border: none; border-radius: 50px; padding: 10px 25px; font-size: 14px; font-weight: bold; cursor: pointer; display: flex; align-items: center; gap: 8px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">☰ Menu</button>
        <div id="dropdownMenu" style="display: none; position: absolute; top: 45px; left: 0; background: white; border-radius: 10px; box-shadow: 0 10px 25px rgba(0,0,0,0.2); min-width: 220px; overflow: hidden; border: 1px solid #eee; z-index: 100;">
            <div style="background: #34495e; color: white; padding: 10px 14px; font-weight: bold;">Navegação Rápida</div>
            <a href="/clientes" style="display: block; padding: 10px 14px; color: #333; text-decoration: none; border-bottom: 1px solid #f1f1f1;">🏠 Menu Principal</a>
            <a href="/empresas" style="display: block; padding: 10px 14px; color: #333; text-decoration: none; border-bottom: 1px solid #f1f1f1;">🏢 Clientes</a>
            <a href="/servicos" style="display: block; padding: 10px 14px; color: #333; text-decoration: none; border-bottom: 1px solid #f1f1f1;">📋 Serviços / OS</a>
            <a href="/orcamentos" style="display: block; padding: 10px 14px; color: #333; text-decoration: none; border-bottom: 1px solid #f1f1f1;">💰 Orçamentos</a>
            <a href="/estoque" style="display: block; padding: 10px 14px; color: #333; text-decoration: none; border-bottom: 1px solid #f1f1f1;">📊 Estoque</a>
            <a href="/materiais" style="display: block; padding: 10px 14px; color: #333; text-decoration: none; border-bottom: 1px solid #f1f1f1;">📦 Materiais</a>
            <a href="/fornecedores" style="display: block; padding: 10px 14px; color: #333; text-decoration: none; border-bottom: 1px solid #f1f1f1;">🚚 Fornecedores</a>
            <a href="/envios" style="display: block; padding: 10px 14px; color: #333; text-decoration: none;">📬 Rastreamento</a>
        </div>
    </div>
</div>
<script>
document.addEventListener('DOMContentLoaded', function(){
    const b = document.getElementById('btnMenu');
    const d = document.getElementById('dropdownMenu');
    if(b && d){
        b.onclick = function(e){
            e.stopPropagation();
            d.style.display = d.style.display === 'block' ? 'none' : 'block';
        };
        document.onclick = function(e){
            if(!b.contains(e.target) && !d.contains(e.target)) d.style.display = 'none';
        };
    }
});
</script>
'''

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'minha_chave_secreta_123')

# ========================
# Dados do Supabase (API)
# ========================
SUPABASE_URL = os.environ.get('SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_KEY')
headers = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json"
}

# ========================
# Funções para acessar o Supabase
# ========================
def buscar_usuario_por_login(username, password):
    try:
        url = f"{SUPABASE_URL}/rest/v1/usuarios?select=*&nome%20de%20usuario=eq.{username}&SENHA=eq.{password}"
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            dados = response.json()
            return dados[0] if len(dados) > 0 else None
        return None
    except Exception as e:
        print("Erro de conexão:", e)
        return None

def buscar_usuarios():
    try:
        url = f"{SUPABASE_URL}/rest/v1/usuarios?select=*"
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            return response.json()
        else:
            print("Erro ao buscar usuários:", response.status_code, response.text)
            return []
    except Exception as e:
        print("Erro de conexão:", e)
        return []

def criar_usuario(username, password, nivel, telefone=None):
    try:
        url = f"{SUPABASE_URL}/rest/v1/usuarios"
        dados = {
            "nome de usuario": username,
            "SENHA": password,
            "NÍVEL": nivel,
            "telefone": telefone
        }
        response = requests.post(url, json=dados, headers=headers)
        if response.status_code == 201:
            return True
        else:
            print("❌ Erro ao criar usuário:")
            print("Status:", response.status_code)
            print("Resposta:", response.text)
            return False
    except Exception as e:
        print("Erro de conexão:", e)
        return False

def excluir_usuario(id):
    try:
        url = f"{SUPABASE_URL}/rest/v1/usuarios?id=eq.{id}"
        response = requests.delete(url, headers=headers)
        if response.status_code == 204:
            return True
        else:
            print("❌ Erro ao excluir usuário:")
            print("Status:", response.status_code)
            print("Resposta:", response.text)
            return False
    except Exception as e:
        print("Erro de conexão:", e)
        return False

def criar_empresa(nome, cnpj, responsavel, telefone, whatsapp, email, endereco, bairro, cidade, estado, cep, numero,
                    entrega_endereco, entrega_numero, entrega_bairro, entrega_cidade, entrega_estado, entrega_cep):
    try:
        url = f"{SUPABASE_URL}/rest/v1/empresas"
        dados = {
            "nome_empresa": nome,
            "cnpj": cnpj,
            "responsavel": responsavel,
            "telefone": telefone,
            "whatsapp": whatsapp,
            "email": email,
            "endereco": endereco,
            "bairro": bairro,
            "cidade": cidade,
            "estado": estado,
            "cep": cep,
            "numero": numero,
            "entrega_endereco": entrega_endereco,
            "entrega_numero": entrega_numero,
            "entrega_bairro": entrega_bairro,
            "entrega_cidade": entrega_cidade,
            "entrega_estado": entrega_estado,
            "entrega_cep": entrega_cep
        }
        response = requests.post(url, json=dados, headers=headers)
        if response.status_code == 201:
            return True
        else:
            print("❌ Erro ao criar empresa:")
            print("Status:", response.status_code)
            print("Resposta:", response.text)
            return False
    except Exception as e:
        print("Erro de conexão:", e)
        return False

def buscar_empresas():
    try:
        url = f"{SUPABASE_URL}/rest/v1/empresas?select=*"
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            return response.json()
        else:
            print("Erro ao buscar empresas:", response.status_code, response.text)
            return []
    except Exception as e:
        print("Erro de conexão:", e)
        return []

def buscar_materiais():
    try:
        url = f"{SUPABASE_URL}/rest/v1/materiais?select=*"
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            return response.json()
        return []
    except:
        return []

def calcular_estoque_atual():
    try:
        url = f"{SUPABASE_URL}/rest/v1/estoque?select=material_id,quantidade,tipo&order=data_movimentacao.asc"
        response = requests.get(url, headers=headers)
        if response.status_code != 200:
            print("❌ Erro ao buscar movimentações:", response.status_code, response.text)
            return {}
        movimentacoes = response.json()
        saldo = {}
        for mov in movimentacoes:
            material_id = mov['material_id']
            quantidade = float(mov.get('quantidade', 0) or 0)
            tipo = str(mov.get('tipo', '')).strip().lower()
            if tipo == 'entrada':
                saldo[material_id] = saldo.get(material_id, 0) + quantidade
            elif tipo == 'saida':
                saldo[material_id] = saldo.get(material_id, 0) - quantidade
            else:
                print(f"⚠️ Tipo desconhecido: {tipo}")
        for mat_id in saldo:
            saldo[mat_id] = max(0, saldo[mat_id])
        print("✅ Saldo final calculado:", saldo)
        return saldo
    except Exception as e:
        print("❌ Erro ao calcular estoque:", str(e))
        return {}

def buscar_movimentacoes_com_materiais(busca=None):
    try:
        url = f"{SUPABASE_URL}/rest/v1/estoque?select=*,materiais(denominacao,unidade_medida)&order=data_movimentacao.desc"
        if busca:
            url += f"&materiais.denominacao=ilike.*{busca}*"
        response = requests.get(url, headers=headers)
        if response.status_code != 200:
            return []
        return response.json()
    except Exception as e:
        print("Erro ao buscar movimentações:", e)
        return []

def excluir_movimentacao_db(id):
    try:
        url = f"{SUPABASE_URL}/rest/v1/estoque?id=eq.{id}"
        response = requests.delete(url, headers=headers)
        return response.status_code == 204
    except:
        return False

def format_data(data_str):
    if data_str is None or not data_str:
        return ''
    return data_str[:16].replace("T", " ")

# ========================
# Funções para Fornecedores
# ========================
def buscar_fornecedores():
    try:
        url = f"{SUPABASE_URL}/rest/v1/fornecedores?select=*&order=nome.asc"
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            return response.json()
        else:
            print("Erro ao buscar fornecedores:", response.status_code, response.text)
            return []
    except Exception as e:
        print("Erro de conexão ao buscar fornecedores:", e)
        return []

def criar_fornecedor(nome, cnpj, contato, telefone, email, endereco):
    try:
        url = f"{SUPABASE_URL}/rest/v1/fornecedores"
        dados = {
            "nome": nome,
            "cnpj": cnpj,
            "contato": contato,
            "telefone": telefone,
            "email": email,
            "endereco": endereco
        }
        response = requests.post(url, json=dados, headers=headers)
        if response.status_code == 201:
            return True
        else:
            print("❌ Erro ao criar fornecedor:", response.status_code, response.text)
            return False
    except Exception as e:
        print("Erro de conexão ao criar fornecedor:", e)
        return False

def atualizar_fornecedor(id, nome, cnpj, contato, telefone, email, endereco):
    try:
        url = f"{SUPABASE_URL}/rest/v1/fornecedores?id=eq.{id}"
        dados = {
            "nome": nome,
            "cnpj": cnpj,
            "contato": contato,
            "telefone": telefone,
            "email": email,
            "endereco": endereco
        }
        response = requests.patch(url, json=dados, headers=headers)
        return response.status_code == 204
    except Exception as e:
        print("Erro de conexão ao atualizar fornecedor:", e)
        return False

def excluir_fornecedor(id):
    try:
        url = f"{SUPABASE_URL}/rest/v1/fornecedores?id=eq.{id}"
        response = requests.delete(url, headers=headers)
        return response.status_code == 204
    except Exception as e:
        print("Erro de conexão ao excluir fornecedor:", e)
        return False

# ========================
# Configurações do sistema (remetente)
# ========================
def buscar_configuracoes():
    try:
        url = f"{SUPABASE_URL}/rest/v1/configuracoes?select=*"
        response = requests.get(url, headers=headers)
        if response.status_code == 200 and response.json():
            return response.json()[0]
        else:
            return {
                "nome_remetente": "Liraprint",
                "endereco_remetente": "R. Dr. Roberto Fernandes, 81",
                "bairro_remetente": "Jardim Palmira",
                "cidade_remetente": "Guarulhos",
                "estado_remetente": "SP",
                "cep_remetente": "07076-070"
            }
    except:
        return {
            "nome_remetente": "Liraprint",
            "endereco_remetente": "R. Dr. Roberto Fernandes, 81",
            "bairro_remetente": "Jardim Palmira",
            "cidade_remetente": "Guarulhos",
            "estado_remetente": "SP",
            "cep_remetente": "07076-070"
        }

def salvar_configuracoes(config):
    try:
        url = f"{SUPABASE_URL}/rest/v1/configuracoes"
        response_check = requests.get(url, headers=headers)
        if response_check.status_code == 200 and response_check.json():
            id = response_check.json()[0]['id']
            response = requests.patch(f"{url}?id=eq.{id}", json=config, headers=headers)
            return response.status_code == 204
        else:
            response = requests.post(url, json=config, headers=headers)
            return response.status_code == 201
    except Exception as e:
        print("Erro ao salvar configurações:", e)
        return False

# ========================
# Páginas do sistema
# ========================
@app.route('/')
def index():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    return redirect(url_for('clientes'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = request.form['username']
        pwd = request.form['password']
        try:
            usuario = buscar_usuario_por_login(user, pwd)
            if usuario:
                session['usuario'] = usuario['nome de usuario']
                session['nivel'] = usuario['NÍVEL']
                session['telefone'] = usuario.get('telefone', '')
                return redirect(url_for('clientes'))
            else:
                flash("Usuário ou senha incorretos!")
        except Exception as e:
            flash("Erro ao conectar ao banco de dados.")
    
    # HTML injetado com concatenação (sem 'f' para não quebrar o CSS com chaves {})
    return '''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Login - Gráfica Rápida</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
    body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: #333; min-height: 100vh; padding: 0; margin: 0; display: flex; justify-content: center; align-items: center; }
    .login-container { background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0,0,0,0.1); width: 100%; max-width: 400px; overflow: hidden; }
    .header { background: #2c3e50; color: white; text-align: center; padding: 30px; }
    h1 { font-size: 24px; margin: 0; font-weight: 600; }
    .form-container { padding: 30px; }
    .form-container label { display: block; margin: 10px 0 5px 0; font-weight: 600; color: #2c3e50; }
    .form-container input { width: 100%; padding: 12px; border: 1px solid #ddd; border-radius: 8px; font-size: 14px; margin-bottom: 20px; }
    .btn { width: 100%; padding: 14px; background: #27ae60; color: white; border: none; border-radius: 8px; font-size: 16px; font-weight: 600; cursor: pointer; }
    .flash { background: #fdf3cd; color: #856404; padding: 12px; border-radius: 8px; margin: 15px 30px; font-size: 14px; }
    .footer { text-align: center; padding: 20px; background: #ecf0f1; color: #7f8c8d; font-size: 13px; border-top: 1px solid #bdc3c7; }
    </style>
    </head>
    <body>
    <div class="login-container">
    <div class="header"><h1>Login</h1></div>
    <form method="post" class="form-container">
    <label>Usuário</label><input type="text" name="username" required>
    <label>Senha</label><input type="password" name="password" required>
    <button type="submit" class="btn">Entrar</button>
    </form>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </div>
    </body>
    </html>
    '''

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/clientes')
def clientes():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Menu da Gráfica</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: #333; min-height: 100vh; padding: 0; margin: 0; }}
    .container {{ max-width: 900px; margin: 50px auto; background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0, 0, 0, 0.1); overflow: hidden; }}
    .header {{ background: #2c3e50; color: white; text-align: center; padding: 25px; }}
    h1 {{ font-size: 26px; margin: 0; font-weight: 600; }}
    .user-info {{ background: #34495e; color: white; padding: 12px 20px; font-size: 15px; display: flex; justify-content: space-between; align-items: center; }}
    .btn-grid {{ display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 12px; padding: 20px; }}
    @media (max-width: 768px) {{ .btn-grid {{ grid-template-columns: 1fr; }} }}
    .btn {{ display: block; width: 100%; padding: 10px 15px; font-size: 14px; font-weight: 600; text-align: center; text-decoration: none; border-radius: 8px; color: white; transition: all 0.3s ease; border: none; cursor: pointer; max-width: 250px; margin: 0 auto; }}
    .btn-green {{ background: #27ae60; }} .btn-blue {{ background: #3498db; }} .btn-purple {{ background: #8e44ad; }} .btn-red {{ background: #e74c3c; }} .btn-orange {{ background: #e67e22; }}
    .btn:hover {{ transform: translateY(-2px); box-shadow: 0 6px 12px rgba(0,0,0,0.1); }}
    .footer {{ text-align: center; padding: 15px; background: #ecf0f1; color: #7f8c8d; font-size: 12px; border-top: 1px solid #bdc3c7; }}
    </style>
    </head>
    <body>\n
    <div class="container">
    <div class="header"><h1>📋 Menu da Gráfica</h1></div>
    <div class="user-info">
    <span>👤 {session['usuario']} ({session['nivel'].upper()})</span>
    <a href="/logout">🚪 Sair</a>
    </div>
    <div class="btn-grid">
    <a href="/empresas" class="btn btn-green">🏢 Clientes / Empresas</a>
    <a href="/servicos" class="btn btn-blue">🔧 Todos os Serviços</a>
    <a href="/orcamentos" class="btn btn-blue">💰 Orçamentos</a>
    {f'<a href="/estoque" class="btn btn-purple">📊 Meu Estoque</a>' if session['nivel'] == 'administrador' else ''}
    <a href="/envios" class="btn btn-blue">📦 Rastreamento</a>
    {f'<a href="/fornecedores" class="btn btn-orange">📦 Fornecedores</a>' if session['nivel'] == 'administrador' else ''}
    {f'<a href="/configuracoes" class="btn btn-red">⚙️ Configurações</a>' if session['nivel'] == 'administrador' else ''}
    {f'<a href="/gerenciar_usuarios" class="btn btn-red">🔐 Gerenciar Usuários</a>' if session['nivel'] == 'administrador' else ''}
    {f'<a href="/exportar_excel" class="btn btn-red">📥 Exportar Backup</a>' if session['nivel'] == 'administrador' else ''}
    {f'<a href="/importar_excel" class="btn btn-red">📤 Importar Excel</a>' if session['nivel'] == 'administrador' else ''}
    </div>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </div>
    </body>
    </html>
    '''

@app.route('/gerenciar_usuarios')
def gerenciar_usuarios():
    if 'usuario' not in session or session['nivel'] != 'administrador':
        flash("Acesso negado!")
        return redirect(url_for('clientes'))
    try:
        usuarios = buscar_usuarios()
        return f'''
        <!DOCTYPE html>
        <html lang="pt-BR">
        <head>
        <meta charset="UTF-8">
        <title>Gerenciar Usuários</title>
        <style>
        body {{ font-family: Arial, sans-serif; padding: 20px; background: #f5f7fa; }}
        .container {{ max-width: 1000px; margin: 0 auto; background: white; padding: 20px; border-radius: 10px; box-shadow: 0 0 10px rgba(0,0,0,0.1); }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        th, td {{ border: 1px solid #ccc; padding: 12px; text-align: left; }}
        th {{ background: #ecf0f1; font-weight: bold; }}
        .form-container {{ background: #f9f9f9; padding: 20px; margin: 20px 0; border-radius: 5px; }}
        input, select {{ padding: 10px; margin: 5px; width: 200px; border: 1px solid #ddd; border-radius: 4px; }}
        button {{ padding: 10px 20px; background: #27ae60; color: white; border: none; border-radius: 4px; cursor: pointer; font-weight: bold; }}
        a {{ color: #3498db; text-decoration: none; }}
        .back-link {{ display: inline-block; margin: 20px 0; color: #3498db; }}
        </style>
        </head>
        <body>\n
        <div class="container">
        <h2>🔐 Gerenciar Usuários</h2>
        {MENU_FLUTUANTE}
        <div class="form-container">
        <h3>➕ Criar Novo Usuário</h3>
        <form method="post" action="/criar_usuario">
        <input type="text" name="username" placeholder="Usuário" required>
        <input type="password" name="password" placeholder="Senha" required>
        <input type="text" name="telefone" placeholder="Telefone (opcional)">
        <select name="nivel" required>
        <option value="">Selecione o nível</option>
        <option value="administrador">Administrador</option>
        <option value="vendedor">Vendedor</option>
        <option value="consulta">Consulta</option>
        </select>
        <button type="submit">Criar Usuário</button>
        </form>
        </div>
        <h3>📋 Usuários Cadastrados</h3>
        <table>
        <thead>
        <tr>
        <th>ID</th>
        <th>Usuário</th>
        <th>Nível</th>
        <th>Telefone</th>
        <th>Ações</th>
        </tr>
        </thead>
        <tbody>
        {''.join(f"""
        <tr>
        <td>{u['id']}</td>
        <td>{u['nome de usuario']}</td>
        <td>{u['NÍVEL']}</td>
        <td>{u.get('telefone', '—')}</td>
        <td><a href="/excluir_usuario/{u['id']}" onclick="return confirm('Tem certeza?')">🗑️ Excluir</a></td>
        </tr>
        """ for u in usuarios)}
        </tbody>
        </table>
        </div>
        </body>
        </html>
        '''
    except Exception as e:
        flash("Erro ao carregar usuários.")
        return redirect(url_for('clientes'))

@app.route('/criar_usuario', methods=['POST'])
def criar_usuario_view():
    if 'usuario' not in session or session['nivel'] != 'administrador':
        flash("Acesso negado!")
        return redirect(url_for('gerenciar_usuarios'))
    username = request.form.get('username')
    password = request.form.get('password')
    nivel = request.form.get('nivel')
    telefone = request.form.get('telefone')
    if not username or not password or not nivel:
        flash("Todos os campos são obrigatórios!")
        return redirect(url_for('gerenciar_usuarios'))
    if nivel not in ['administrador', 'vendedor', 'consulta']:
        flash("Nível inválido!")
        return redirect(url_for('gerenciar_usuarios'))
    try:
        if criar_usuario(username, password, nivel, telefone):
            flash("Usuário criado com sucesso!")
        else:
            flash("Erro ao criar usuário.")
    except Exception as e:
        print("Erro ao criar usuário:", e)
        flash("Erro interno no servidor.")
    return redirect(url_for('gerenciar_usuarios'))

@app.route('/excluir_usuario/<int:id>')
def excluir_usuario_view(id):
    if 'usuario' not in session or session['nivel'] != 'administrador':
        flash("Acesso negado!")
        return redirect(url_for('clientes'))
    try:
        if excluir_usuario(id):
            flash("Usuário excluído!")
        else:
            flash("Erro ao excluir usuário.")
    except Exception as e:
        print("Erro ao excluir usuário:", e)
        flash("Erro interno no servidor.")
    return redirect(url_for('gerenciar_usuarios'))

@app.route('/cadastrar_cliente', methods=['GET', 'POST'])
def cadastrar_cliente():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        nome = request.form.get('nome')
        cnpj = request.form.get('cnpj')
        responsavel = request.form.get('responsavel')
        telefone = request.form.get('telefone')
        whatsapp = request.form.get('whatsapp')
        email = request.form.get('email')
        endereco = request.form.get('endereco')
        bairro = request.form.get('bairro')
        cidade = request.form.get('cidade')
        estado = request.form.get('estado')
        cep = request.form.get('cep')
        numero = request.form.get('numero')
        tem_entrega = request.form.get('tem_entrega') == 'on'
        entrega_endereco = request.form.get('entrega_endereco') if tem_entrega else None
        entrega_numero = request.form.get('entrega_numero') if tem_entrega else None
        entrega_bairro = request.form.get('entrega_bairro') if tem_entrega else None
        entrega_cidade = request.form.get('entrega_cidade') if tem_entrega else None
        entrega_estado = request.form.get('entrega_estado') if tem_entrega else None
        entrega_cep = request.form.get('entrega_cep') if tem_entrega else None
        
        if not nome or not cnpj:
            flash("Nome e CNPJ são obrigatórios!")
            return redirect(url_for('cadastrar_cliente'))
        
        if criar_empresa(nome, cnpj, responsavel, telefone, whatsapp, email, endereco, bairro, cidade, estado, cep, numero,
                        entrega_endereco, entrega_numero, entrega_bairro, entrega_cidade, entrega_estado, entrega_cep):
            flash("✅ Empresa cadastrada com sucesso!")
            return redirect(url_for('listar_empresas'))
        else:
            flash("❌ Erro ao cadastrar empresa.")
            return redirect(url_for('cadastrar_cliente'))
    
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Cadastrar Empresa - Sua Gráfica</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: #333; min-height: 100vh; padding: 0; margin: 0; }}
    .container {{ max-width: 800px; margin: 30px auto; background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0,0,0,0.1); overflow: hidden; }}
    .header {{ background: #2c3e50; color: white; text-align: center; padding: 30px; }}
    h1 {{ font-size: 28px; margin: 0; font-weight: 600; }}
    .user-info {{ background: #34495e; color: white; padding: 15px 20px; font-size: 15px; display: flex; justify-content: space-between; align-items: center; }}
    .form-container {{ padding: 30px; }}
    .grid-2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 15px; }}
    .grid-3 {{ display: grid; grid-template-columns: 1fr 1fr 2fr; gap: 15px; }}
    .form-container label {{ display: block; margin: 10px 0 5px 0; font-weight: 600; color: #2c3e50; }}
    .form-container input, .form-container select {{ width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 6px; font-size: 14px; }}
    .btn {{ padding: 12px 20px; background: #27ae60; color: white; border: none; border-radius: 8px; font-size: 16px; font-weight: 600; cursor: pointer; }}
    .back-link {{ display: inline-block; margin: 20px 30px; color: #3498db; text-decoration: none; font-weight: 500; }}
    .footer {{ text-align: center; padding: 20px; background: #ecf0f1; color: #7f8c8d; font-size: 13px; border-top: 1px solid #bdc3c7; }}
    </style>
    </head>
    <body>
    <div class="container">
    <div class="header"><h1>➕ Cadastrar Nova Empresa</h1></div>
    <div class="user-info"><span>👤 {session['usuario']} ({session['nivel'].upper()})</span><a href="/logout">🚪 Sair</a></div>
    <a href="/empresas" class="back-link">← Voltar à lista</a>
    <form method="post" class="form-container">
    <div class="grid-2"><div><label>Nome da Empresa *</label><input type="text" name="nome" required></div><div><label>CNPJ *</label><input type="text" name="cnpj" class="mascara-cnpj" placeholder="00.000.000/0000-00" required></div></div>
    <div class="grid-2"><div><label>Nome do Responsável</label><input type="text" name="responsavel"></div><div><label>WhatsApp</label><input type="text" name="whatsapp"></div></div>
    <div class="grid-2"><div><label>Telefone</label><input type="text" name="telefone"></div><div><label>E-mail</label><input type="email" name="email"></div></div>
    <div class="grid-3"><div><label>CEP</label><input type="text" name="cep" id="cep" onblur="buscarEnderecoPorCEP()" placeholder="00000-000" style="width: 150px;"></div><div><label>Bairro</label><input type="text" name="bairro" id="bairro" style="width: 150px;"></div><div><label>Endereço</label><input type="text" name="endereco" id="endereco" style="width: 100%; max-width: 350px;"></div></div>
    <div class="grid-3"><div><label>Número</label><input type="text" name="numero" placeholder="Ex: 123"></div><div><label>Cidade</label><input type="text" name="cidade" id="cidade"></div><div><label>Estado</label><select name="estado" id="estado"><option value="">Selecione</option><option value="AC">AC</option><option value="AL">AL</option><option value="AP">AP</option><option value="AM">AM</option><option value="BA">BA</option><option value="CE">CE</option><option value="DF">DF</option><option value="ES">ES</option><option value="GO">GO</option><option value="MA">MA</option><option value="MT">MT</option><option value="MS">MS</option><option value="MG">MG</option><option value="PA">PA</option><option value="PB">PB</option><option value="PR">PR</option><option value="PE">PE</option><option value="PI">PI</option><option value="RJ">RJ</option><option value="RN">RN</option><option value="RS">RS</option><option value="RO">RO</option><option value="RR">RR</option><option value="SC">SC</option><option value="SP">SP</option><option value="SE">SE</option><option value="TO">TO</option></select></div></div>
    <div style="margin: 20px 0; padding: 15px; border: 1px dashed #3498db; border-radius: 8px; line-height: 1;"><input type="checkbox" name="tem_entrega" id="tem_entrega" onchange="toggleEntrega()" style="margin-right: 8px; vertical-align: middle;"><label for="tem_entrega" style="font-weight: 600; font-size: 16px; vertical-align: middle;">Endereço de entrega diferente do endereço da empresa?</label></div>
    <div id="campos-entrega" style="display: none;">
    <div class="grid-3"><div><label>CEP de Entrega</label><input type="text" name="entrega_cep" id="entrega_cep" placeholder="00000-000" style="width: 150px;"></div><div><label>Bairro de Entrega</label><input type="text" name="entrega_bairro" id="entrega_bairro" style="width: 150px;"></div><div><label>Endereço de Entrega</label><input type="text" name="entrega_endereco" id="entrega_endereco" style="width: 100%; max-width: 350px;"></div></div>
    <div class="grid-3"><div><label>Número de Entrega</label><input type="text" name="entrega_numero" placeholder="Ex: 123"></div><div><label>Cidade de Entrega</label><input type="text" name="entrega_cidade" id="entrega_cidade"></div><div><label>Estado de Entrega</label><select name="entrega_estado" id="entrega_estado"><option value="">Selecione</option><option value="AC">AC</option><option value="AL">AL</option><option value="AP">AP</option><option value="AM">AM</option><option value="BA">BA</option><option value="CE">CE</option><option value="DF">DF</option><option value="ES">ES</option><option value="GO">GO</option><option value="MA">MA</option><option value="MT">MT</option><option value="MS">MS</option><option value="MG">MG</option><option value="PA">PA</option><option value="PB">PB</option><option value="PR">PR</option><option value="PE">PE</option><option value="PI">PI</option><option value="RJ">RJ</option><option value="RN">RN</option><option value="RS">RS</option><option value="RO">RO</option><option value="RR">RR</option><option value="SC">SC</option><option value="SP">SP</option><option value="SE">SE</option><option value="TO">TO</option></select></div></div>
    </div>
    <button type="submit" class="btn">💾 Salvar Empresa</button>
    </form>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </div>
    <script>
    function buscarEnderecoPorCEP() {{ const cep = document.getElementById('cep').value.replace(/\\D/g, ''); if (cep.length !== 8) {{ alert('CEP inválido!'); return; }} fetch(`https://viacep.com.br/ws/${{cep}}/json/`).then(response => response.json()).then(data => {{ if (data.erro) {{ alert('CEP não encontrado!'); return; }} document.getElementById('endereco').value = data.logradouro; document.getElementById('bairro').value = data.bairro; document.getElementById('cidade').value = data.localidade; document.getElementById('estado').value = data.uf; }}).catch(error => {{ console.error('Erro ao buscar CEP:', error); alert('Erro ao buscar CEP. Tente novamente.'); }}); }}
    function toggleEntrega() {{ const campos = document.getElementById('campos-entrega'); campos.style.display = document.getElementById('tem_entrega').checked ? 'block' : 'none'; }}
    document.getElementById('entrega_cep').onblur = function() {{ const cep = this.value.replace(/\\D/g, ''); if (cep.length !== 8) return; fetch(`https://viacep.com.br/ws/${{cep}}/json/`).then(r => r.json()).then(data => {{ if (!data.erro) {{ document.getElementById('entrega_endereco').value = data.logradouro; document.getElementById('entrega_bairro').value = data.bairro; document.getElementById('entrega_cidade').value = data.localidade; document.getElementById('entrega_estado').value = data.uf; }} }}); }};
    </script>
    {SCRIPT_MASCARA_CNPJ}
    </body>
    </html>
    '''

@app.route('/empresas')
def listar_empresas():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    
    busca = request.args.get('q', '').strip()
    try:
        if busca:
            url = f"{SUPABASE_URL}/rest/v1/empresas?or=(nome_empresa.ilike.*{busca}*,cnpj.ilike.*{busca}*)"
        else:
            url = f"{SUPABASE_URL}/rest/v1/empresas?select=*"
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            empresas = response.json()
        else:
            flash("Erro ao carregar empresas.")
            empresas = []
    except Exception as e:
        flash("Erro de conexão.")
        empresas = []
    
    # Botão Voltar só aparece quando há busca
    btn_voltar = f'<a href="/empresas" style="margin-right: 10px; padding: 10px 15px; background: #95a5a6; color: white; text-decoration: none; border-radius: 8px; font-size: 14px;">← Voltar / Limpar</a>' if busca else ''
    
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Empresas Cadastradas</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f5f7fa; color: #333; min-height: 100vh; padding: 0; margin: 0; }}
    .container {{ max-width: 1100px; margin: 30px auto; background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0, 0, 0, 0.1); overflow: hidden; }}
    .header {{ background: #2c3e50; color: white; text-align: center; padding: 30px; }}
    h1 {{ font-size: 28px; margin: 0; font-weight: 600; }}
    .user-info {{ background: #34495e; color: white; padding: 15px 20px; font-size: 15px; display: flex; justify-content: space-between; align-items: center; }}
    .search-box {{ padding: 20px 30px; text-align: center; }}
    .search-box input {{ width: 70%; padding: 12px; border: 1px solid #ddd; border-radius: 8px; font-size: 16px; }}
    table {{ width: 100%; border-collapse: collapse; }}
    th, td {{ padding: 16px 20px; text-align: left; }}
    th {{ background: #ecf0f1; color: #2c3e50; font-weight: 600; text-transform: uppercase; font-size: 14px; }}
    tr:nth-child(even) {{ background: #f9f9f9; }}
    tr:hover {{ background: #f1f7fb; }}
    .back-link {{ display: inline-block; margin: 20px 30px; color: #3498db; text-decoration: none; font-weight: 500; }}
    .footer {{ text-align: center; padding: 20px; background: #ecf0f1; color: #7f8c8d; font-size: 13px; border-top: 1px solid #bdc3c7; }}
    </style>
    </head>
    <body>
    <div class="container">
    <div class="header"><h1>📋 Empresas Cadastradas</h1></div>
    <div class="user-info"><span>👤 {session['usuario']} ({session['nivel'].upper()})</span><a href="/logout">🚪 Sair</a></div>
    {MENU_FLUTUANTE}
    <div class="search-box">
    <form method="get" style="display: inline;">
        {btn_voltar}
        <input type="text" name="q" class="mascara-cnpj" placeholder="Pesquisar por nome ou CNPJ..." value="{busca}">
        <button type="submit" style="padding: 10px 20px; background: #3498db; color: white; border: none; border-radius: 8px; cursor: pointer;">🔍 Pesquisar</button>
    </form>
    </div>
    <table>
    <thead><tr><th>ID</th><th>Empresa</th><th>CNPJ</th><th>Responsável</th><th>WhatsApp</th><th>Ações</th></tr></thead>
    <tbody>
    {''.join(f'''<tr><td>{e["id"]}</td><td>{e["nome_empresa"]}</td><td>{e["cnpj"]}</td><td>{e["responsavel"]}</td><td>{e["whatsapp"]}</td><td><a href="/empresa/{e["id"]}" style="color: #3498db; text-decoration: none;">👁️ Ver Detalhes</a></td></tr>''' for e in empresas)}
    </tbody>
    </table>
    <div style="text-align: center; padding: 20px;"><a href="/cadastrar_cliente" class="btn" style="padding: 12px 20px; background: #27ae60; color: white; border-radius: 8px; text-decoration: none; font-weight: 600;">➕ Cadastrar Nova Empresa</a></div>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </div>
    
    <!-- SCRIPT DA MÁSCARA DE CNPJ -->
    <script>
    document.addEventListener('DOMContentLoaded', function() {{
        const inputs = document.querySelectorAll('.mascara-cnpj');
        inputs.forEach(input => {{
            input.addEventListener('input', function(e) {{
                let value = e.target.value.replace(/\D/g, "");
                value = value.replace(/^(\\d{{2}})(\\d)/, "$1.$2");
                value = value.replace(/^(\\d{{2}})\\.(\\d{{3}})(\\d)/, "$1.$2.$3");
                value = value.replace(/\\.(\\d{{3}})(\\d)/, ".$1/$2");
                value = value.replace(/(\\d{{4}})(\\d)/, "$1-$2");
                e.target.value = value;
            }});
        }});
    }});
    </script>
    </body>
    </html>
    '''

@app.route('/empresa/<int:id>')
def detalhes_empresa(id):
    if 'usuario' not in session:
        return redirect(url_for('login'))
    try:
        url = f"{SUPABASE_URL}/rest/v1/empresas?id=eq.{id}"
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            empresa = response.json()[0] if response.json() else None
            if not empresa:
                flash("Empresa não encontrada.")
                return redirect(url_for('listar_empresas'))
        else:
            flash("Erro ao carregar empresa.")
            return redirect(url_for('listar_empresas'))
    except Exception as e:
        flash("Erro de conexão.")
        return redirect(url_for('listar_empresas'))
    try:
        url_amostras = f"{SUPABASE_URL}/rest/v1/envios?select=*,empresas(nome_empresa)&empresa_id=eq.{id}&tipo_envio=eq.Amostra&order=data_envio.desc"
        response_amostras = requests.get(url_amostras, headers=headers)
        amostras = response_amostras.json() if response_amostras.status_code == 200 else []
    except Exception as e:
        amostras = []
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{empresa['nome_empresa']} - Detalhes</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f5f7fa; color: #333; min-height: 100vh; padding: 0; margin: 0; }}
    .container {{ max-width: 800px; margin: 30px auto; background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0, 0, 0, 0.1); overflow: hidden; }}
    .header {{ background: #2c3e50; color: white; text-align: center; padding: 30px; }}
    h1 {{ font-size: 28px; margin: 0; font-weight: 600; }}
    .user-info {{ background: #34495e; color: white; padding: 15px 20px; font-size: 15px; display: flex; justify-content: space-between; align-items: center; }}
    .details {{ padding: 30px; }}
    .details p {{ margin: 10px 0; font-size: 16px; }}
    .details strong {{ color: #2c3e50; }}
    .btn {{ padding: 12px 20px; background: #27ae60; color: white; border: none; border-radius: 8px; font-size: 16px; font-weight: 600; text-decoration: none; margin: 10px 30px; }}
    .btn-blue {{ background: #3498db; }}
    .back-link {{ display: inline-block; margin: 20px 30px; color: #3498db; text-decoration: none; font-weight: 500; }}
    .footer {{ text-align: center; padding: 20px; background: #ecf0f1; color: #7f8c8d; font-size: 13px; border-top: 1px solid #bdc3c7; }}
    .section {{ padding: 20px 30px; }}
    .section h3 {{ margin: 20px 0 15px 0; color: #2c3e50; font-size: 20px; }}
    table {{ width: 100%; border-collapse: collapse; margin: 10px 0; }}
    th, td {{ padding: 12px 15px; text-align: left; border: 1px solid #ddd; }}
    th {{ background: #ecf0f1; font-weight: 600; }}
    </style>
    </head>
    <body>\n
    <div class="container">
    <div class="header"><h1>🏢 {empresa['nome_empresa']}</h1></div>
    <div class="user-info"><span>👤 {session['usuario']} ({session['nivel'].upper()})</span><a href="/logout">🚪 Sair</a></div>
    <a href="/empresas" class="back-link">← Voltar à Lista</a>
    <div class="details">
    <p><strong>CNPJ:</strong> {empresa['cnpj']}</p>
    <p><strong>Responsável:</strong> {empresa['responsavel']}</p>
    <p><strong>Telefone:</strong> {empresa['telefone']}</p>
    <p><strong>WhatsApp:</strong> {empresa['whatsapp']}</p>
    <p><strong>E-mail:</strong> {empresa['email']}</p>
    <p><strong>Endereço:</strong> {empresa['endereco']}, {empresa['numero']} - {empresa['bairro']}, {empresa['cidade']} - {empresa['estado']} ({empresa['cep']})</p>
    {f'<p><strong>Endereço de Entrega:</strong> {empresa["entrega_endereco"]}, {empresa["entrega_numero"]} - {empresa["entrega_bairro"]}, {empresa["entrega_cidade"]} - {empresa["entrega_estado"]} ({empresa["entrega_cep"]})</p>' if empresa.get("entrega_endereco") else ''}
    </div>
    <div style="display: flex; gap: 15px; margin: 20px 0;">
    <a href="/servicos_empresa/{id}" class="btn">📋 Serviços desta empresa</a>
    <a href="/editar_empresa/{empresa['id']}" class="btn" style="background: #f39c12;">✏️ Editar Empresa</a>
    <a href="/gerar_etiqueta/{id}" class="btn" style="background: #8e44ad;">📬 Etiqueta de Postagem</a>
    </div>
    <div class="section">
    <h3>📦 Amostras Enviadas</h3>
    <table>
    <thead><tr><th>Data Envio</th><th>O que foi enviado</th><th>Código Rastreio</th><th>Status</th><th>Ações</th></tr></thead>
    <tbody>
    {''.join(f"""<tr><td>{format_data(a.get('data_envio'))}</td><td>{a['descricao']}</td><td>{a['codigo_rastreio']}</td><td><span style="color: {'#27ae60' if a['status'] == 'Entregue' else '#e67e22'}; font-weight: bold;">{a['status']}</span></td><td><a href="https://www.linkcorreios.com.br/{a['codigo_rastreio']}" target="_blank" class="btn btn-blue" style="padding: 5px 10px; font-size: 12px;">🔍 Rastrear</a></td></tr>""" for a in amostras)}
    </tbody>
    </table>
    {f'<p style="text-align: center; color: #95a5a6;">Nenhuma amostra enviada ainda.</p>' if not amostras else ''}
    </div>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </div>
    </body>
    </html>
    '''

@app.route('/editar_empresa/<int:id>', methods=['GET', 'POST'])
def editar_empresa(id):
    if 'usuario' not in session:
        return redirect(url_for('login'))
    try:
        url = f"{SUPABASE_URL}/rest/v1/empresas?id=eq.{id}"
        response = requests.get(url, headers=headers)
        if response.status_code != 200 or not response.json():
            flash("Empresa não encontrada.")
            return redirect(url_for('listar_empresas'))
        empresa = response.json()[0]
    except Exception as e:
        flash("Erro ao carregar empresa.")
        return redirect(url_for('listar_empresas'))
    if request.method == 'POST':
        nome = request.form.get('nome')
        cnpj = request.form.get('cnpj')
        responsavel = request.form.get('responsavel')
        telefone = request.form.get('telefone')
        whatsapp = request.form.get('whatsapp')
        email = request.form.get('email')
        endereco = request.form.get('endereco')
        bairro = request.form.get('bairro')
        cidade = request.form.get('cidade')
        estado = request.form.get('estado')
        cep = request.form.get('cep')
        numero = request.form.get('numero')
        tem_entrega = request.form.get('tem_entrega') == 'on'
        entrega_endereco = request.form.get('entrega_endereco') if tem_entrega else None
        entrega_numero = request.form.get('entrega_numero') if tem_entrega else None
        entrega_bairro = request.form.get('entrega_bairro') if tem_entrega else None
        entrega_cidade = request.form.get('entrega_cidade') if tem_entrega else None
        entrega_estado = request.form.get('entrega_estado') if tem_entrega else None
        entrega_cep = request.form.get('entrega_cep') if tem_entrega else None
        if not nome or not cnpj:
            flash("Nome e CNPJ são obrigatórios!")
            return redirect(url_for('editar_empresa', id=id))
        try:
            url = f"{SUPABASE_URL}/rest/v1/empresas?id=eq.{id}"
            dados = {
                "nome_empresa": nome, "cnpj": cnpj, "responsavel": responsavel, "telefone": telefone,
                "whatsapp": whatsapp, "email": email, "endereco": endereco, "bairro": bairro,
                "cidade": cidade, "estado": estado, "cep": cep, "numero": numero,
                "entrega_endereco": entrega_endereco, "entrega_numero": entrega_numero,
                "entrega_bairro": entrega_bairro, "entrega_cidade": entrega_cidade,
                "entrega_estado": entrega_estado, "entrega_cep": entrega_cep
            }
            response = requests.patch(url, json=dados, headers=headers)
            if response.status_code == 204:
                flash("✅ Empresa atualizada com sucesso!")
                return redirect(url_for('detalhes_empresa', id=id))
            else:
                flash("❌ Erro ao atualizar empresa.")
        except Exception as e:
            flash("❌ Erro de conexão.")
        return redirect(url_for('editar_empresa', id=id))
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Editar Empresa - Sua Gráfica</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f5f7fa; color: #333; min-height: 100vh; padding: 0; margin: 0; }}
    .container {{ max-width: 800px; margin: 30px auto; background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0, 0, 0, 0.1); overflow: hidden; }}
    .header {{ background: #2c3e50; color: white; text-align: center; padding: 30px; }}
    h1 {{ font-size: 28px; margin: 0; font-weight: 600; }}
    .user-info {{ background: #34495e; color: white; padding: 15px 20px; font-size: 15px; display: flex; justify-content: space-between; align-items: center; }}
    .form-container {{ padding: 30px; }}
    .grid-2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 15px; }}
    .grid-3 {{ display: grid; grid-template-columns: 1fr 1fr 2fr; gap: 15px; }}
    .form-container label {{ display: block; margin: 10px 0 5px 0; font-weight: 600; color: #2c3e50; }}
    .form-container input, .form-container select {{ width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 6px; font-size: 14px; }}
    .btn {{ padding: 12px 20px; background: #27ae60; color: white; border: none; border-radius: 8px; font-size: 16px; font-weight: 600; cursor: pointer; }}
    .back-link {{ display: inline-block; margin: 20px 30px; color: #3498db; text-decoration: none; font-weight: 500; }}
    .footer {{ text-align: center; padding: 20px; background: #ecf0f1; color: #7f8c8d; font-size: 13px; border-top: 1px solid #bdc3c7; }}
    </style>
    </head>
    <body>\n
    <div class="container">
    <div class="header"><h1>✏️ Editar {empresa['nome_empresa']}</h1></div>
    <div class="user-info"><span>👤 {session['usuario']} ({session['nivel'].upper()})</span><a href="/logout">🚪 Sair</a></div>
    <a href="/empresa/{id}" class="back-link">← Voltar aos Detalhes</a>
    <form method="post" class="form-container">
    <div class="grid-2"><div><label>Nome da Empresa *</label><input type="text" name="nome" value="{empresa['nome_empresa']}" required></div><div><label>CNPJ *</label><input type="text" name="cnpj" value="{empresa['cnpj']}" required></div></div>
    <div class="grid-2"><div><label>Nome do Responsável</label><input type="text" name="responsavel" value="{empresa['responsavel']}"></div><div><label>WhatsApp</label><input type="text" name="whatsapp" value="{empresa['whatsapp']}"></div></div>
    <div class="grid-2"><div><label>Telefone</label><input type="text" name="telefone" value="{empresa['telefone']}"></div><div><label>E-mail</label><input type="email" name="email" value="{empresa['email']}"></div></div>
    <div class="grid-3"><div><label>CEP</label><input type="text" name="cep" id="cep" onblur="buscarEnderecoPorCEP()" placeholder="00000-000" value="{empresa['cep']}" style="width: 150px;"></div><div><label>Bairro</label><input type="text" name="bairro" id="bairro" value="{empresa['bairro']}" style="width: 150px;"></div><div><label>Endereço</label><input type="text" name="endereco" id="endereco" value="{empresa['endereco']}" style="width: 100%; max-width: 350px;"></div></div>
    <div class="grid-3"><div><label>Número</label><input type="text" name="numero" value="{empresa['numero']}"></div><div><label>Cidade</label><input type="text" name="cidade" id="cidade" value="{empresa['cidade']}"></div><div><label>Estado</label><select name="estado" id="estado"><option value="">Selecione</option><option value="AC" {"selected" if empresa['estado'] == "AC" else ""}>AC</option><option value="AL" {"selected" if empresa['estado'] == "AL" else ""}>AL</option><option value="AP" {"selected" if empresa['estado'] == "AP" else ""}>AP</option><option value="AM" {"selected" if empresa['estado'] == "AM" else ""}>AM</option><option value="BA" {"selected" if empresa['estado'] == "BA" else ""}>BA</option><option value="CE" {"selected" if empresa['estado'] == "CE" else ""}>CE</option><option value="DF" {"selected" if empresa['estado'] == "DF" else ""}>DF</option><option value="ES" {"selected" if empresa['estado'] == "ES" else ""}>ES</option><option value="GO" {"selected" if empresa['estado'] == "GO" else ""}>GO</option><option value="MA" {"selected" if empresa['estado'] == "MA" else ""}>MA</option><option value="MT" {"selected" if empresa['estado'] == "MT" else ""}>MT</option><option value="MS" {"selected" if empresa['estado'] == "MS" else ""}>MS</option><option value="MG" {"selected" if empresa['estado'] == "MG" else ""}>MG</option><option value="PA" {"selected" if empresa['estado'] == "PA" else ""}>PA</option><option value="PB" {"selected" if empresa['estado'] == "PB" else ""}>PB</option><option value="PR" {"selected" if empresa['estado'] == "PR" else ""}>PR</option><option value="PE" {"selected" if empresa['estado'] == "PE" else ""}>PE</option><option value="PI" {"selected" if empresa['estado'] == "PI" else ""}>PI</option><option value="RJ" {"selected" if empresa['estado'] == "RJ" else ""}>RJ</option><option value="RN" {"selected" if empresa['estado'] == "RN" else ""}>RN</option><option value="RS" {"selected" if empresa['estado'] == "RS" else ""}>RS</option><option value="RO" {"selected" if empresa['estado'] == "RO" else ""}>RO</option><option value="RR" {"selected" if empresa['estado'] == "RR" else ""}>RR</option><option value="SC" {"selected" if empresa['estado'] == "SC" else ""}>SC</option><option value="SP" {"selected" if empresa['estado'] == "SP" else ""}>SP</option><option value="SE" {"selected" if empresa['estado'] == "SE" else ""}>SE</option><option value="TO" {"selected" if empresa['estado'] == "TO" else ""}>TO</option></select></div></div>
    <div style="margin: 20px 0; padding: 15px; border: 1px dashed #3498db; border-radius: 8px; line-height: 1;"><input type="checkbox" name="tem_entrega" id="tem_entrega" onchange="toggleEntrega()" {"checked" if empresa.get("entrega_endereco") else ""} style="margin-right: 8px; vertical-align: middle;"><label for="tem_entrega" style="font-weight: 600; font-size: 16px; vertical-align: middle;">Endereço de entrega diferente do endereço da empresa?</label></div>
    <div id="campos-entrega" style="display: {'block' if empresa.get('entrega_endereco') else 'none'};">
    <div class="grid-3"><div><label>CEP de Entrega</label><input type="text" name="entrega_cep" id="entrega_cep" placeholder="00000-000" value="{empresa.get('entrega_cep', '')}" style="width: 150px;"></div><div><label>Bairro de Entrega</label><input type="text" name="entrega_bairro" id="entrega_bairro" value="{empresa.get('entrega_bairro', '')}" style="width: 150px;"></div><div><label>Endereço de Entrega</label><input type="text" name="entrega_endereco" id="entrega_endereco" value="{empresa.get('entrega_endereco', '')}" style="width: 100%; max-width: 350px;"></div></div>
    <div class="grid-3"><div><label>Número de Entrega</label><input type="text" name="entrega_numero" placeholder="Ex: 123" value="{empresa.get('entrega_numero', '')}"></div><div><label>Cidade de Entrega</label><input type="text" name="entrega_cidade" id="entrega_cidade" value="{empresa.get('entrega_cidade', '')}"></div><div><label>Estado de Entrega</label><select name="entrega_estado" id="entrega_estado"><option value="">Selecione</option><option value="AC" {"selected" if empresa.get('entrega_estado') == "AC" else ""}>AC</option><option value="AL" {"selected" if empresa.get('entrega_estado') == "AL" else ""}>AL</option><option value="AP" {"selected" if empresa.get('entrega_estado') == "AP" else ""}>AP</option><option value="AM" {"selected" if empresa.get('entrega_estado') == "AM" else ""}>AM</option><option value="BA" {"selected" if empresa.get('entrega_estado') == "BA" else ""}>BA</option><option value="CE" {"selected" if empresa.get('entrega_estado') == "CE" else ""}>CE</option><option value="DF" {"selected" if empresa.get('entrega_estado') == "DF" else ""}>DF</option><option value="ES" {"selected" if empresa.get('entrega_estado') == "ES" else ""}>ES</option><option value="GO" {"selected" if empresa.get('entrega_estado') == "GO" else ""}>GO</option><option value="MA" {"selected" if empresa.get('entrega_estado') == "MA" else ""}>MA</option><option value="MT" {"selected" if empresa.get('entrega_estado') == "MT" else ""}>MT</option><option value="MS" {"selected" if empresa.get('entrega_estado') == "MS" else ""}>MS</option><option value="MG" {"selected" if empresa.get('entrega_estado') == "MG" else ""}>MG</option><option value="PA" {"selected" if empresa.get('entrega_estado') == "PA" else ""}>PA</option><option value="PB" {"selected" if empresa.get('entrega_estado') == "PB" else ""}>PB</option><option value="PR" {"selected" if empresa.get('entrega_estado') == "PR" else ""}>PR</option><option value="PE" {"selected" if empresa.get('entrega_estado') == "PE" else ""}>PE</option><option value="PI" {"selected" if empresa.get('entrega_estado') == "PI" else ""}>PI</option><option value="RJ" {"selected" if empresa.get('entrega_estado') == "RJ" else ""}>RJ</option><option value="RN" {"selected" if empresa.get('entrega_estado') == "RN" else ""}>RN</option><option value="RS" {"selected" if empresa.get('entrega_estado') == "RS" else ""}>RS</option><option value="RO" {"selected" if empresa.get('entrega_estado') == "RO" else ""}>RO</option><option value="RR" {"selected" if empresa.get('entrega_estado') == "RR" else ""}>RR</option><option value="SC" {"selected" if empresa.get('entrega_estado') == "SC" else ""}>SC</option><option value="SP" {"selected" if empresa.get('entrega_estado') == "SP" else ""}>SP</option><option value="SE" {"selected" if empresa.get('entrega_estado') == "SE" else ""}>SE</option><option value="TO" {"selected" if empresa.get('entrega_estado') == "TO" else ""}>TO</option></select></div></div>
    </div>
    <button type="submit" class="btn">💾 Salvar Alterações</button>
    </form>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </div>
    <script>
    function buscarEnderecoPorCEP() {{ const cep = document.getElementById('cep').value.replace(/\\D/g, ''); if (cep.length !== 8) {{ alert('CEP inválido!'); return; }} fetch(`https://viacep.com.br/ws/${{cep}}/json/`).then(response => response.json()).then(data => {{ if (data.erro) {{ alert('CEP não encontrado!'); return; }} document.getElementById('endereco').value = data.logradouro; document.getElementById('bairro').value = data.bairro; document.getElementById('cidade').value = data.localidade; document.getElementById('estado').value = data.uf; }}).catch(error => {{ console.error('Erro ao buscar CEP:', error); alert('Erro ao buscar CEP. Tente novamente.'); }}); }}
    function toggleEntrega() {{ const campos = document.getElementById('campos-entrega'); campos.style.display = document.getElementById('tem_entrega').checked ? 'block' : 'none'; }}
    document.getElementById('entrega_cep').onblur = function() {{ const cep = this.value.replace(/\\D/g, ''); if (cep.length !== 8) return; fetch(`https://viacep.com.br/ws/${{cep}}/json/`).then(r => r.json()).then(data => {{ if (!data.erro) {{ document.getElementById('entrega_endereco').value = data.logradouro; document.getElementById('entrega_bairro').value = data.bairro; document.getElementById('entrega_cidade').value = data.localidade; document.getElementById('entrega_estado').value = data.uf; }} }}); }};
    </script>
    </body>
    </html>
    '''

@app.route('/servicos_empresa/<int:id>')
def servicos_empresa(id):
    if 'usuario' not in session:
        return redirect(url_for('login'))
    try:
        url_empresa = f"{SUPABASE_URL}/rest/v1/empresas?id=eq.{id}"
        response = requests.get(url_empresa, headers=headers)
        if response.status_code != 200 or not response.json():
            flash("Empresa não encontrada.")
            return redirect(url_for('listar_empresas'))
        empresa = response.json()[0]
        url_servicos = f"{SUPABASE_URL}/rest/v1/servicos?select=*,materiais_usados(*,materiais(denominacao))&empresa_id=eq.{id}&order=codigo_servico.desc"
        response_serv = requests.get(url_servicos, headers=headers)
        servicos = response_serv.json() if response_serv.status_code == 200 else []
    except Exception as e:
        print(f"Erro ao carregar serviços da empresa: {e}")
        flash("Erro ao carregar serviços.")
        return redirect(url_for('listar_empresas'))
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Serviços - {empresa['nome_empresa']}</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f5f7fa; color: #333; min-height: 100vh; padding: 0; margin: 0; }}
    .container {{ max-width: 1200px; margin: 30px auto; background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0,0,0,0.1); overflow: hidden; }}
    .header {{ background: #2c3e50; color: white; text-align: center; padding: 30px; }}
    h1 {{ font-size: 28px; margin: 0; font-weight: 600; }}
    .user-info {{ background: #34495e; color: white; padding: 15px 20px; font-size: 15px; display: flex; justify-content: space-between; align-items: center; }}
    table {{ width: 100%; border-collapse: collapse; }}
    th, td {{ padding: 12px 15px; text-align: left; border-bottom: 1px solid #eee; }}
    th {{ background: #ecf0f1; color: #2c3e50; font-weight: 600; }}
    .back-link {{ display: inline-block; margin: 20px 30px; color: #3498db; text-decoration: none; font-weight: 500; }}
    .btn {{ padding: 8px 12px; background: #3498db; color: white; text-decoration: none; border-radius: 6px; font-size: 14px; }}
    .footer {{ text-align: center; padding: 20px; background: #ecf0f1; color: #7f8c8d; font-size: 13px; border-top: 1px solid #bdc3c7; }}
    </style>
    </head>
    <body>\n
    <div class="container">
    <div class="header"><h1>📋 Serviços - {empresa['nome_empresa']}</h1></div>
    <div class="user-info"><span>👤 {session['usuario']} ({session['nivel'].upper()})</span><a href="/logout">🚪 Sair</a></div>
    <a href="/empresa/{id}" class="back-link">← Voltar à empresa</a>
    <div style="padding: 30px;">
    <h2 style="margin-bottom: 20px;">Total de serviços: {len(servicos)}</h2>
    <table>
    <thead><tr><th>Código</th><th>Título</th><th>Status</th><th>Valor</th><th>Data</th><th>Ações</th></tr></thead>
    <tbody>
    {''.join(f"""
    <tr>
    <td>{s['codigo_servico']}</td>
    <td>{s['titulo']}</td>
    <td>{s.get('status', '—')}</td>
    <td>R$ {float(s.get('valor_cobrado', 0) or 0):.2f}</td>
    <td>{s.get('data_abertura', '—')[:10] if s.get('data_abertura') else '—'}</td>
    <td><a href="/os/{s['id']}" class="btn">📄 Ver OS</a></td>
    </tr>
    """ for s in servicos)}
    </tbody>
    </table>
    {f'<p style="text-align: center; color: #95a5a6; margin-top: 30px;">Nenhum serviço encontrado para esta empresa.</p>' if not servicos else ''}
    </div>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </div>
    </body>
    </html>
    '''
@app.route('/servicos')
def listar_servicos():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    busca = request.args.get('q', '').strip()
    try:
        url = f"{SUPABASE_URL}/rest/v1/servicos?select=*,empresas(nome_empresa),materiais_usados(*,materiais(denominacao))&order=codigo_servico.desc&tipo=neq.Orçamento"
        if busca:
            url += f"&titulo=ilike.*{busca}*"
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            servicos = response.json()
        else:
            flash("Erro ao carregar serviços.")
            servicos = []
    except Exception as e:
        flash("Erro de conexão.")
        servicos = []

    def calcular_custo(servico_id):
        try:
            url_mat = f"{SUPABASE_URL}/rest/v1/materiais_usados?select=valor_total&servico_id=eq.{servico_id}"
            resp = requests.get(url_mat, headers=headers)
            if resp.status_code == 200:
                itens = resp.json()
                return sum(float(i['valor_total']) for i in itens)
            return 0.0
        except:
            return 0.0

    def calcular_prazo_restante(previsao, status):
        if not previsao:
            return {"dias": 0, "cor": "#95a5a6", "texto": "Sem prazo"}
        try:
            data_prev = datetime.strptime(previsao.split("T")[0], "%Y-%m-%d")
            hoje = datetime.now().date()
            dias = (data_prev.date() - hoje).days
            if status in ['Concluído', 'Entregue']:
                return {"dias": 0, "cor": "#27ae60", "texto": "Finalizado"}
            if dias < 0:
                return {"dias": abs(dias), "cor": "#e74c3c", "texto": f"Atrasado há {abs(dias)} dia(s)"}
            elif dias <= 3:
                return {"dias": dias, "cor": "#e67e22", "texto": f"Faltam {dias} dias"}
            elif dias <= 5:
                return {"dias": dias, "cor": "#f39c12", "texto": f"Faltam {dias} dias"}
            else:
                return {"dias": dias, "cor": "#27ae60", "texto": f"Faltam {dias} dias"}
        except:
            return {"dias": 0, "cor": "#95a5a6", "texto": "Erro"}

    html_todos = ""
    html_andamento = ""
    html_concluidos = ""

    for s in servicos:
        empresa_nome = s['empresas']['nome_empresa'] if s.get('empresas') else "Sem cliente"
        custo_materiais = calcular_custo(s['id'])
        valor_cobrado = float(s.get('valor_cobrado', 0) or 0)
        lucro = valor_cobrado - custo_materiais
        status_class = {
            'Pendente': 'status-pendente',
            'Em Produção': 'status-producao',
            'Concluído': 'status-concluido',
            'Entregue': 'status-entregue'
        }.get(s.get('status', ''), 'status-pendente')
        prazo = calcular_prazo_restante(s.get('previsao_entrega'), s.get('status'))

        # ✅ CORREÇÃO: Botões com espaçamento adequado e nowrap para não quebrar
        botoes_html = f'''
        <div style="display: flex; gap: 8px; align-items: center; white-space: nowrap;">
            <a href="/os/{s['id']}" class="btn btn-blue" style="padding: 6px 12px; font-size: 12px;">📄 OS</a>
            <a href="/editar_servico/{s['id']}" class="btn btn-edit" style="padding: 6px 12px; font-size: 12px;">✏️ Editar</a>
            <a href="/excluir_servico/{s['id']}" class="btn btn-delete" style="padding: 6px 12px; font-size: 12px;" onclick="return confirm('Tem certeza?')">🗑️ Excluir</a>
        </div>
        '''

        linha = f'''
        <tr>
        <td>{s['codigo_servico']}</td>
        <td>{s['titulo']}</td>
        <td>{empresa_nome}</td>
        <td>{s.get('quantidade', '-')}</td>
        <td>{s.get('dimensao', '-')}</td>
        <td>R$ {custo_materiais:.2f}</td>
        <td>R$ {valor_cobrado:.2f}</td>
        <td>R$ {lucro:.2f}</td>
        <td><span class="{status_class}">{s.get('status', 'Pendente')}</span></td>
        <td><span style="color: {prazo['cor']}; font-weight: bold;">{prazo['texto']}</span></td>
        <td>{botoes_html}</td>
        </tr>
        '''
        html_todos += linha
        if s.get('status') in ['Pendente', 'Em Produção']:
            html_andamento += linha
        elif s.get('status') in ['Concluído', 'Entregue']:
            html_concluidos += linha

    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Serviços / Ordens de Serviço</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f5f7fa; color: #333; min-height: 100vh; padding: 0; margin: 0; }}
    .container {{ max-width: 1400px; margin: 30px auto; background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0,0,0,0.1); overflow: hidden; }}
    .header {{ background: #2c3e50; color: white; text-align: center; padding: 30px; }}
    h1 {{ font-size: 28px; margin: 0; font-weight: 600; }}
    .user-info {{ background: #34495e; color: white; padding: 15px 20px; font-size: 15px; display: flex; justify-content: space-between; align-items: center; }}
    .btn {{ padding: 10px 15px; background: #27ae60; color: white; border: none; border-radius: 8px; font-size: 14px; font-weight: 600; text-decoration: none; margin: 5px; display: inline-block; }}
    .btn-blue {{ background: #3498db; }}
    .btn-edit {{ background: #f39c12; }}
    .btn-delete {{ background: #e74c3c; }}
    .tabs {{ display: flex; margin: 0 30px; border-bottom: 1px solid #ddd; }}
    .tab {{ padding: 15px 20px; background: #ecf0f1; color: #7f8c8d; cursor: pointer; font-weight: 600; }}
    .tab.active {{ background: #3498db; color: white; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    th, td {{ padding: 12px 10px; text-align: left; border-bottom: 1px solid #eee; }}
    th {{ background: #f8f9fa; color: #2c3e50; font-weight: 600; white-space: nowrap; }}
    tr:nth-child(even) {{ background: #fafbfc; }}
    .status-pendente {{ color: #e67e22; font-weight: bold; }}
    .status-producao {{ color: #3498db; font-weight: bold; }}
    .status-concluido {{ color: #27ae60; font-weight: bold; }}
    .status-entregue {{ color: #2c3e50; font-weight: bold; }}
    .back-link {{ display: inline-block; margin: 20px 30px; color: #3498db; text-decoration: none; font-weight: 500; }}
    .footer {{ text-align: center; padding: 20px; background: #ecf0f1; color: #7f8c8d; font-size: 13px; border-top: 1px solid #bdc3c7; }}
    .tab-content {{ display: none; }}
    .tab-content.active {{ display: table-row-group; }}
    .search-box {{ text-align: center; padding: 20px; }}
    .search-box input {{ padding: 10px; width: 300px; border: 1px solid #ddd; border-radius: 8px; }}
    </style>
    </head>
    <body>\n
    <div class="container">
    <div class="header"><h1>📋 Todos os Serviços</h1></div>
    <div class="user-info"><span>👤 {session['usuario']} ({session['nivel'].upper()})</span><a href="/logout">🚪 Sair</a></div>
    {MENU_FLUTUANTE}
    <a href="/adicionar_servico" class="btn">➕ Adicionar Novo Serviço</a>
    <div class="search-box">
    <form method="get" style="display: inline;"><input type="text" name="q" placeholder="Pesquisar por título..." value="{busca}"><button type="submit" style="padding: 10px 20px; background: #3498db; color: white; border: none; border-radius: 8px; cursor: pointer;">🔍 Pesquisar</button></form>
    </div>
    <div class="tabs">
    <div class="tab active" onclick="mostrarTab('todos')">Todos os Serviços</div>
    <div class="tab" onclick="mostrarTab('andamento')">Em Andamento</div>
    <div class="tab" onclick="mostrarTab('concluidos')">Concluídos / Entregues</div>
    </div>
    <div style="overflow-x: auto;">
    <table>
    <thead><tr><th>Código</th><th>Título</th><th>Cliente</th><th>Qtd</th><th>Dimensão</th><th>Custo Mat.</th><th>Valor Cobrado</th><th>Lucro</th><th>Status</th><th>Prazo Restante</th><th>Ações</th></tr></thead>
    <tbody id="tab-todos" class="tab-content active">{html_todos}</tbody>
    <tbody id="tab-andamento" class="tab-content">{html_andamento}</tbody>
    <tbody id="tab-concluidos" class="tab-content">{html_concluidos}</tbody>
    </table>
    </div>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </div>
    <script>
    function mostrarTab(nome) {{ document.querySelectorAll('.tab-content').forEach(t => t.classList.remove('active')); document.querySelectorAll('.tab').forEach(t => t.classList.remove('active')); document.getElementById('tab-' + nome).classList.add('active'); document.querySelector(`[onclick="mostrarTab('${{nome}}')"]`).classList.add('active'); }}
    </script>
    </body>
    </html>
    '''

@app.route('/adicionar_servico', methods=['GET', 'POST'])
def adicionar_servico():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    if request.method == 'POST':
        titulo = request.form.get('titulo')
        empresa_id = request.form.get('empresa_id')
        tipo = request.form.get('tipo')
        quantidade = request.form.get('quantidade')
        dimensao = request.form.get('dimensao')
        numero_cores = request.form.get('numero_cores')
        aplicacao = request.form.get('aplicacao')
        status = request.form.get('status') or 'Pendente'
        data_abertura = request.form.get('data_abertura')
        previsao_entrega = request.form.get('previsao_entrega')
        valor_cobrado = request.form.get('valor_cobrado') or 0.0
        observacoes = request.form.get('observacoes')
        if not titulo or not empresa_id:
            flash("Título e Cliente são obrigatórios!")
            return redirect(url_for('adicionar_servico'))
        try:
            valor_cobrado = float(valor_cobrado)
        except:
            valor_cobrado = 0.0
        try:
            url_seq = f"{SUPABASE_URL}/rest/v1/servicos?select=codigo_servico&order=codigo_servico.desc&limit=1"
            response = requests.get(url_seq, headers=headers)
            if response.status_code == 200 and response.json():
                ultimo = response.json()[0]['codigo_servico']
                numero = int(ultimo.split('-')[1]) + 1
            else:
                numero = 1
            codigo_servico = f"OS-{numero:03d}"
        except:
            codigo_servico = "OS-001"
        try:
            url = f"{SUPABASE_URL}/rest/v1/servicos"
            dados = {
                "codigo_servico": codigo_servico,
                "titulo": titulo,
                "empresa_id": int(empresa_id),
                "tipo": tipo,
                "quantidade": quantidade,
                "dimensao": dimensao,
                "numero_cores": numero_cores,
                "aplicacao": aplicacao,
                "status": status,
                "data_abertura": data_abertura,
                "previsao_entrega": previsao_entrega,
                "valor_cobrado": valor_cobrado,
                "observacoes": observacoes
            }
            response = requests.post(url, json=dados, headers=headers)
            if response.status_code == 201:
                servico_id = response.json()['id']
                flash("✅ Serviço criado com sucesso!")
                materiais_ids = request.form.getlist('material_id[]')
                quantidades = request.form.getlist('quantidade_usada[]')
                valores_unitarios = request.form.getlist('valor_unitario[]')
                for i in range(len(materiais_ids)):
                    try:
                        material_id = int(materiais_ids[i])
                        qtd = float(quantidades[i])
                        vlr = float(valores_unitarios[i])
                        total = qtd * vlr
                        dados_mat = {
                            "servico_id": servico_id,
                            "material_id": material_id,
                            "quantidade_usada": qtd,
                            "valor_unitario": vlr,
                            "valor_total": total
                        }
                        requests.post(f"{SUPABASE_URL}/rest/v1/materiais_usados", json=dados_mat, headers=headers)
                    except:
                        continue
                return redirect(url_for('listar_servicos'))
            else:
                flash("❌ Erro ao salvar serviço.")
        except Exception as e:
            flash("❌ Erro de conexão.")
        empresas = buscar_empresas()
        materiais = buscar_materiais()
        return f'''
        <!DOCTYPE html>
        <html lang="pt-BR">
        <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Adicionar Serviço</title>
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f5f7fa; color: #333; min-height: 100vh; padding: 0; margin: 0; }}
        .container {{ max-width: 1000px; margin: 30px auto; background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0,0,0,0.1); overflow: hidden; }}
        .header {{ background: #2c3e50; color: white; text-align: center; padding: 30px; }}
        h1 {{ font-size: 28px; margin: 0; font-weight: 600; }}
        .user-info {{ background: #34495e; color: white; padding: 15px 20px; font-size: 15px; display: flex; justify-content: space-between; align-items: center; }}
        .form-container {{ padding: 30px; }}
        .grid-2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 15px; }}
        .grid-3 {{ display: grid; grid-template-columns: 1fr 1fr 2fr; gap: 15px; }}
        .form-container label {{ display: block; margin: 10px 0 5px 0; font-weight: 600; color: #2c3e50; }}
        .form-container input, .form-container select, .form-container textarea {{ width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 6px; font-size: 14px; }}
        .btn {{ padding: 12px 20px; background: #27ae60; color: white; border: none; border-radius: 8px; font-size: 16px; font-weight: 600; cursor: pointer; }}
        .back-link {{ display: inline-block; margin: 20px 30px; color: #3498db; text-decoration: none; font-weight: 500; }}
        .footer {{ text-align: center; padding: 20px; background: #ecf0f1; color: #7f8c8d; font-size: 13px; border-top: 1px solid #bdc3c7; }}
        </style>
        </head>
        <body>\n
        <div class="container">
        <div class="header"><h1>➕ Adicionar Novo Serviço</h1></div>
        <div class="user-info"><span>👤 {session['usuario']} ({session['nivel'].upper()})</span><a href="/logout">🚪 Sair</a></div>
        <a href="/servicos" class="back-link">← Voltar à lista</a>
        <form method="post" class="form-container">
        <label>Código do Serviço (OS)</label><input type="text" readonly value="(será gerado automaticamente)" style="background: #eee;">
        <label>Título do Serviço *</label><input type="text" name="titulo" required>
        <label>Cliente *</label><select name="empresa_id" required><option value="">Selecione uma empresa</option>{''.join(f'<option value="{e["id"]}">{e["nome_empresa"]}</option>' for e in empresas)}</select>
        <div class="grid-2">
            <div><label>Tipo</label><select name="tipo"><option value="">Selecione</option><option value="Orçamento">Orçamento</option><option value="Produção">Produção</option><option value="Equipamento">Equipamento</option></select></div>
            <div><label>Status</label><select name="status"><option value="Pendente">Pendente</option><option value="Em Produção">Em Produção</option><option value="Concluído">Concluído</option><option value="Entregue">Entregue</option></select></div>
        </div>
        <div class="grid-2">
            <div><label>Quantidade / Lote</label><input type="number" name="quantidade" step="1"></div>
            <div><label>Nº de Cores</label><input type="number" name="numero_cores" step="1"></div>
        </div>
        <div class="grid-2">
            <div><label>Dimensão (ex: 60x90 cm)</label><input type="text" name="dimensao"></div>
            <div><label>Valor Cobrado (R$)</label><input type="number" name="valor_cobrado" step="0.01"></div>
        </div>
        <div class="grid-2">
            <div><label>Data de Abertura</label><input type="date" name="data_abertura"></div>
            <div><label>Previsão de Entrega</label><input type="date" name="previsao_entrega"></div>
        </div>
        <label>Aplicação / Uso / Ambiente</label><textarea name="aplicacao" rows="3"></textarea>
        <label>Observações</label><textarea name="observacoes" rows="3"></textarea>
        <h3>Materiais Usados</h3>
        <div id="materiais-lista">
            <div class="grid-3">
                <div><label>Material</label><select name="material_id[]" required><option value="">Selecione</option>{''.join(f'<option value="{m["id"]}">{m["denominacao"]} ({m["unidade_medida"]})</option>' for m in materiais)}</select></div>
                <div><label>Qtd Usada</label><input type="number" name="quantidade_usada[]" step="0.01" required></div>
                <div><label>Valor Unitário (R$)</label><input type="number" name="valor_unitario[]" step="0.01" required></div>
            </div>
        </div>
        <button type="button" onclick="adicionarMaterial()" style="margin: 10px 0;">+ Adicionar outro material</button>
        <button type="submit" class="btn">💾 Salvar Serviço</button>
        </form>
        <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
        </div>
        <script>
        function adicionarMaterial() {{
            const container = document.getElementById('materiais-lista');
            const div = document.createElement('div');
            div.className = 'grid-3';
            div.innerHTML = `
            <div><label>Material</label><select name="material_id[]" required><option value="">Selecione</option>{''.join(f'<option value="{m["id"]}">{m["denominacao"]} ({m["unidade_medida"]})</option>' for m in materiais)}</select></div>
            <div><label>Qtd Usada</label><input type="number" name="quantidade_usada[]" step="0.01" required></div>
            <div><label>Valor Unitário (R$)</label><input type="number" name="valor_unitario[]" step="0.01" required></div>
            `;
            container.appendChild(div);
        }}
        </script>
        </body>
        </html>
        '''

@app.route('/editar_servico/<int:id>', methods=['GET', 'POST'])
def editar_servico(id):
    if 'usuario' not in session:
        return redirect(url_for('login'))
    try:
        url = f"{SUPABASE_URL}/rest/v1/servicos?id=eq.{id}"
        response = requests.get(url, headers=headers)
        if response.status_code != 200 or not response.json():
            flash("Serviço não encontrado.")
            return redirect(url_for('listar_servicos'))
        servico = response.json()[0]
    except Exception as e:
        flash("Erro ao carregar serviço.")
        return redirect(url_for('listar_servicos'))
    try:
        url_mats = f"{SUPABASE_URL}/rest/v1/materiais_usados?select=*,materiais(denominacao,unidade_medida)&servico_id=eq.{id}"
        response_mats = requests.get(url_mats, headers=headers)
        materiais_usados = response_mats.json() if response_mats.status_code == 200 else []
    except:
        materiais_usados = []
    if request.method == 'POST':
        titulo = request.form.get('titulo')
        empresa_id = request.form.get('empresa_id')
        tipo = request.form.get('tipo')
        quantidade = request.form.get('quantidade')
        dimensao = request.form.get('dimensao')
        numero_cores = request.form.get('numero_cores')
        aplicacao = request.form.get('aplicacao')
        status = request.form.get('status')
        data_abertura = request.form.get('data_abertura')
        previsao_entrega = request.form.get('previsao_entrega')
        valor_cobrado = request.form.get('valor_cobrado') or 0.0
        observacoes = request.form.get('observacoes')
        if not titulo or not empresa_id:
            flash("Título e Cliente são obrigatórios!")
            return redirect(request.url)
        try:
            valor_cobrado = float(valor_cobrado)
        except:
            valor_cobrado = 0.0
        try:
            dados = {
                "titulo": titulo,
                "empresa_id": int(empresa_id),
                "tipo": tipo,
                "quantidade": quantidade,
                "dimensao": dimensao,
                "numero_cores": numero_cores,
                "aplicacao": aplicacao,
                "status": status,
                "data_abertura": data_abertura,
                "previsao_entrega": previsao_entrega,
                "valor_cobrado": valor_cobrado,
                "observacoes": observacoes
            }
            response = requests.patch(url, json=dados, headers=headers)
            if response.status_code == 204:
                flash("✅ Serviço atualizado com sucesso!")
                ids_materiais = request.form.getlist('material_usado_id[]')
                for i in range(len(ids_materiais)):
                    try:
                        mat_id = ids_materiais[i]
                        qtd = float(request.form[f'quantidade_usada_{mat_id}'])
                        vlr = float(request.form[f'valor_unitario_{mat_id}'])
                        total = qtd * vlr
                        dados_mat = {
                            "quantidade_usada": qtd,
                            "valor_unitario": vlr,
                            "valor_total": total
                        }
                        requests.patch(f"{SUPABASE_URL}/rest/v1/materiais_usados?id=eq.{mat_id}", json=dados_mat, headers=headers)
                    except:
                        continue
                return redirect(url_for('listar_servicos'))
            else:
                flash("❌ Erro ao atualizar serviço.")
        except Exception as e:
            flash("❌ Erro de conexão.")
        return redirect(request.url)
    empresas = buscar_empresas()
    materiais = buscar_materiais()
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Editar Serviço</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f5f7fa; color: #333; min-height: 100vh; padding: 0; margin: 0; }}
    .container {{ max-width: 1000px; margin: 30px auto; background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0,0,0,0.1); overflow: hidden; }}
    .header {{ background: #2c3e50; color: white; text-align: center; padding: 30px; }}
    h1 {{ font-size: 28px; margin: 0; font-weight: 600; }}
    .user-info {{ background: #34495e; color: white; padding: 15px 20px; font-size: 15px; display: flex; justify-content: space-between; align-items: center; }}
    .form-container {{ padding: 30px; }}
    .grid-2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 15px; }}
    .form-container label {{ display: block; margin: 10px 0 5px 0; font-weight: 600; color: #2c3e50; }}
    .form-container input, .form-container select, .form-container textarea {{ width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 6px; font-size: 14px; }}
    .btn {{ padding: 12px 20px; background: #f39c12; color: white; border: none; border-radius: 8px; font-size: 16px; font-weight: 600; cursor: pointer; }}
    .back-link {{ display: inline-block; margin: 20px 30px; color: #3498db; text-decoration: none; font-weight: 500; }}
    .footer {{ text-align: center; padding: 20px; background: #ecf0f1; color: #7f8c8d; font-size: 13px; border-top: 1px solid #bdc3c7; }}
    </style>
    </head>
    <body>\n
    <div class="container">
    <div class="header"><h1>✏️ Editar Serviço: {servico['codigo_servico']}</h1></div>
    <div class="user-info"><span>👤 {session['usuario']} ({session['nivel'].upper()})</span><a href="/logout">🚪 Sair</a></div>
    <a href="/servicos" class="back-link">← Voltar à lista</a>
    <form method="post" class="form-container">
    <label>Título do Serviço *</label><input type="text" name="titulo" value="{servico['titulo']}" required>
    <label>Cliente *</label><select name="empresa_id" required><option value="">Selecione uma empresa</option>{''.join(f'<option value="{e["id"]}" {"selected" if e["id"] == servico["empresa_id"] else ""}>{e["nome_empresa"]}</option>' for e in empresas)}</select>
    <div class="grid-2">
        <div><label>Tipo</label><select name="tipo"><option value="">Selecione</option><option value="Orçamento" {"selected" if servico["tipo"] == "Orçamento" else ""}>Orçamento</option><option value="Produção" {"selected" if servico["tipo"] == "Produção" else ""}>Produção</option><option value="Equipamento" {"selected" if servico["tipo"] == "Equipamento" else ""}>Equipamento</option></select></div>
        <div><label>Status</label><select name="status"><option value="Pendente" {"selected" if servico["status"] == "Pendente" else ""}>Pendente</option><option value="Em Produção" {"selected" if servico["status"] == "Em Produção" else ""}>Em Produção</option><option value="Concluído" {"selected" if servico["status"] == "Concluído" else ""}>Concluído</option><option value="Entregue" {"selected" if servico["status"] == "Entregue" else ""}>Entregue</option></select></div>
    </div>
    <div class="grid-2">
        <div><label>Quantidade / Lote</label><input type="number" name="quantidade" value="{servico.get('quantidade', '')}" step="1"></div>
        <div><label>Nº de Cores</label><input type="number" name="numero_cores" value="{servico.get('numero_cores', '')}" step="1"></div>
    </div>
    <div class="grid-2">
        <div><label>Dimensão (ex: 60x90 cm)</label><input type="text" name="dimensao" value="{servico.get('dimensao', '')}"></div>
        <div><label>Valor Cobrado (R$)</label><input type="number" name="valor_cobrado" value="{servico.get('valor_cobrado', 0)}" step="0.01"></div>
    </div>
    <div class="grid-2">
        <div><label>Data de Abertura</label><input type="date" name="data_abertura" value="{servico.get('data_abertura', '')[:10] if servico.get('data_abertura') else ''}"></div>
        <div><label>Previsão de Entrega</label><input type="date" name="previsao_entrega" value="{servico.get('previsao_entrega', '')[:10] if servico.get('previsao_entrega') else ''}"></div>
    </div>
    <label>Aplicação / Uso / Ambiente</label><textarea name="aplicacao" rows="3">{servico.get('aplicacao', '')}</textarea>
    <label>Observações</label><textarea name="observacoes" rows="3">{servico.get('observacoes', '')}</textarea>
    <h3>Materiais Usados</h3>
    {''.join(f'''
    <input type="hidden" name="material_usado_id[]" value="{m['id']}">
    <div class="grid-3">
        <div><label>Material</label><input type="text" value="{m['materiais']['denominacao']} ({m['materiais']['unidade_medida']})" readonly></div>
        <div><label>Qtd Usada</label><input type="number" name="quantidade_usada_{m['id']}" value="{m['quantidade_usada']}" step="0.01" required></div>
        <div><label>Valor Unitário (R$)</label><input type="number" name="valor_unitario_{m['id']}" value="{m['valor_unitario']}" step="0.01" required></div>
    </div>
    ''' for m in materiais_usados)}
    <button type="submit" class="btn">💾 Salvar Alterações</button>
    </form>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </div>
    </body>
    </html>
    '''

@app.route('/gerar_etiqueta/<int:id>')
def gerar_etiqueta(id):
    if 'usuario' not in session:
        return redirect(url_for('login'))
    
    try:
        url = f"{SUPABASE_URL}/rest/v1/empresas?id=eq.{id}"
        resp = requests.get(url, headers=headers)
        empresa = resp.json()[0] if resp.json() else None
        
        if not empresa:
            flash("Empresa não encontrada!")
            return redirect(url_for('listar_empresas'))
        
        # Dados do remetente (Lira Print)
        remetente_padrao = {
            "nome": "LIRAPRINT",
            "endereco": "R. Dr. Roberto Fernandes, 81",
            "bairro": "Jardim Palmira",
            "cidade": "Guarulhos",
            "estado": "SP",
            "cep": "07076-070",
            "telefone": "(11) 2468-1234"
        }
        
        # Dados do destinatário
        if empresa.get('entrega_endereco'):
            dest_endereco = empresa['entrega_endereco']
            dest_numero = empresa.get('entrega_numero', '')
            dest_bairro = empresa.get('entrega_bairro', '')
            dest_cidade = empresa.get('entrega_cidade', '')
            dest_estado = empresa.get('entrega_estado', '')
            dest_cep = empresa.get('entrega_cep', '')
        else:
            dest_endereco = empresa['endereco']
            dest_numero = empresa.get('numero', '')
            dest_bairro = empresa['bairro']
            dest_cidade = empresa['cidade']
            dest_estado = empresa['estado']
            dest_cep = empresa['cep']
        
        destinatario_padrao = {
            "nome": empresa['nome_empresa'],
            "responsavel": empresa.get('responsavel', ''),
            "endereco": f"{dest_endereco}, {dest_numero}",
            "bairro": dest_bairro,
            "cidade": dest_cidade,
            "estado": dest_estado,
            "cep": dest_cep,
            "telefone": empresa.get('whatsapp', empresa.get('telefone', ''))
        }
        
    except Exception as e:
        print(f"Erro: {e}")
        flash("Erro ao carregar dados.")
        return redirect(url_for('listar_empresas'))
    
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Etiqueta - {empresa['nome_empresa']}</title>
    <style>
    * {{ margin: 0; padding: 0; box-sizing: border-box; }}
    body {{ 
        font-family: Arial, sans-serif; 
        background: #f5f7fa; 
        padding: 20px;
        font-size: 14px;
    }}
    .container {{ 
        max-width: 900px; 
        margin: 0 auto; 
        background: white; 
        border-radius: 10px; 
        box-shadow: 0 2px 10px rgba(0,0,0,0.1); 
        padding: 30px;
    }}
    .header {{ 
        background: #2c3e50; 
        color: white; 
        padding: 15px; 
        text-align: center; 
        border-radius: 10px 10px 0 0; 
        margin: -30px -30px 20px -30px;
        font-size: 18px;
    }}
    .section {{ 
        margin-bottom: 20px; 
        border: 2px solid #ecf0f1; 
        border-radius: 8px; 
        padding: 15px; 
    }}
    .section-title {{ 
        font-size: 14px; 
        font-weight: bold; 
        color: #2c3e50; 
        margin-bottom: 10px;
        text-transform: uppercase;
        border-bottom: 2px solid #3498db;
        padding-bottom: 5px;
    }}
    .grid-2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 10px; }}
    .form-group label {{ 
        display: block; margin-bottom: 4px; font-weight: bold; color: #2c3e50; font-size: 13px; 
    }}
    .form-group input {{ 
        width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 4px; font-size: 14px; color: #333;
    }}
    .btn {{ 
        padding: 10px 20px; background: #27ae60; color: white; border: none; border-radius: 5px; cursor: pointer; font-size: 15px; font-weight: bold; margin: 5px; 
    }}
    .btn-purple {{ background: #8e44ad; }}
    .btn-blue {{ background: #3498db; }}
    .back-link {{ color: #3498db; text-decoration: none; display: inline-block; margin-bottom: 10px; font-size: 14px; }}
    
    /* ETIQUETA COMPACTADA PARA CABER EM UMA FOLHA */
    .etiqueta {{ 
        border: 2px solid #2c3e50; 
        padding: 10px 15px; /* Padding reduzido */
        margin: 10px auto; 
        background: #fff;
        width: 100%; /* Flexível */
        max-width: 100mm; /* Largura padrão etiqueta */
        height: 100%; /* Ocupa a altura necessária */
        page-break-inside: avoid;
        overflow: hidden; /* Garante que não estoure */
    }}
    
    /* Remetente e Destinatário mais próximos */
    .etiqueta-bloco {{ 
        margin-bottom: 10px; /* Espaço reduzido entre blocos */
        padding: 0;
    }}
    
    .etiqueta-label {{ 
        font-size: 10px; 
        text-transform: uppercase; 
        color: #7f8c8d; 
        font-weight: bold; 
        margin-bottom: 2px; 
    }}
    
    /* Fontes ajustadas para serem legíveis mas compactas */
    .etiqueta-valor {{ 
        font-size: 13px; /* Tamanho médio */
        color: #2c3e50; 
        font-weight: bold; 
        margin-bottom: 2px; 
        line-height: 1.2; /* Linhas mais juntas */
    }}
    
    /* A/C compacto */
    .etiqueta-ac {{
        font-size: 12px;
        color: #e74c3c;
        font-weight: bold;
        margin-top: 0;
        margin-bottom: 4px;
        font-style: italic;
    }}
    
    /* CEP ajustado */
    .etiqueta-cep {{ 
        font-size: 16px; /* CEP menor para caber */
        font-weight: 900; 
        color: #2c3e50; 
        background: #ecf0f1; 
        padding: 4px 10px; 
        border-radius: 4px; 
        display: inline-block; 
        margin-top: 5px; 
    }}
    
    /* IMPRESSÃO */
    @media print {{
        @page {{ 
            size: auto;
            margin: 0mm;
        }}
        body {{ 
            background: white; 
            padding: 0; 
            margin: 0; 
        }}
        .no-print {{ display: none !important; }}
        .container {{ 
            box-shadow: none; border-radius: 0; padding: 0; margin: 0; width: 100%; 
        }}
        .etiqueta {{ 
            border: 2px solid #000;
            margin: 0;
            width: 100mm; /* Força largura exata */
            height: auto; /* Altura automática baseada no conteúdo */
        }}
    }}
    </style>
    </head>
    <body>
    <div class="container">
        <div class="header no-print">📬 Etiqueta de Postagem</div>
        
        <div class="no-print">
            <a href="/empresa/{id}" class="back-link">← Voltar</a>
            
            <form id="formEtiqueta">
                <!-- REMETENTE -->
                <div class="section">
                    <div class="section-title">📤 Remetente (Lira Print)</div>
                    <div class="form-group"><label>Nome</label><input type="text" id="rem_nome" value="{remetente_padrao['nome']}"></div>
                    <div class="form-group"><label>Endereço</label><input type="text" id="rem_endereco" value="{remetente_padrao['endereco']}"></div>
                    <div class="grid-2">
                        <div class="form-group"><label>Bairro</label><input type="text" id="rem_bairro" value="{remetente_padrao['bairro']}"></div>
                        <div class="form-group"><label>Cidade/UF</label><input type="text" id="rem_cidade_uf" value="{remetente_padrao['cidade']} - {remetente_padrao['estado']}"></div>
                    </div>
                    <div class="grid-2">
                        <div class="form-group"><label>CEP</label><input type="text" id="rem_cep" value="{remetente_padrao['cep']}"></div>
                        <div class="form-group"><label>Telefone</label><input type="text" id="rem_telefone" value="{remetente_padrao['telefone']}"></div>
                    </div>
                </div>
                
                <!-- DESTINATÁRIO -->
                <div class="section">
                    <div class="section-title">📥 Destinatário (Cliente)</div>
                    <div class="form-group"><label>Nome/Razão Social</label><input type="text" id="dest_nome" value="{destinatario_padrao['nome']}"></div>
                    <div class="form-group"><label>A/C (Aos Cuidados de)</label><input type="text" id="dest_responsavel" value="{destinatario_padrao['responsavel']}"></div>
                    <div class="form-group"><label>Endereço</label><input type="text" id="dest_endereco" value="{destinatario_padrao['endereco']}"></div>
                    <div class="grid-2">
                        <div class="form-group"><label>Bairro</label><input type="text" id="dest_bairro" value="{destinatario_padrao['bairro']}"></div>
                        <div class="form-group"><label>Cidade/UF</label><input type="text" id="dest_cidade_uf" value="{destinatario_padrao['cidade']} - {destinatario_padrao['estado']}"></div>
                    </div>
                    <div class="grid-2">
                        <div class="form-group"><label>CEP</label><input type="text" id="dest_cep" value="{destinatario_padrao['cep']}"></div>
                        <div class="form-group"><label>Telefone</label><input type="text" id="dest_telefone" value="{destinatario_padrao['telefone']}"></div>
                    </div>
                </div>
                
                <div style="text-align: center; margin: 20px 0;">
                    <button type="button" class="btn btn-purple" onclick="imprimirEtiqueta()">🖨️ Imprimir Etiqueta</button>
                    <button type="button" class="btn btn-blue" onclick="window.location.href='/empresa/{id}'">← Voltar</button>
                </div>
            </form>
        </div>
        
        <!-- ÁREA DE IMPRESSÃO DA ETIQUETA -->
        <div id="areaEtiqueta" class="etiqueta">
            
            <!-- REMETENTE -->
            <div class="etiqueta-bloco">
                <div class="etiqueta-label">REMETENTE</div>
                <div class="etiqueta-valor" id="view_rem_nome">{remetente_padrao['nome']}</div>
                <div class="etiqueta-valor" id="view_rem_endereco">{remetente_padrao['endereco']}</div>
                <div class="etiqueta-valor" id="view_rem_bairro">{remetente_padrao['bairro']}</div>
                <div class="etiqueta-valor" id="view_rem_cidade_uf">{remetente_padrao['cidade']} - {remetente_padrao['estado']}</div>
                <div class="etiqueta-cep" id="view_rem_cep">{remetente_padrao['cep']}</div>
            </div>
            
            <!-- DESTINATÁRIO -->
            <div class="etiqueta-bloco">
                <div class="etiqueta-label">DESTINATÁRIO</div>
                <div class="etiqueta-valor" id="view_dest_nome">{destinatario_padrao['nome']}</div>
                <div class="etiqueta-ac" id="view_dest_ac">A/C: {destinatario_padrao['responsavel']}</div>
                <div class="etiqueta-valor" id="view_dest_endereco">{destinatario_padrao['endereco']}</div>
                <div class="etiqueta-valor" id="view_dest_bairro">{destinatario_padrao['bairro']}</div>
                <div class="etiqueta-valor" id="view_dest_cidade_uf">{destinatario_padrao['cidade']} - {destinatario_padrao['estado']}</div>
                <div class="etiqueta-cep" id="view_dest_cep">{destinatario_padrao['cep']}</div>
            </div>
            
        </div>
    </div>
    
    <script>
    function atualizarEtiqueta() {{
        document.getElementById('view_rem_nome').textContent = document.getElementById('rem_nome').value;
        document.getElementById('view_rem_endereco').textContent = document.getElementById('rem_endereco').value;
        document.getElementById('view_rem_bairro').textContent = document.getElementById('rem_bairro').value;
        document.getElementById('view_rem_cidade_uf').textContent = document.getElementById('rem_cidade_uf').value;
        document.getElementById('view_rem_cep').textContent = document.getElementById('rem_cep').value;
        
        document.getElementById('view_dest_nome').textContent = document.getElementById('dest_nome').value;
        const resp = document.getElementById('dest_responsavel').value;
        const ac = document.getElementById('view_dest_ac');
        ac.textContent = resp ? 'A/C: ' + resp : '';
        ac.style.display = resp ? 'block' : 'none';
        
        document.getElementById('view_dest_endereco').textContent = document.getElementById('dest_endereco').value;
        document.getElementById('view_dest_bairro').textContent = document.getElementById('dest_bairro').value;
        document.getElementById('view_dest_cidade_uf').textContent = document.getElementById('dest_cidade_uf').value;
        document.getElementById('view_dest_cep').textContent = document.getElementById('dest_cep').value;
    }}
    
    document.querySelectorAll('#formEtiqueta input').forEach(input => input.addEventListener('input', atualizarEtiqueta));
    
    function imprimirEtiqueta() {{ window.print(); }}
    
    document.getElementById('dest_cep').addEventListener('input', function(e) {{
        let v = e.target.value.replace(/\D/g, '');
        if(v.length > 5) v = v.slice(0,5) + '-' + v.slice(5,8);
        e.target.value = v;
    }});
    </script>
    </body>
    </html>
    '''

@app.route('/complementar_orcamento/<int:id>')
def complementar_orcamento(id):
    if 'usuario' not in session:
        return redirect(url_for('login'))
    try:
        # Converte orçamento em OS
        requests.patch(f"{SUPABASE_URL}/rest/v1/servicos?id=eq.{id}", 
            json={"tipo": "Produção", "status": "Pendente"}, headers=headers)
        flash("✅ Orçamento convertido em OS!")
    except:
        flash("❌ Erro ao converter")
    return redirect(url_for('listar_orcamentos'))  # Volta para ORÇAMENTOS, não serviços


@app.route('/excluir_servico/<int:id>')
def excluir_servico(id):
    if session.get('nivel') != 'administrador':
        flash("Acesso negado!")
        return redirect(url_for('listar_orcamentos'))  # Ou listar_servicos, conforme origem
    
    try:
        # Verifica o tipo antes de excluir
        url = f"{SUPABASE_URL}/rest/v1/servicos?id=eq.{id}&select=tipo"
        resp = requests.get(url, headers=headers).json()
        eh_orcamento = resp and resp[0].get('tipo') == 'Orçamento'
        
        if eh_orcamento:
            requests.delete(f"{SUPABASE_URL}/rest/v1/itens_orcamento?orcamento_id=eq.{id}", headers=headers)
        
        requests.delete(f"{SUPABASE_URL}/rest/v1/servicos?id=eq.{id}", headers=headers)
        flash("🗑️ Excluído com sucesso!")
        
        # Redireciona conforme o tipo
        return redirect(url_for('listar_orcamentos') if eh_orcamento else url_for('listar_servicos'))
    except:
        flash("❌ Erro ao excluir")
        return redirect(url_for('listar_orcamentos'))


@app.route('/os/<int:id>')
def imprimir_os(id):
    if 'usuario' not in session:
        return redirect(url_for('login'))
    try:
        url_serv = f"{SUPABASE_URL}/rest/v1/servicos?id=eq.{id}&select=*,empresas(nome_empresa),materiais_usados(*,materiais(denominacao,unidade_medida))&order=codigo_servico.desc"
        response = requests.get(url_serv, headers=headers)
        if response.status_code != 200 or not response.json():
            flash("Serviço não encontrado.")
            return redirect(url_for('listar_servicos'))
        servico = response.json()[0]
    except Exception as e:
        flash("Erro ao carregar serviço.")
        return redirect(url_for('listar_servicos'))
    def calcular_custo():
        try:
            url_mat = f"{SUPABASE_URL}/rest/v1/materiais_usados?select=valor_total&servico_id=eq.{id}"
            resp = requests.get(url_mat, headers=headers)
            if resp.status_code == 200:
                itens = resp.json()
                return sum(float(i['valor_total']) for i in itens)
            return 0.0
        except:
            return 0.0
    custo_materiais = calcular_custo()
    valor_cobrado = float(servico.get('valor_cobrado', 0) or 0)
    lucro = valor_cobrado - custo_materiais
    empresa_nome = servico['empresas']['nome_empresa'] if servico.get('empresas') else "Sem cliente"
    logo_url = "https://i.postimg.cc/RVqcJzzQ/logo.png"
    html = f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>OS {servico['codigo_servico']} - Impressão</title>
    <style>
    body {{ font-family: Arial, sans-serif; padding: 40px; color: #333; background: white; }}
    .header {{ text-align: center; margin-bottom: 20px; border-bottom: 2px solid #2c3e50; padding-bottom: 15px; }}
    .header img {{ max-width: 200px; margin-bottom: 10px; }}
    .header h1 {{ margin: 0; color: #2c3e50; font-size: 24px; text-transform: uppercase; letter-spacing: 1px; }}
    .info-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 15px; margin-bottom: 20px; }}
    .info-item strong {{ display: block; font-size: 14px; color: #555; }}
    table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
    th, td {{ border: 1px solid #ccc; padding: 8px; text-align: left; }}
    th {{ background-color: #ecf0f1; color: #2c3e50; }}
    .total-box {{ text-align: right; font-size: 16px; margin-top: 20px; }}
    .status {{ font-weight: bold; color: {'#27ae60' if servico['status'] == 'Concluído' else '#e67e22' if servico['status'] == 'Em Produção' else '#95a5a6'}; }}
    @media print {{ .no-print {{ display: none; }} body {{ background: white; }} }}
    .footer {{ margin-top: 40px; text-align: center; padding: 20px; border-top: 1px solid #ddd; font-size: 12px; color: #7f8c8d; }}
    </style>
    </head>
    <body>\n
    <div class="header">
    <img src="{logo_url}" alt="Logo da Empresa">
    <h1>ORDEM DE SERVIÇO</h1>
    <p><strong>Código:</strong> {servico['codigo_servico']}</p>
    </div>
    <div class="info-grid">
    <div class="info-item"><strong>Cliente:</strong> {empresa_nome}</div>
    <div class="info-item"><strong>Status:</strong> <span class="status">{servico['status']}</span></div>
    <div class="info-item"><strong>Título:</strong> {servico['titulo']}</div>
    <div class="info-item"><strong>Data de Abertura:</strong> {format_data(servico.get('data_abertura'))}</div>
    <div class="info-item"><strong>Previsão de Entrega:</strong> {format_data(servico.get('previsao_entrega'))}</div>
    <div class="info-item"><strong>Quantidade:</strong> {servico.get('quantidade', '-')}</div>
    <div class="info-item"><strong>Dimensão:</strong> {servico.get('dimensao', '-')}</div>
    <div class="info-item"><strong>Nº de Cores:</strong> {servico.get('numero_cores', '-')}</div>
    <div class="info-item"><strong>Aplicação:</strong> {servico.get('aplicacao', '-')}</div>
    <div class="info-item"><strong>Observações:</strong> {servico.get('observacoes', '-')}</div>
    </div>
    <h3>Materiais Utilizados</h3>
    <table>
    <thead><tr><th>Material</th><th>Unidade</th><th>Qtd Usada</th><th>Valor Unit.</th><th>Valor Total</th></tr></thead>
    <tbody>
    {''.join(f'''
    <tr>
    <td>{m['materiais']['denominacao']}</td>
    <td>{m['materiais']['unidade_medida']}</td>
    <td>{m['quantidade_usada']}</td>
    <td>R$ {m['valor_unitario']:.2f}</td>
    <td>R$ {m['valor_total']:.2f}</td>
    </tr>
    ''' for m in servico.get('materiais_usados', []) if m.get('materiais'))}
    </tbody>
    </table>
    <div class="total-box">
    <p><strong>Custo com Materiais:</strong> R$ {custo_materiais:.2f}</p>
    <p><strong>Valor Cobrado:</strong> R$ {valor_cobrado:.2f}</p>
    <p><strong>Lucro Estimado:</strong> R$ {lucro:.2f}</p>
    </div>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    <div style="text-align: center; margin-top: 40px;">
    <button onclick="window.print()" class="no-print" style="padding: 12px 20px; background: #27ae60; color: white; border: none; border-radius: 8px; cursor: pointer;">🖨️ Imprimir</button>
    <a href="/pdf_os/{id}" class="no-print" style="margin-left: 10px; padding: 12px 20px; background: #e67e22; color: white; text-decoration: none; border-radius: 8px;">📄 Gerar PDF</a>
    <a href="/servicos" class="no-print" style="margin-left: 10px; color: #3498db;">← Voltar</a>
    </div>
    </body>
    </html>
    '''
    return html

@app.route('/pdf_os/<int:id>')
def pdf_os(id):
    if 'usuario' not in session:
        return redirect(url_for('login'))
    try:
        url_serv = f"{SUPABASE_URL}/rest/v1/servicos?id=eq.{id}&select=*,empresas(nome_empresa),materiais_usados(*,materiais(denominacao,unidade_medida))&order=codigo_servico.desc"
        response = requests.get(url_serv, headers=headers)
        if response.status_code != 200 or not response.json():
            flash("Serviço não encontrado.")
            return redirect(url_for('listar_servicos'))
        servico = response.json()[0]
    except Exception as e:
        flash("Erro ao carregar serviço.")
        return redirect(url_for('listar_servicos'))
    def calcular_custo():
        try:
            url_mat = f"{SUPABASE_URL}/rest/v1/materiais_usados?select=valor_total&servico_id=eq.{id}"
            resp = requests.get(url_mat, headers=headers)
            if resp.status_code == 200:
                itens = resp.json()
                return sum(float(i['valor_total']) for i in itens)
            return 0.0
        except:
            return 0.0
    custo_materiais = calcular_custo()
    valor_cobrado = float(servico.get('valor_cobrado', 0) or 0)
    lucro = valor_cobrado - custo_materiais
    empresa_nome = servico['empresas']['nome_empresa'] if servico.get('empresas') else "Sem cliente"
    logo_url = "https://i.postimg.cc/RVqcJzzQ/logo.png"
    html = f'''
    <!DOCTYPE html>
    <html>
    <head>
    <meta charset="UTF-8">
    <title>OS {servico['codigo_servico']}</title>
    <style>
    body {{ font-family: Arial, sans-serif; padding: 40px; background: white; }}
    .header {{ text-align: center; margin-bottom: 20px; border-bottom: 2px solid #2c3e50; padding-bottom: 15px; }}
    .header img {{ max-width: 200px; margin-bottom: 10px; }}
    .header h1 {{ margin: 0; color: #2c3e50; font-size: 24px; text-transform: uppercase; letter-spacing: 1px; }}
    table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
    th, td {{ border: 1px solid #ccc; padding: 8px; text-align: left; }}
    th {{ background-color: #ecf0f1; color: #2c3e50; }}
    .total-box {{ text-align: right; font-size: 18px; margin-top: 20px; }}
    .footer {{ margin-top: 40px; text-align: center; padding: 20px; border-top: 1px solid #ddd; font-size: 12px; color: #7f8c8d; }}
    </style>
    </head>
    <body>\n
    <div class="header">
    <img src="{logo_url}" alt="Logo da Empresa">
    <h1>ORDEM DE SERVIÇO</h1>
    <p><strong>Código:</strong> {servico['codigo_servico']}</p>
    </div>
    <table>
    <tr><th>Cliente</th><td>{empresa_nome}</td></tr>
    <tr><th>Título</th><td>{servico['titulo']}</td></tr>
    <tr><th>Status</th><td>{servico['status']}</td></tr>
    <tr><th>Quantidade</th><td>{servico.get('quantidade', '—')}</td></tr>
    <tr><th>Dimensão</th><td>{servico.get('dimensao', '—')}</td></tr>
    <tr><th>Nº de Cores</th><td>{servico.get('numero_cores', '—')}</td></tr>
    <tr><th>Aplicação</th><td>{servico.get('aplicacao', '—')}</td></tr>
    <tr><th>Data Abertura</th><td>{format_data(servico.get('data_abertura'))}</td></tr>
    <tr><th>Previsão Entrega</th><td>{format_data(servico.get('previsao_entrega'))}</td></tr>
    <tr><th>Valor Cobrado</th><td>R$ {valor_cobrado:.2f}</td></tr>
    <tr><th>Custo Materiais</th><td>R$ {custo_materiais:.2f}</td></tr>
    <tr><th>Lucro</th><td>R$ {lucro:.2f}</td></tr>
    <tr><th>Observações</th><td>{servico.get('observacoes', '—')}</td></tr>
    </table>
    <h3>Materiais Utilizados</h3>
    <table>
    <thead><tr><th>Material</th><th>Qtd</th><th>Valor Unit.</th><th>Valor Total</th></tr></thead>
    <tbody>
    {''.join(f"""
    <tr>
    <td>{m['materiais']['denominacao']}</td>
    <td>{m['quantidade_usada']}</td>
    <td>R$ {m['valor_unitario']:.2f}</td>
    <td>R$ {m['valor_total']:.2f}</td>
    </tr>
    """ for m in servico.get('materiais_usados', []) if m.get('materiais'))}
    </tbody>
    </table>
    <div class="total-box"><p><strong>Lucro Final:</strong> R$ {lucro:.2f}</p></div>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </body>
    </html>
    '''
    pdf = pdfkit.from_string(html, False)
    return send_file(BytesIO(pdf), as_attachment=True, download_name=f"os_{servico['codigo_servico']}.pdf", mimetype="application/pdf")

@app.route('/configuracoes')
def configuracoes():
    if 'usuario' not in session or session['nivel'] != 'administrador':
        flash("Acesso negado!")
        return redirect(url_for('clientes'))
    config = buscar_configuracoes()
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Configurações do Sistema</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f5f7fa; color: #333; min-height: 100vh; padding: 0; margin: 0; }}
    .container {{ max-width: 800px; margin: 30px auto; background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0,0,0,0.1); overflow: hidden; }}
    .header {{ background: #2c3e50; color: white; text-align: center; padding: 30px; }}
    h1 {{ font-size: 28px; margin: 0; font-weight: 600; }}
    .user-info {{ background: #34495e; color: white; padding: 15px 20px; font-size: 15px; display: flex; justify-content: space-between; align-items: center; }}
    .form-container {{ padding: 30px; }}
    .form-container label {{ display: block; margin: 10px 0 5px 0; font-weight: 600; color: #2c3e50; }}
    .form-container input {{ width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 6px; font-size: 14px; }}
    .btn {{ padding: 12px 20px; background: #27ae60; color: white; border: none; border-radius: 8px; font-size: 16px; font-weight: 600; cursor: pointer; }}
    .back-link {{ display: inline-block; margin: 20px 30px; color: #3498db; text-decoration: none; font-weight: 500; }}
    .footer {{ text-align: center; padding: 20px; background: #ecf0f1; color: #7f8c8d; font-size: 13px; border-top: 1px solid #bdc3c7; }}
    </style>
    </head>
    <body>\n
    <div class="container">
    <div class="header"><h1>⚙️ Configurações do Sistema</h1></div>
    <div class="user-info"><span>👤 {session['usuario']} ({session['nivel'].upper()})</span><a href="/logout">🚪 Sair</a></div>
    {MENU_FLUTUANTE}
    <form method="post" action="/salvar_configuracoes" class="form-container">
    <h3>Remetente (Etiquetas)</h3>
    <div><label>Nome do Remetente</label><input type="text" name="nome_remetente" value="{config['nome_remetente']}" required></div>
    <div><label>Endereço Completo</label><input type="text" name="endereco_remetente" value="{config['endereco_remetente']}" required></div>
    <div><label>Bairro</label><input type="text" name="bairro_remetente" value="{config['bairro_remetente']}" required></div>
    <div><label>Cidade</label><input type="text" name="cidade_remetente" value="{config['cidade_remetente']}" required></div>
    <div><label>Estado</label><input type="text" name="estado_remetente" value="{config['estado_remetente']}" required maxlength="2"></div>
    <div><label>CEP</label><input type="text" name="cep_remetente" value="{config['cep_remetente']}" required></div>
    <button type="submit" class="btn">💾 Salvar Configurações</button>
    </form>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </div>
    </body>
    </html>
    '''

@app.route('/salvar_configuracoes', methods=['POST'])
def salvar_configuracoes_view():
    if 'usuario' not in session or session['nivel'] != 'administrador':
        flash("Acesso negado!")
        return redirect(url_for('clientes'))
    config = {
        "nome_remetente": request.form.get('nome_remetente'),
        "endereco_remetente": request.form.get('endereco_remetente'),
        "bairro_remetente": request.form.get('bairro_remetente'),
        "cidade_remetente": request.form.get('cidade_remetente'),
        "estado_remetente": request.form.get('estado_remetente'),
        "cep_remetente": request.form.get('cep_remetente')
    }
    if salvar_configuracoes(config):
        flash("✅ Configurações salvas com sucesso!")
    else:
        flash("❌ Erro ao salvar configurações.")
    return redirect(url_for('configuracoes'))

@app.route('/materiais')
def listar_materiais():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    busca = request.args.get('q', '').strip()
    try:
        if busca:
            url = f"{SUPABASE_URL}/rest/v1/materiais?denominacao=ilike.*{busca}*"
        else:
            url = f"{SUPABASE_URL}/rest/v1/materiais?select=*"
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            materiais = response.json()
        else:
            flash("Erro ao carregar materiais.")
            materiais = []
    except Exception as e:
        flash("Erro de conexão.")
        materiais = []
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Materiais Cadastrados</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f5f7fa; color: #333; min-height: 100vh; padding: 0; margin: 0; }}
    .container {{ max-width: 1100px; margin: 30px auto; background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0, 0, 0, 0.1); overflow: hidden; }}
    .header {{ background: #2c3e50; color: white; text-align: center; padding: 30px; }}
    h1 {{ font-size: 28px; margin: 0; font-weight: 600; }}
    .user-info {{ background: #34495e; color: white; padding: 15px 20px; font-size: 15px; display: flex; justify-content: space-between; align-items: center; }}
    .search-box {{ padding: 20px 30px; text-align: center; }}
    .search-box input {{ width: 70%; padding: 12px; border: 1px solid #ddd; border-radius: 8px; font-size: 16px; }}
    table {{ width: 100%; border-collapse: collapse; }}
    th, td {{ padding: 16px 20px; text-align: left; }}
    th {{ background: #ecf0f1; color: #2c3e50; font-weight: 600; text-transform: uppercase; font-size: 14px; }}
    tr:nth-child(even) {{ background: #f9f9f9; }}
    tr:hover {{ background: #f1f7fb; }}
    .back-link {{ display: inline-block; margin: 20px 30px; color: #3498db; text-decoration: none; font-weight: 500; }}
    .footer {{ text-align: center; padding: 20px; background: #ecf0f1; color: #7f8c8d; font-size: 13px; border-top: 1px solid #bdc3c7; }}
    </style>
    </head>
    <body>\n
    <div class="container">
    <div class="header"><h1>📦 Materiais Cadastrados</h1></div>
    <div class="user-info"><span>👤 {session['usuario']} ({session['nivel'].upper()})</span><a href="/logout">🚪 Sair</a></div>
    {MENU_FLUTUANTE}
    <a href="/cadastrar_material" class="btn">➕ Cadastrar Novo Material</a>
    <div class="search-box"><form method="get" style="display: inline;"><input type="text" name="q" placeholder="Pesquisar por denominação..." value="{busca}"><button type="submit" style="padding: 10px 20px; background: #3498db; color: white; border: none; border-radius: 8px; cursor: pointer;">🔍 Pesquisar</button></form></div>
    <table>
    <thead><tr><th>ID</th><th>Denominação</th><th>Marca</th><th>Grupo</th><th>Unidade</th><th>Fornecedor</th><th>Ações</th></tr></thead>
    <tbody>
    {''.join(f'''
    <tr>
    <td>{m["id"]}</td>
    <td><a href="/material/{m["id"]}" style="color: #3498db; text-decoration: none;">{m["denominacao"]}</a></td>
    <td>{m["marca"] or "—"}</td>
    <td>{m["grupo_material"] or "—"}</td>
    <td>{m["unidade_medida"]}</td>
    <td>{m["fornecedor"] or "—"}</td>
    <td>
    {f'<a href="/editar_material/{m["id"]}" style="color: #f39c12; text-decoration: none;">✏️ Editar</a>' if session['nivel'] in ['administrador', 'vendedor'] else ''}
    {f'<a href="/excluir_material/{m["id"]}" style="color: #e74c3c; text-decoration: none; margin-left: 10px;" onclick="return confirm(\'Tem certeza que deseja excluir?\')">🗑️ Excluir</a>' if session['nivel'] == 'administrador' else ''}
    </td>
    </tr>
    ''' for m in materiais)}
    </tbody>
    </table>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </div>
    </body>
    </html>
    '''

@app.route('/material/<int:id>')
def detalhes_material(id):
    if 'usuario' not in session:
        return redirect(url_for('login'))
    try:
        url = f"{SUPABASE_URL}/rest/v1/materiais?id=eq.{id}"
        response = requests.get(url, headers=headers)
        if response.status_code != 200 or not response.json():
            flash("Material não encontrado.")
            return redirect(url_for('listar_materiais'))
        material = response.json()[0] if response.json() else None
        if not material:
            flash("Material não encontrado.")
            return redirect(url_for('listar_materiais'))
    except Exception as e:
        flash("Erro ao carregar material.")
        return redirect(url_for('listar_materiais'))
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{material['denominacao']} - Detalhes</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f5f7fa; color: #333; min-height: 100vh; padding: 0; margin: 0; }}
    .container {{ max-width: 800px; margin: 30px auto; background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0, 0, 0, 0.1); overflow: hidden; }}
    .header {{ background: #2c3e50; color: white; text-align: center; padding: 30px; }}
    h1 {{ font-size: 28px; margin: 0; font-weight: 600; }}
    .user-info {{ background: #34495e; color: white; padding: 15px 20px; font-size: 15px; display: flex; justify-content: space-between; align-items: center; }}
    .details {{ padding: 30px; }}
    .details p {{ margin: 10px 0; font-size: 16px; }}
    .details strong {{ color: #2c3e50; }}
    .btn {{ padding: 12px 20px; background: #27ae60; color: white; border: none; border-radius: 8px; font-size: 16px; font-weight: 600; text-decoration: none; margin: 10px 30px; }}
    .back-link {{ display: inline-block; margin: 20px 30px; color: #3498db; text-decoration: none; font-weight: 500; }}
    .footer {{ text-align: center; padding: 20px; background: #ecf0f1; color: #7f8c8d; font-size: 13px; border-top: 1px solid #bdc3c7; }}
    </style>
    </head>
    <body>\n
    <div class="container">
    <div class="header"><h1>📦 {material['denominacao']}</h1></div>
    <div class="user-info"><span>👤 {session['usuario']} ({session['nivel'].upper()})</span><a href="/logout">🚪 Sair</a></div>
    <a href="/materiais" class="back-link">← Voltar à Lista</a>
    <div class="details">
    <p><strong>Marca:</strong> {material['marca'] or "—"}</p>
    <p><strong>Grupo de Material:</strong> {material['grupo_material'] or "—"}</p>
    <p><strong>Unidade de Medida:</strong> {material['unidade_medida']}</p>
    <p><strong>Valor Unitário:</strong> R$ {material['valor_unitario']:.2f}</p>
    <p><strong>Especificação:</strong> {material['especificacao'] or "—"}</p>
    <p><strong>Fornecedor:</strong> {material['fornecedor'] or "—"}</p>
    </div>
    <div style="display: flex; gap: 15px; margin: 20px 0;">
    <a href="/editar_material/{id}" class="btn" style="background: #f39c12;">✏️ Editar Material</a>
    <a href="/excluir_material/{id}" class="btn" style="background: #e74c3c;" onclick="return confirm('Tem certeza que deseja excluir este material?')">🗑️ Excluir Material</a>
    </div>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </div>
    </body>
    </html>
    '''

@app.route('/editar_material/<int:id>', methods=['GET', 'POST'])
def editar_material(id):
    if 'usuario' not in session:
        return redirect(url_for('login'))
    if session['nivel'] not in ['administrador', 'vendedor']:
        flash("Acesso negado!")
        return redirect(url_for('listar_materiais'))
    try:
        url = f"{SUPABASE_URL}/rest/v1/materiais?id=eq.{id}"
        response = requests.get(url, headers=headers)
        if response.status_code != 200 or not response.json():
            flash("Material não encontrado.")
            return redirect(url_for('listar_materiais'))
        material = response.json()[0]
    except Exception as e:
        flash("Erro ao carregar material.")
        return redirect(url_for('listar_materiais'))
    if request.method == 'POST':
        denominacao = request.form.get('denominacao')
        marca = request.form.get('marca')
        grupo_material = request.form.get('grupo_material')
        unidade_medida = request.form.get('unidade_medida')
        unidade_outro = request.form.get('unidade_outro')
        valor_unitario = request.form.get('valor_unitario')
        especificacao = request.form.get('especificacao')
        fornecedor_id = request.form.get('fornecedor_id')
        if unidade_medida == 'outro' and unidade_outro:
            unidade_medida = unidade_outro.strip()
        elif not unidade_medida:
            flash("Unidade de Medida é obrigatória!")
            return redirect(request.url)
        if not denominacao or not valor_unitario:
            flash("Denominação e Valor Unitário são obrigatórios!")
            return redirect(request.url)
        try:
            valor_unitario = float(valor_unitario)
        except:
            flash("Valor unitário deve ser um número!")
            return redirect(request.url)
        try:
            url = f"{SUPABASE_URL}/rest/v1/materiais?id=eq.{id}"
            dados = {
                "denominacao": denominacao,
                "marca": marca,
                "grupo_material": grupo_material,
                "unidade_medida": unidade_medida,
                "valor_unitario": valor_unitario,
                "especificacao": especificacao,
                "fornecedor": None
            }
            if fornecedor_id:
                fornecedores = buscar_fornecedores()
                fornecedor = next((f for f in fornecedores if f['id'] == int(fornecedor_id)), None)
                if fornecedor:
                    dados["fornecedor"] = fornecedor["nome"]
            response = requests.patch(url, json=dados, headers=headers)
            if response.status_code == 204:
                flash("✅ Material atualizado com sucesso!")
                return redirect(url_for('detalhes_material', id=id))
            else:
                flash("❌ Erro ao atualizar material.")
        except Exception as e:
            flash("❌ Erro de conexão.")
        return redirect(request.url)
    fornecedores = buscar_fornecedores()
    fornecedor_selecionado = None
    if material.get('fornecedor'):
        fornecedor_selecionado = next((f for f in fornecedores if f['nome'] == material['fornecedor']), None)
    def get_selected_attr(f_id):
        if fornecedor_selecionado and f_id == fornecedor_selecionado['id']:
            return 'selected'
        return ''
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Editar Material</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f5f7fa; color: #333; min-height: 100vh; padding: 0; margin: 0; }}
    .container {{ max-width: 800px; margin: 30px auto; background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0, 0, 0, 0.1); overflow: hidden; }}
    .header {{ background: #2c3e50; color: white; text-align: center; padding: 30px; }}
    h1 {{ font-size: 28px; margin: 0; font-weight: 600; }}
    .user-info {{ background: #34495e; color: white; padding: 15px 20px; font-size: 15px; display: flex; justify-content: space-between; align-items: center; }}
    .form-container {{ padding: 30px; }}
    .form-container label {{ display: block; margin: 10px 0 5px 0; font-weight: 600; color: #2c3e50; }}
    .form-container input, .form-container select, .form-container textarea {{ width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 6px; font-size: 14px; }}
    .btn {{ padding: 12px 20px; background: #27ae60; color: white; border: none; border-radius: 8px; font-size: 16px; font-weight: 600; cursor: pointer; }}
    .back-link {{ display: inline-block; margin: 20px 30px; color: #3498db; text-decoration: none; font-weight: 500; }}
    .footer {{ text-align: center; padding: 20px; background: #ecf0f1; color: #7f8c8d; font-size: 13px; border-top: 1px solid #bdc3c7; }}
    </style>
    </head>
    <body>\n
    <div class="container">
    <div class="header"><h1>✏️ Editar {material['denominacao']}</h1></div>
    <div class="user-info"><span>👤 {session['usuario']} ({session['nivel'].upper()})</span><a href="/logout">🚪 Sair</a></div>
    <a href="/material/{id}" class="back-link">← Voltar aos Detalhes</a>
    <form method="post" class="form-container">
    <div><label>Denominação *</label><input type="text" name="denominacao" value="{material['denominacao']}" required></div>
    <div><label>Marca</label><input type="text" name="marca" value="{material['marca']}"></div>
    <div><label>Grupo de Material</label><input type="text" name="grupo_material" value="{material['grupo_material']}"></div>
    <div>
    <label>Unidade de Medida *</label>
    <select name="unidade_medida" id="unidade_medida" onchange="toggleOutro()" required>
    <option value="">Selecione</option>
    <option value="folha" {"selected" if material['unidade_medida'] == 'folha' else ""}>folha</option>
    <option value="metro" {"selected" if material['unidade_medida'] == 'metro' else ""}>metro</option>
    <option value="centímetro" {"selected" if material['unidade_medida'] == 'centímetro' else ""}>centímetro</option>
    <option value="milímetro" {"selected" if material['unidade_medida'] == 'milímetro' else ""}>milímetro</option>
    <option value="grama" {"selected" if material['unidade_medida'] == 'grama' else ""}>grama</option>
    <option value="quilograma" {"selected" if material['unidade_medida'] == 'quilograma' else ""}>quilograma</option>
    <option value="rolo" {"selected" if material['unidade_medida'] == 'rolo' else ""}>rolo</option>
    <option value="litro" {"selected" if material['unidade_medida'] == 'litro' else ""}>litro</option>
    <option value="unidade" {"selected" if material['unidade_medida'] == 'unidade' else ""}>unidade</option>
    <option value="conjunto" {"selected" if material['unidade_medida'] == 'conjunto' else ""}>conjunto</option>
    <option value="m²" {"selected" if material['unidade_medida'] == 'm²' else ""}>m²</option>
    <option value="cm²" {"selected" if material['unidade_medida'] == 'cm²' else ""}>cm²</option>
    <option value="outro" {"selected" if material['unidade_medida'] == 'outro' else ""}>Outro (especifique)</option>
    </select>
    <input type="text" name="unidade_outro" id="unidade_outro" placeholder="Digite a unidade" style="display: none; margin-top: 10px;" oninput="this.value = this.value.toLowerCase()" value="{material['unidade_medida'] if material['unidade_medida'] not in ['folha', 'metro', 'centímetro', 'milímetro', 'grama', 'quilograma', 'rolo', 'litro', 'unidade', 'conjunto', 'm²', 'cm²'] else ''}">
    </div>
    <div><label>Valor Unitário *</label><input type="number" name="valor_unitario" step="0.01" value="{material['valor_unitario']}" required></div>
    <div><label>Especificação</label><textarea name="especificacao" rows="3">{material['especificacao']}</textarea></div>
    <div>
    <label>Fornecedor</label>
    <select name="fornecedor_id" id="fornecedor_id">
    <option value="">Selecione um fornecedor</option>
    {''.join(f'<option value="{f["id"]}" {get_selected_attr(f["id"])}>{f["nome"]}</option>' for f in fornecedores)}
    </select>
    </div>
    <button type="submit" class="btn">💾 Salvar Alterações</button>
    </form>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </div>
    <script>
    function toggleOutro() {{
        const select = document.getElementById('unidade_medida');
        const input = document.getElementById('unidade_outro');
        if (select.value === 'outro') {{
            input.style.display = 'block';
            input.setAttribute('required', 'required');
        }} else {{
            input.style.display = 'none';
            input.removeAttribute('required');
        }}
    }}
    window.onload = function() {{
        const select = document.getElementById('unidade_medida');
        if (select.value === 'outro') {{
            document.getElementById('unidade_outro').style.display = 'block';
        }}
    }};
    </script>
    </body>
    </html>
    '''

@app.route('/cadastrar_material', methods=['GET', 'POST'])
def cadastrar_material():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    if request.method == 'POST':
        denominacao = request.form.get('denominacao')
        marca = request.form.get('marca')
        grupo_material = request.form.get('grupo_material')
        unidade_medida = request.form.get('unidade_medida')
        unidade_outro = request.form.get('unidade_outro')
        valor_unitario = request.form.get('valor_unitario')
        especificacao = request.form.get('especificacao')
        fornecedor_id = request.form.get('fornecedor_id')
        if unidade_medida == 'outro' and unidade_outro:
            unidade_medida = unidade_outro.strip()
        elif not unidade_medida:
            flash("Unidade de Medida é obrigatória!")
            return redirect(request.url)
        if not denominacao or not valor_unitario:
            flash("Denominação e Valor Unitário são obrigatórios!")
            return redirect(request.url)
        try:
            valor_unitario = float(valor_unitario)
        except:
            flash("Valor unitário deve ser um número!")
            return redirect(request.url)
        try:
            url = f"{SUPABASE_URL}/rest/v1/materiais"
            dados = {
                "denominacao": denominacao,
                "marca": marca,
                "grupo_material": grupo_material,
                "unidade_medida": unidade_medida,
                "valor_unitario": valor_unitario,
                "especificacao": especificacao,
                "fornecedor": None
            }
            if fornecedor_id:
                fornecedores = buscar_fornecedores()
                fornecedor = next((f for f in fornecedores if f['id'] == int(fornecedor_id)), None)
                if fornecedor:
                    dados["fornecedor"] = fornecedor["nome"]
            response = requests.post(url, json=dados, headers=headers)
            if response.status_code == 201:
                flash("✅ Material cadastrado com sucesso!")
                return redirect(url_for('listar_materiais'))
            else:
                flash("❌ Erro ao cadastrar material.")
        except Exception as e:
            flash("❌ Erro de conexão.")
        return redirect(request.url)
    fornecedores = buscar_fornecedores()
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Cadastrar Material</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: #333; min-height: 100vh; padding: 0; margin: 0; }}
    .container {{ max-width: 800px; margin: 30px auto; background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0,0,0,0.1); overflow: hidden; }}
    .header {{ background: #2c3e50; color: white; text-align: center; padding: 30px; }}
    h1 {{ font-size: 28px; margin: 0; font-weight: 600; }}
    .user-info {{ background: #34495e; color: white; padding: 15px 20px; font-size: 15px; display: flex; justify-content: space-between; align-items: center; }}
    .form-container {{ padding: 30px; }}
    .form-container label {{ display: block; margin: 10px 0 5px 0; font-weight: 600; color: #2c3e50; }}
    .form-container input, .form-container select, .form-container textarea {{ width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 6px; font-size: 14px; }}
    .btn {{ padding: 12px 20px; background: #27ae60; color: white; border: none; border-radius: 8px; font-size: 16px; font-weight: 600; cursor: pointer; }}
    .back-link {{ display: inline-block; margin: 20px 30px; color: #3498db; text-decoration: none; font-weight: 500; }}
    .footer {{ text-align: center; padding: 20px; background: #ecf0f1; color: #7f8c8d; font-size: 13px; border-top: 1px solid #bdc3c7; }}
    </style>
    </head>
    <body>\n
    <div class="container">
    <div class="header"><h1>➕ Cadastrar Novo Material</h1></div>
    <div class="user-info"><span>👤 {session['usuario']} ({session['nivel'].upper()})</span><a href="/logout">🚪 Sair</a></div>
    <a href="/materiais" class="back-link">← Voltar à Lista</a>
    <form method="post" class="form-container">
    <div><label>Denominação *</label><input type="text" name="denominacao" required></div>
    <div><label>Marca</label><input type="text" name="marca"></div>
    <div><label>Grupo de Material</label><input type="text" name="grupo_material"></div>
    <div>
    <label>Unidade de Medida *</label>
    <select name="unidade_medida" id="unidade_medida" onchange="toggleOutro()" required>
    <option value="">Selecione</option>
    <option value="folha">folha</option><option value="metro">metro</option><option value="centímetro">centímetro</option><option value="milímetro">milímetro</option><option value="grama">grama</option><option value="quilograma">quilograma</option><option value="rolo">rolo</option><option value="litro">litro</option><option value="unidade">unidade</option><option value="conjunto">conjunto</option><option value="m²">m²</option><option value="cm²">cm²</option><option value="outro">Outro (especifique)</option>
    </select>
    <input type="text" name="unidade_outro" id="unidade_outro" placeholder="Digite a unidade" style="display: none; margin-top: 10px;" oninput="this.value = this.value.toLowerCase()">
    </div>
    <div><label>Valor Unitário *</label><input type="number" name="valor_unitario" step="0.01" required></div>
    <div><label>Especificação</label><textarea name="especificacao" rows="3"></textarea></div>
    <div>
    <label>Fornecedor</label>
    <select name="fornecedor_id" id="fornecedor_id">
    <option value="">Selecione um fornecedor</option>
    {''.join(f'<option value="{f["id"]}">{f["nome"]}</option>' for f in fornecedores)}
    </select>
    </div>
    <button type="submit" class="btn">💾 Salvar Material</button>
    </form>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </div>
    <script>
    function toggleOutro() {{
        const select = document.getElementById('unidade_medida');
        const input = document.getElementById('unidade_outro');
        if (select.value === 'outro') {{
            input.style.display = 'block';
            input.setAttribute('required', 'required');
        }} else {{
            input.style.display = 'none';
            input.removeAttribute('required');
        }}
    }}
    </script>
    </body>
    </html>
    '''

@app.route('/excluir_material/<int:id>')
def excluir_material(id):
    if 'usuario' not in session:
        return redirect(url_for('login'))
    try:
        url = f"{SUPABASE_URL}/rest/v1/materiais?id=eq.{id}"
        response = requests.delete(url, headers=headers)
        if response.status_code == 204:
            flash("🗑️ Material excluído com sucesso!")
        else:
            flash("❌ Erro ao excluir material.")
    except Exception as e:
        flash("❌ Erro de conexão.")
    return redirect(url_for('listar_materiais'))

@app.route('/estoque')
def estoque():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    if session['nivel'] != 'administrador':
        flash("Acesso negado!")
        return redirect(url_for('clientes'))
    busca_mov = request.args.get('q', '').strip()
    try:
        materiais_catalogo = buscar_materiais()
        saldo = calcular_estoque_atual()
        materiais_em_estoque = []
        for m in materiais_catalogo:
            qtd = saldo.get(m['id'], 0)
            m['quantidade_estoque'] = qtd
            materiais_em_estoque.append(m)
        movimentacoes = buscar_movimentacoes_com_materiais(busca_mov)
    except Exception as e:
        flash("Erro de conexão.")
        materiais_em_estoque = []
        movimentacoes = []
    movimentacoes_html = ""
    for m in movimentacoes:
        data = format_data(m.get("data_movimentacao"))
        tipo = m["tipo"]
        classe_tipo = "tipo-entrada" if tipo == "entrada" else "tipo-saida"
        material_info = m.get("materiais")
        if material_info is None:
            nome_material = "<em>Material excluído</em>"
            unidade = ""
        else:
            nome_material = material_info.get("denominacao", "Desconhecido")
            unidade = material_info.get("unidade_medida", "")
        valor_unitario = m.get("valor_unitario", 0.0) or 0.0
        valor_total = m.get("valor_total", 0.0) or 0.0
        qtd = m.get("quantidade", 0) or 0
        acoes = f'<a href="/excluir_movimentacao/{m["id"]}" class="btn btn-delete" onclick="return confirm(\'Tem certeza que deseja excluir?\')">🗑️ Excluir</a>'
        movimentacoes_html += f'''
        <tr data-id="{m["id"]}">
        <td>{data}</td>
        <td>{nome_material}</td>
        <td class="{classe_tipo}">{tipo.upper()}</td>
        <td>{qtd} {unidade}</td>
        <td>R$ {valor_unitario:.2f}</td>
        <td>R$ {valor_total:.2f}</td>
        <td>{acoes}</td>
        </tr>
        '''
    materiais_html = ""
    for m in materiais_em_estoque:
        classe_estoque = "estoque-baixo" if m["quantidade_estoque"] < 5 else ""
        materiais_html += f'''
        <tr>
        <td>{m["id"]}</td>
        <td>{m["denominacao"]}</td>
        <td>{m["unidade_medida"]}</td>
        <td class="{classe_estoque}">{m["quantidade_estoque"]}</td>
        <td>
        {f'<a href="/registrar_entrada_form?material_id={m["id"]}" class="btn btn-green">📥 Entrada</a>' if session['nivel'] == 'administrador' else ''}
        {f'<a href="/registrar_saida_form?material_id={m["id"]}" class="btn btn-red">📤 Saída</a>' if session['nivel'] == 'administrador' else ''}
        {f'<a href="/editar_material/{m["id"]}" class="btn btn-edit">✏️ Editar</a>' if session['nivel'] in ['administrador', 'vendedor'] else ''}
        </td>
        </tr>
        '''
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Meu Estoque</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f5f7fa; color: #333; min-height: 100vh; padding: 0; margin: 0; }}
    .container {{ max-width: 1200px; margin: 30px auto; background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0, 0, 0, 0.1); overflow: hidden; }}
    .header {{ background: #2c3e50; color: white; text-align: center; padding: 30px; }}
    h1 {{ font-size: 28px; margin: 0; font-weight: 600; }}
    .user-info {{ background: #34495e; color: white; padding: 15px 20px; font-size: 15px; display: flex; justify-content: space-between; align-items: center; }}
    .section {{ padding: 20px 30px; }}
    .section-title {{ font-size: 20px; margin: 0 0 15px 0; color: #2c3e50; border-bottom: 1px solid #ddd; padding-bottom: 10px; }}
    .search-box {{ text-align: center; margin-bottom: 20px; }}
    .search-box input {{ width: 70%; padding: 12px; border: 1px solid #ddd; border-radius: 8px; font-size: 16px; }}
    table {{ width: 100%; border-collapse: collapse; }}
    th, td {{ padding: 12px 15px; text-align: left; }}
    th {{ background: #ecf0f1; color: #2c3e50; font-weight: 600; text-transform: uppercase; font-size: 14px; }}
    tr:nth-child(even) {{ background: #f9f9f9; }}
    tr:hover {{ background: #f1f7fb; }}
    .back-link {{ display: inline-block; margin: 20px 30px; color: #3498db; text-decoration: none; font-weight: 500; }}
    .btn {{ padding: 8px 12px; border: none; border-radius: 6px; font-size: 14px; cursor: pointer; text-decoration: none; margin-right: 5px; }}
    .btn-green {{ background: #27ae60; color: white; }} .btn-red {{ background: #e74c3c; color: white; }} .btn-delete {{ background: #95a5a6; color: white; }} .btn-edit {{ background: #f39c12; color: white; }}
    .estoque-baixo {{ color: #e74c3c; font-weight: bold; }} .tipo-entrada {{ color: #27ae60; font-weight: bold; }} .tipo-saida {{ color: #e74c3c; font-weight: bold; }}
    .footer {{ text-align: center; padding: 20px; background: #ecf0f1; color: #7f8c8d; font-size: 13px; border-top: 1px solid #bdc3c7; }}
    </style>
    </head>
    <body>\n
    <div class="container">
    <div class="header"><h1>📊 Meu Estoque</h1></div>
    <div class="user-info"><span>👤 {session['usuario']} ({session['nivel'].upper()})</span><a href="/logout">🚪 Sair</a></div>
    {MENU_FLUTUANTE}
    <div class="section">
    <h2 class="section-title">Adicionar ao Estoque</h2>
    <p style="margin: 10px 0;"><a href="/registrar_entrada_form" class="btn btn-green">➕ Registrar Nova Entrada</a><a href="/cadastrar_material" class="btn btn-blue">📦 Cadastrar Novo Material</a></p>
    </div>
    <div class="section">
    <h2 class="section-title">Itens em Estoque</h2>
    <table><thead><tr><th>ID</th><th>Material</th><th>Unidade</th><th>Qtd. em Estoque</th><th>Ações</th></tr></thead><tbody>{materiais_html}</tbody></table>
    </div>
    <div class="section">
    <h2 class="section-title">Últimas Movimentações</h2>
    <div class="search-box"><form method="get" style="display: inline;"><input type="text" name="q" placeholder="Pesquisar por material..." value="{busca_mov}"><button type="submit" style="padding: 10px 20px; background: #3498db; color: white; border: none; border-radius: 8px; cursor: pointer;">🔍 Pesquisar</button></form></div>
    <table><thead><tr><th>Data</th><th>Material</th><th>Tipo</th><th>Quantidade</th><th>Valor Unit.</th><th>Valor Total</th><th>Ações</th></tr></thead><tbody>{movimentacoes_html}</tbody></table>
    </div>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </div>
    </body>
    </html>
    '''

@app.route('/registrar_entrada_form')
def registrar_entrada_form():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    if session['nivel'] != 'administrador':
        flash("Acesso negado!")
        return redirect(url_for('clientes'))
    material_id = request.args.get('material_id')
    material = None
    try:
        materiais = buscar_materiais()
        if material_id:
            material = next((m for m in materiais if m['id'] == int(material_id)), None)
    except:
        flash("Erro ao carregar material.")
        return redirect(url_for('estoque'))
    import json
    materiais_js = json.dumps(materiais, ensure_ascii=False)
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Registrar Entrada</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f5f7fa; color: #333; min-height: 100vh; padding: 0; margin: 0; }}
    .container {{ max-width: 900px; margin: 30px auto; background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0, 0, 0, 0.1); overflow: hidden; }}
    .header {{ background: #2c3e50; color: white; text-align: center; padding: 30px; }}
    h1 {{ font-size: 28px; margin: 0; font-weight: 600; }}
    .user-info {{ background: #34495e; color: white; padding: 15px 20px; font-size: 15px; display: flex; justify-content: space-between; align-items: center; }}
    .form-container {{ padding: 30px; }}
    .grid-2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 15px; }}
    .form-container label {{ display: block; margin: 10px 0 5px 0; font-weight: 600; color: #2c3e50; }}
    .form-container input, .form-container select {{ width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 6px; font-size: 14px; }}
    .btn {{ padding: 12px 20px; background: #27ae60; color: white; border: none; border-radius: 8px; font-size: 16px; font-weight: 600; cursor: pointer; }}
    .back-link {{ display: inline-block; margin: 20px 30px; color: #3498db; text-decoration: none; font-weight: 500; }}
    .footer {{ text-align: center; padding: 20px; background: #ecf0f1; color: #7f8c8d; font-size: 13px; border-top: 1px solid #bdc3c7; }}
    </style>
    </head>
    <body>\n
    <div class="container">
    <div class="header"><h1>📥 Registrar Entrada de Material</h1></div>
    <div class="user-info"><span>👤 {session['usuario']} ({session['nivel'].upper()})</span><a href="/logout">🚪 Sair</a></div>
    <a href="/estoque" class="back-link">← Voltar ao Estoque</a>
    <div class="form-container">
    <form method="post" action="/registrar_entrada" onsubmit="return validarFormulario()">
    <div><label>Material *</label><select name="material_id" id="material_id" onchange="carregarDadosMaterial()" required><option value="">Selecione um material</option>{''.join(f'<option value="{m["id"]}" {"selected" if material and m["id"] == material["id"] else ""}>{m["denominacao"]}</option>' for m in materiais)}</select></div>
    <div class="grid-2"><div><label>Unidade de Medida (do cadastro)</label><input type="text" id="unidade_medida" readonly></div><div><label>Valor Unitário Cadastrado</label><input type="text" id="valor_unitario_cadastrado" readonly></div></div>
    <div class="grid-2"><div><label>Quantidade Comprada *</label><input type="number" name="quantidade" id="quantidade" step="0.01" required oninput="calcularValorUnitario()"></div><div><label>Tamanho (ex: 66x96 cm)</label><input type="text" name="tamanho" placeholder="Opcional"></div></div>
    <div><label>Valor Total Pago *</label><input type="number" name="valor_total" id="valor_total" step="0.01" required oninput="calcularValorUnitario()"></div>
    <div><label>Valor Unitário Calculado</label><input type="text" id="valor_unitario_calculado" readonly></div>
    <button type="submit" class="btn">➕ Registrar Entrada</button>
    </form>
    </div>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </div>
    <script>
    const materiais = {materiais_js};
    function carregarDadosMaterial() {{ const select = document.getElementById('material_id'); const id = select.value; const material = materiais.find(m => m.id == id); if (material) {{ document.getElementById('unidade_medida').value = material.unidade_medida; document.getElementById('valor_unitario_cadastrado').value = parseFloat(material.valor_unitario).toFixed(2); document.getElementById('quantidade').value = ''; document.getElementById('valor_total').value = ''; document.getElementById('valor_unitario_calculado').value = ''; }} else {{ document.getElementById('unidade_medida').value = ''; document.getElementById('valor_unitario_cadastrado').value = ''; }} }}
    function calcularValorUnitario() {{ const quantidade = parseFloat(document.getElementById('quantidade').value) || 0; const valor_total = parseFloat(document.getElementById('valor_total').value) || 0; if (quantidade > 0 && valor_total > 0) {{ const valor_calculado = (valor_total / quantidade).toFixed(2); document.getElementById('valor_unitario_calculado').value = valor_calculado; }} else {{ document.getElementById('valor_unitario_calculado').value = ''; }} }}
    function validarFormulario() {{ const quantidade = parseFloat(document.getElementById('quantidade').value); const valor_total = parseFloat(document.getElementById('valor_total').value); if (quantidade <= 0 || valor_total <= 0) {{ alert('Quantidade e valor total devem ser maiores que zero.'); return false; }} return true; }}
    window.onload = function() {{ if ('{material_id}') {{ carregarDadosMaterial(); }} }};
    </script>
    </body>
    </html>
    '''

@app.route('/registrar_entrada', methods=['POST'])
def registrar_entrada():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    if session['nivel'] not in ['administrador', 'vendedor']:
        flash("Acesso negado!")
        return redirect(url_for('estoque'))
    material_id = request.form.get('material_id')
    quantidade = request.form.get('quantidade')
    valor_total = request.form.get('valor_total')
    tamanho = request.form.get('tamanho')
    if not material_id or not quantidade or not valor_total:
        flash("Preencha todos os campos obrigatórios!")
        return redirect(url_for('estoque'))
    try:
        quantidade = float(quantidade)
        valor_total = float(valor_total)
        if quantidade <= 0 or valor_total <= 0:
            flash("Quantidade e valor total devem ser maiores que zero.")
            return redirect(url_for('estoque'))
        valor_unitario = round(valor_total / quantidade, 2)
    except:
        flash("Quantidade e valor devem ser números válidos.")
        return redirect(url_for('estoque'))
    try:
        url = f"{SUPABASE_URL}/rest/v1/estoque"
        dados = {
            "material_id": int(material_id),
            "tipo": "entrada",
            "quantidade": quantidade,
            "valor_unitario": valor_unitario,
            "valor_total": valor_total,
            "tamanho": tamanho,
            "data_movimentacao": datetime.now().isoformat(),
            "motivo": None
        }
        response = requests.post(url, json=dados, headers=headers)
        if response.status_code == 201:
            flash("✅ Entrada registrada com sucesso!")
        else:
            print("❌ Erro ao registrar entrada:", response.status_code, response.text)
            flash("❌ Erro ao registrar entrada. Verifique os dados.")
    except Exception as e:
        print("❌ Erro de conexão:", str(e))
        flash("❌ Erro ao conectar ao banco de dados.")
    return redirect(url_for('estoque'))

@app.route('/registrar_saida_form')
def registrar_saida_form():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    if session['nivel'] != 'administrador':
        flash("Acesso negado!")
        return redirect(url_for('clientes'))
    material_id = request.args.get('material_id')
    material = None
    saldo_atual = 0
    try:
        if material_id:
            url = f"{SUPABASE_URL}/rest/v1/materiais?id=eq.{material_id}"
            response = requests.get(url, headers=headers)
            if response.status_code == 200 and response.json():
                material = response.json()[0]
            saldo = calcular_estoque_atual()
            saldo_atual = saldo.get(int(material_id), 0)
    except:
        flash("Erro ao carregar material.")
        return redirect(url_for('estoque'))
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Registrar Saída</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f5f7fa; color: #333; min-height: 100vh; padding: 0; margin: 0; }}
    .container {{ max-width: 900px; margin: 30px auto; background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0, 0, 0, 0.1); overflow: hidden; }}
    .header {{ background: #2c3e50; color: white; text-align: center; padding: 30px; }}
    h1 {{ font-size: 28px; margin: 0; font-weight: 600; }}
    .user-info {{ background: #34495e; color: white; padding: 15px 20px; font-size: 15px; display: flex; justify-content: space-between; align-items: center; }}
    .form-container {{ padding: 30px; }}
    .form-container label {{ display: block; margin: 10px 0 5px 0; font-weight: 600; color: #2c3e50; }}
    .form-container input, .form-container select, .form-container textarea {{ width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 6px; font-size: 14px; }}
    .btn {{ padding: 12px 20px; background: #e74c3c; color: white; border: none; border-radius: 8px; font-size: 16px; font-weight: 600; cursor: pointer; }}
    .back-link {{ display: inline-block; margin: 20px 30px; color: #3498db; text-decoration: none; font-weight: 500; }}
    .footer {{ text-align: center; padding: 20px; background: #ecf0f1; color: #7f8c8d; font-size: 13px; border-top: 1px solid #bdc3c7; }}
    .alert {{ background: #fdf3cd; color: #856404; padding: 15px; border-radius: 8px; margin: 15px 0; font-size: 14px; }}
    </style>
    </head>
    <body>\n
    <div class="container">
    <div class="header"><h1>📤 Registrar Saída de Material</h1></div>
    <div class="user-info"><span>👤 {session['usuario']} ({session['nivel'].upper()})</span><a href="/logout">🚪 Sair</a></div>
    <a href="/estoque" class="back-link">← Voltar ao Estoque</a>
    <div class="form-container">
    <form method="post" action="/registrar_saida" onsubmit="return validarSaida()">
    <input type="hidden" name="material_id" value="{material['id']}">
    <div><label>Material</label><input type="text" value="{material['denominacao']}" readonly></div>
    <div><label>Unidade de Medida</label><input type="text" value="{material['unidade_medida']}" readonly></div>
    <div><label>Saldo Atual em Estoque</label><input type="text" id="saldo_atual" value="{saldo_atual}" readonly style="font-weight: bold;"></div>
    <div><label>Quantidade a Retirar *</label><input type="number" name="quantidade" id="quantidade" step="0.01" required oninput="verificarSaldo()"></div>
    <div><label>Motivo da Saída *</label><textarea name="motivo" rows="3" required></textarea></div>
    <div id="alerta_saldo" class="alert" style="display: none;">⚠️ A quantidade retirada é maior que o saldo em estoque!</div>
    <button type="submit" class="btn">📤 Registrar Saída</button>
    </form>
    </div>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </div>
    <script>
    function verificarSaldo() {{ const saldo = parseFloat(document.getElementById('saldo_atual').value); const qtd = parseFloat(document.getElementById('quantidade').value) || 0; const alerta = document.getElementById('alerta_saldo'); if (qtd > saldo) {{ alerta.style.display = 'block'; }} else {{ alerta.style.display = 'none'; }} }}
    function validarSaida() {{ const saldo = parseFloat(document.getElementById('saldo_atual').value); const qtd = parseFloat(document.getElementById('quantidade').value) || 0; if (qtd <= 0) {{ alert('A quantidade deve ser maior que zero.'); return false; }} if (qtd > saldo) {{ if (!confirm('⚠️ A quantidade é maior que o saldo. Deseja continuar mesmo assim?')) {{ return false; }} }} return true; }}
    </script>
    </body>
    </html>
    '''

@app.route('/registrar_saida', methods=['POST'])
def registrar_saida():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    if session['nivel'] not in ['administrador', 'vendedor']:
        flash("Acesso negado!")
        return redirect(url_for('estoque'))
    material_id = request.form.get('material_id')
    quantidade = request.form.get('quantidade')
    motivo = request.form.get('motivo')
    if not material_id or not quantidade or not motivo:
        flash("Preencha todos os campos!")
        return redirect(url_for('estoque'))
    try:
        quantidade = float(quantidade)
        if quantidade <= 0:
            flash("Quantidade deve ser maior que zero.")
            return redirect(url_for('estoque'))
    except:
        flash("Quantidade inválida.")
        return redirect(url_for('estoque'))
    try:
        url = f"{SUPABASE_URL}/rest/v1/estoque"
        dados = {
            "material_id": int(material_id),
            "tipo": "saida",
            "quantidade": quantidade,
            "motivo": motivo,
            "data_movimentacao": datetime.now().isoformat()
        }
        response = requests.post(url, json=dados, headers=headers)
        if response.status_code == 201:
            flash("📤 Saída registrada com sucesso!")
        else:
            flash("❌ Erro ao registrar saída.")
    except Exception as e:
        flash("❌ Erro ao registrar saída.")
    return redirect(url_for('estoque'))

@app.route('/excluir_movimentacao/<int:id>')
def excluir_movimentacao(id):
    if 'usuario' not in session or session['nivel'] != 'administrador':
        flash("Acesso negado!")
        return redirect(url_for('estoque'))
    if excluir_movimentacao_db(id):
        flash("🗑️ Movimentação excluída com sucesso!")
    else:
        flash("❌ Erro ao excluir movimentação.")
    return redirect(url_for('estoque'))

@app.route('/fornecedores')
def listar_fornecedores():
    if 'usuario' not in session or session['nivel'] != 'administrador':
        flash("Acesso negado!")
        return redirect(url_for('clientes'))
    busca = request.args.get('q', '').strip()
    try:
        if busca:
            url = f"{SUPABASE_URL}/rest/v1/fornecedores?or=(nome.ilike.*{busca}*,cnpj.ilike.*{busca}*)"
        else:
            url = f"{SUPABASE_URL}/rest/v1/fornecedores?select=*"
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            fornecedores = response.json()
        else:
            flash("Erro ao carregar fornecedores.")
            fornecedores = []
    except Exception as e:
        flash("Erro de conexão.")
        fornecedores = []
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Fornecedores Cadastrados</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f5f7fa; color: #333; min-height: 100vh; padding: 0; margin: 0; }}
    .container {{ max-width: 1100px; margin: 30px auto; background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0, 0, 0, 0.1); overflow: hidden; }}
    .header {{ background: #2c3e50; color: white; text-align: center; padding: 30px; }}
    h1 {{ font-size: 28px; margin: 0; font-weight: 600; }}
    .user-info {{ background: #34495e; color: white; padding: 15px 20px; font-size: 15px; display: flex; justify-content: space-between; align-items: center; }}
    .search-box {{ padding: 20px 30px; text-align: center; }}
    .search-box input {{ width: 70%; padding: 12px; border: 1px solid #ddd; border-radius: 8px; font-size: 16px; }}
    table {{ width: 100%; border-collapse: collapse; }}
    th, td {{ padding: 16px 20px; text-align: left; }}
    th {{ background: #ecf0f1; color: #2c3e50; font-weight: 600; text-transform: uppercase; font-size: 14px; }}
    tr:nth-child(even) {{ background: #f9f9f9; }}
    tr:hover {{ background: #f1f7fb; }}
    .back-link {{ display: inline-block; margin: 20px 30px; color: #3498db; text-decoration: none; font-weight: 500; }}
    .footer {{ text-align: center; padding: 20px; background: #ecf0f1; color: #7f8c8d; font-size: 13px; border-top: 1px solid #bdc3c7; }}
    </style>
    </head>
    <body>\n
    <div class="container">
    <div class="header"><h1>📋 Fornecedores Cadastrados</h1></div>
    <div class="user-info"><span>👤 {session['usuario']} ({session['nivel'].upper()})</span><a href="/logout">🚪 Sair</a></div>
    {MENU_FLUTUANTE}
    <a href="/cadastrar_fornecedor" class="btn" style="padding: 12px 20px; background: #27ae60; color: white; border-radius: 8px; text-decoration: none; font-weight: 600; margin: 0 30px;">➕ Cadastrar Novo Fornecedor</a>
    <div class="search-box"><form method="get" style="display: inline;"><input type="text" name="q" placeholder="Pesquisar por nome ou CNPJ..." value="{busca}"><button type="submit" style="padding: 10px 20px; background: #3498db; color: white; border: none; border-radius: 8px; cursor: pointer;">🔍 Pesquisar</button></form></div>
    <table><thead><tr><th>ID</th><th>Nome</th><th>CNPJ</th><th>Contato</th><th>Telefone</th><th>E-mail</th><th>Ações</th></tr></thead><tbody>{''.join(f"""
    <tr>
    <td>{f["id"]}</td><td>{f["nome"]}</td><td>{f["cnpj"]}</td><td>{f.get("contato", "—")}</td><td>{f.get("telefone", "—")}</td><td>{f.get("email", "—")}</td>
    <td><div style="display: flex; gap: 10px;"><a href="/editar_fornecedor/{f["id"]}" style="color: #f39c12; text-decoration: none;">✏️ Editar</a><a href="/excluir_fornecedor/{f["id"]}" style="color: #e74c3c; text-decoration: none;" onclick="return confirm('Tem certeza que deseja excluir?')">🗑️ Excluir</a></div></td>
    </tr>
    """ for f in fornecedores)}</tbody></table>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </div>
    </body>
    </html>
    '''

@app.route('/cadastrar_fornecedor', methods=['GET', 'POST'])
def cadastrar_fornecedor():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    if session['nivel'] != 'administrador':
        flash("Acesso negado!")
        return redirect(url_for('clientes'))
    if request.method == 'POST':
        nome = request.form.get('nome')
        cnpj = request.form.get('cnpj')
        contato = request.form.get('contato')
        telefone = request.form.get('telefone')
        email = request.form.get('email')
        endereco = request.form.get('endereco')
        if not nome:
            flash("Nome do fornecedor é obrigatório!")
            return redirect(request.url)
        if criar_fornecedor(nome, cnpj, contato, telefone, email, endereco):
            flash("✅ Fornecedor cadastrado com sucesso!")
            return redirect(url_for('listar_fornecedores'))
        else:
            flash("❌ Erro ao cadastrar fornecedor.")
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Cadastrar Fornecedor</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: #333; min-height: 100vh; padding: 0; margin: 0; }}
    .container {{ max-width: 800px; margin: 30px auto; background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0,0,0,0.1); overflow: hidden; }}
    .header {{ background: #2c3e50; color: white; text-align: center; padding: 30px; }}
    h1 {{ font-size: 28px; margin: 0; font-weight: 600; }}
    .user-info {{ background: #34495e; color: white; padding: 15px 20px; font-size: 15px; display: flex; justify-content: space-between; align-items: center; }}
    .form-container {{ padding: 30px; }}
    .form-container label {{ display: block; margin: 10px 0 5px 0; font-weight: 600; color: #2c3e50; }}
    .form-container input, .form-container textarea {{ width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 6px; font-size: 14px; }}
    .btn {{ padding: 12px 20px; background: #27ae60; color: white; border: none; border-radius: 8px; font-size: 16px; font-weight: 600; cursor: pointer; }}
    .back-link {{ display: inline-block; margin: 20px 30px; color: #3498db; text-decoration: none; font-weight: 500; }}
    .footer {{ text-align: center; padding: 20px; background: #ecf0f1; color: #7f8c8d; font-size: 13px; border-top: 1px solid #bdc3c7; }}
    </style>
    </head>
    <body>\n
    <div class="container">
    <div class="header"><h1>➕ Cadastrar Novo Fornecedor</h1></div>
    <div class="user-info"><span>👤 {session['usuario']} ({session['nivel'].upper()})</span><a href="/logout">🚪 Sair</a></div>
    <a href="/fornecedores" class="back-link">← Voltar à lista</a>
    <form method="post" class="form-container">
    <div><label>Nome *</label><input type="text" name="nome" required></div>
    <div><label>CNPJ</label><input type="text" name="cnpj"></div>
    <div><label>Contato</label><input type="text" name="contato"></div>
    <div><label>Telefone</label><input type="text" name="telefone"></div>
    <div><label>E-mail</label><input type="email" name="email"></div>
    <div><label>Endereço</label><textarea name="endereco" rows="3"></textarea></div>
    <button type="submit" class="btn">💾 Salvar Fornecedor</button>
    </form>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </div>
    </body>
    </html>
    '''

@app.route('/editar_fornecedor/<int:id>', methods=['GET', 'POST'])
def editar_fornecedor(id):
    if 'usuario' not in session:
        return redirect(url_for('login'))
    if session['nivel'] != 'administrador':
        flash("Acesso negado!")
        return redirect(url_for('clientes'))
    try:
        url = f"{SUPABASE_URL}/rest/v1/fornecedores?id=eq.{id}"
        response = requests.get(url, headers=headers)
        if response.status_code != 200 or not response.json():
            flash("Fornecedor não encontrado.")
            return redirect(url_for('listar_fornecedores'))
        fornecedor = response.json()[0]
    except Exception as e:
        flash("Erro ao carregar fornecedor.")
        return redirect(url_for('listar_fornecedores'))
    if request.method == 'POST':
        nome = request.form.get('nome')
        cnpj = request.form.get('cnpj')
        contato = request.form.get('contato')
        telefone = request.form.get('telefone')
        email = request.form.get('email')
        endereco = request.form.get('endereco')
        if not nome:
            flash("Nome do fornecedor é obrigatório!")
            return redirect(request.url)
        if atualizar_fornecedor(id, nome, cnpj, contato, telefone, email, endereco):
            flash("✅ Fornecedor atualizado com sucesso!")
            return redirect(url_for('listar_fornecedores'))
        else:
            flash("❌ Erro ao atualizar fornecedor.")
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Editar Fornecedor</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f5f7fa; color: #333; min-height: 100vh; padding: 0; margin: 0; }}
    .container {{ max-width: 800px; margin: 30px auto; background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0,0,0,0.1); overflow: hidden; }}
    .header {{ background: #2c3e50; color: white; text-align: center; padding: 30px; }}
    h1 {{ font-size: 28px; margin: 0; font-weight: 600; }}
    .user-info {{ background: #34495e; color: white; padding: 15px 20px; font-size: 15px; display: flex; justify-content: space-between; align-items: center; }}
    .form-container {{ padding: 30px; }}
    .form-container label {{ display: block; margin: 10px 0 5px 0; font-weight: 600; color: #2c3e50; }}
    .form-container input, .form-container textarea {{ width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 6px; font-size: 14px; }}
    .btn {{ padding: 12px 20px; background: #f39c12; color: white; border: none; border-radius: 8px; font-size: 16px; font-weight: 600; cursor: pointer; }}
    .back-link {{ display: inline-block; margin: 20px 30px; color: #3498db; text-decoration: none; font-weight: 500; }}
    .footer {{ text-align: center; padding: 20px; background: #ecf0f1; color: #7f8c8d; font-size: 13px; border-top: 1px solid #bdc3c7; }}
    </style>
    </head>
    <body>\n
    <div class="container">
    <div class="header"><h1>✏️ Editar Fornecedor: {fornecedor['nome']}</h1></div>
    <div class="user-info"><span>👤 {session['usuario']} ({session['nivel'].upper()})</span><a href="/logout">🚪 Sair</a></div>
    <a href="/fornecedores" class="back-link">← Voltar à lista</a>
    <form method="post" class="form-container">
    <div><label>Nome *</label><input type="text" name="nome" value="{fornecedor['nome']}" required></div>
    <div><label>CNPJ</label><input type="text" name="cnpj" value="{fornecedor.get('cnpj', '')}"></div>
    <div><label>Contato</label><input type="text" name="contato" value="{fornecedor.get('contato', '')}"></div>
    <div><label>Telefone</label><input type="text" name="telefone" value="{fornecedor.get('telefone', '')}"></div>
    <div><label>E-mail</label><input type="email" name="email" value="{fornecedor.get('email', '')}"></div>
    <div><label>Endereço</label><textarea name="endereco" rows="3">{fornecedor.get('endereco', '')}</textarea></div>
    <button type="submit" class="btn">💾 Salvar Alterações</button>
    </form>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </div>
    </body>
    </html>
    '''

@app.route('/excluir_fornecedor/<int:id>')
def excluir_fornecedor_view(id):
    if 'usuario' not in session:
        flash("Acesso negado!")
        return redirect(url_for('login'))
    if session['nivel'] != 'administrador':
        flash("Acesso negado!")
        return redirect(url_for('clientes'))
    if excluir_fornecedor(id):
        flash("🗑️ Fornecedor excluído com sucesso!")
    else:
        flash("❌ Erro ao excluir fornecedor.")
    return redirect(url_for('listar_fornecedores'))

# ========================
# ROTAS DE ORÇAMENTOS
# ========================
@app.route('/orcamentos')
def listar_orcamentos():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    busca = request.args.get('q', '').strip()
    try:
        url = f"{SUPABASE_URL}/rest/v1/servicos?select=*,empresas(nome_empresa)&tipo=eq.Orçamento&order=codigo_servico.desc"
        if busca:
            url += f"&titulo=ilike.*{busca}*"
        response = requests.get(url, headers=headers)
        orcamentos = response.json() if response.status_code == 200 else []
    except:
        orcamentos = []
    
    rows = ""
    for o in orcamentos:
        cliente = o.get('empresas', {}).get('nome_empresa', '—') if o.get('empresas') else '—'
        rows += f'''
        <tr>
            <td>{o.get('codigo_servico', '—')}</td>
            <td>{o.get('titulo', '—')}</td>
            <td>{cliente}</td>
            <td>R$ {float(o.get('valor_cobrado', 0) or 0):.2f}</td>
            <td>{format_data(o.get('data_abertura'))}</td>
            <td>
                <a href="/pdf_orcamento/{o['id']}" style="padding:6px 12px;background:#e67e22;color:white;text-decoration:none;border-radius:4px;margin-right:5px;">📄 PDF</a>
                <a href="/complementar_orcamento/{o['id']}" style="padding:6px 12px;background:#27ae60;color:white;text-decoration:none;border-radius:4px;margin-right:5px;">✅ Aceito</a>
                <a href="/editar_orcamento/{o['id']}" style="padding:6px 12px;background:#f39c12;color:white;text-decoration:none;border-radius:4px;margin-right:5px;">✏️ Editar</a>
                <a href="/excluir_servico/{o['id']}" onclick="return confirm('Tem certeza?')" style="padding:6px 12px;background:#e74c3c;color:white;text-decoration:none;border-radius:4px;">🗑️</a>
            </td>
        </tr>'''
    
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Orçamentos</title>
    <style>
    body {{ font-family: Arial, sans-serif; background: #f5f7fa; margin: 0; padding: 20px; }}
    .container {{ max-width: 1200px; margin: 0 auto; background: white; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
    .header {{ background: #2c3e50; color: white; padding: 20px; text-align: center; border-radius: 10px 10px 0 0; }}
    .user-bar {{ background: #34495e; color: white; padding: 10px 20px; display: flex; justify-content: space-between; }}
    .content {{ padding: 30px; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
    th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #ecf0f1; }}
    th {{ background: #f8f9fa; font-weight: bold; color: #2c3e50; }}
    .btn {{ padding: 10px 20px; background: #27ae60; color: white; text-decoration: none; border-radius: 5px; display: inline-block; margin: 10px 0; }}
    .search {{ padding: 10px; width: 300px; border: 1px solid #ddd; border-radius: 5px; }}
    .menu-container {{ position: relative; z-index: 9999; margin: 20px 0; }}
    </style>
    </head>
    <body>
    <div class="container">
        <div class="header"><h1 style="margin:0;">💰 Orçamentos</h1></div>
        <div class="user-bar">
            <span>👤 {session['usuario']} ({session['nivel'].upper()})</span>
            <a href="/logout" style="color: #3498db;">🚪 Sair</a>
        </div>
        <div class="content">
            <div class="menu-container">{MENU_FLUTUANTE}</div>
            <a href="/adicionar_orcamento" class="btn">➕ Novo Orçamento</a>
            <div style="margin: 20px 0;">
                <form method="get" style="display: inline-block;">
                    <input type="text" name="q" class="search" placeholder="Pesquisar por título..." value="{busca}">
                    <button type="submit" style="padding: 10px 20px; background: #3498db; color: white; border: none; border-radius: 5px; cursor: pointer;">🔍 Pesquisar</button>
                </form>
            </div>
            <table>
                <thead><tr><th>Código</th><th>Título</th><th>Cliente</th><th>Valor</th><th>Data</th><th>Ações</th></tr></thead>
                <tbody>{rows if rows else '<tr><td colspan="6" style="text-align: center; color: #95a5a6;">Nenhum orçamento encontrado</td></tr>'}</tbody>
            </table>
        </div>
    </div>
    </body>
    </html>
    '''

def adicionar_dias_uteis(data_inicio, dias):
    """Adiciona dias úteis pulando finais de semana e feriados"""
    from datetime import timedelta
    
    # Lista de feriados nacionais (você pode adicionar mais)
    feriados = [
        "2026-01-01",  # Confraternização Universal
        "2026-02-17",  # Carnaval
        "2026-04-03",  # Sexta-feira Santa
        "2026-04-21",  # Tiradentes
        "2026-05-01",  # Dia do Trabalho
        "2026-06-04",  # Corpus Christi
        "2026-09-07",  # Independência
        "2026-10-12",  # Nossa Senhora
        "2026-11-02",  # Finados
        "2026-11-15",  # Proclamação República
        "2026-11-20",  # Consciência Negra
        "2026-12-25",  # Natal
    ]
    
    data_atual = data_inicio
    dias_adicionados = 0
    
    while dias_adicionados < dias:
        data_atual += timedelta(days=1)
        # Verifica se é sábado (5) ou domingo (6)
        if data_atual.weekday() < 5:
            # Verifica se não é feriado
            if data_atual.strftime("%Y-%m-%d") not in feriados:
                dias_adicionados += 1
    
    return data_atual

@app.route('/adicionar_orcamento', methods=['GET', 'POST'])
def adicionar_orcamento():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        empresa_id = request.form.get('empresa_id')
        data_abertura = request.form.get('data_abertura')
        prazo_dias = request.form.get('prazo_dias', '7')
        observacoes = request.form.get('observacoes_gerais', '')
        
        if not empresa_id:
            flash("❌ Cliente é obrigatório!")
            return redirect(url_for('adicionar_orcamento'))
        
        # Gera código OR-XXX
        try:
            ult = requests.get(f"{SUPABASE_URL}/rest/v1/servicos?select=codigo_servico&order=codigo_servico.desc&limit=1", headers=headers).json()
            n = int(ult[0]['codigo_servico'].split('-')[1]) + 1 if ult and len(ult) > 0 else 1
            cod = f"OR-{n:03d}"
        except Exception as e:
            print(f"Erro ao gerar código: {e}")
            cod = f"OR-001"
        
        try:
            # Cria o orçamento
            json_data = {
                "codigo_servico": cod,
                "titulo": "Orçamento Múltiplo",
                "empresa_id": int(empresa_id),
                "tipo": "Orçamento",
                "status": "Pendente",
                "data_abertura": data_abertura,
                "previsao_entrega": None,
                "valor_cobrado": 0.0,
                "observacoes": f"Prazo: {prazo_dias} dias úteis após aprovação da arte. {observacoes}"
            }
            
            resp = requests.post(f"{SUPABASE_URL}/rest/v1/servicos", json=json_data, headers=headers)
            
            if resp.status_code in [200, 201]:
                try:
                    oid = resp.json().get('id')
                except:
                    busca = requests.get(f"{SUPABASE_URL}/rest/v1/servicos?select=id&codigo_servico=eq.{cod}&order=id.desc&limit=1", headers=headers)
                    if busca.status_code == 200 and busca.json():
                        oid = busca.json()[0].get('id')
                    else:
                        flash("❌ Erro ao obter ID do orçamento.")
                        return redirect(url_for('listar_orcamentos'))
                
                if not oid:
                    flash("❌ Erro ao criar orçamento.")
                    return redirect(url_for('adicionar_orcamento'))
                
                # Processa os itens
                vt = 0.0
                titulos = request.form.getlist('item_titulo[]')
                quantidades = request.form.getlist('item_quantidade[]')
                valores_unit = request.form.getlist('item_valor_unit[]')
                dimensoes = request.form.getlist('item_dimensao[]')
                cores = request.form.getlist('item_cores[]')
                
                for i in range(len(titulos)):
                    t = titulos[i].strip()
                    if not t:
                        continue
                    
                    try:
                        q_str = quantidades[i].strip().replace(',', '.') if i < len(quantidades) else '1'
                        vu_str = valores_unit[i].strip().replace(',', '.') if i < len(valores_unit) else '0'
                        
                        q = float(q_str) if q_str else 1.0
                        vu = float(vu_str) if vu_str else 0.0
                    except Exception as e:
                        print(f"Erro ao converter: {e}")
                        continue
                    
                    dim = dimensoes[i].strip() if i < len(dimensoes) else ''
                    cor = cores[i].strip() if i < len(cores) else ''
                    total_item = q * vu
                    vt += total_item
                    
                    requests.post(f"{SUPABASE_URL}/rest/v1/itens_orcamento", json={
                        "orcamento_id": oid,
                        "titulo": t,
                        "quantidade": q,
                        "dimensao": dim,
                        "numero_cores": cor,
                        "valor_unitario": vu,
                        "valor_total": total_item
                    }, headers=headers)
                
                # Atualiza valor total
                requests.patch(f"{SUPABASE_URL}/rest/v1/servicos?id=eq.{oid}", 
                    json={"valor_cobrado": vt}, headers=headers)
                
                flash("✅ Orçamento criado com sucesso!")
                return redirect(url_for('listar_orcamentos'))
            else:
                flash(f"❌ Erro ao criar orçamento. Status: {resp.status_code}")
                return redirect(url_for('adicionar_orcamento'))
                
        except Exception as e:
            print(f"ERRO GERAL: {str(e)}")
            import traceback
            traceback.print_exc()
            flash("❌ Erro ao criar orçamento.")
    
    # GET - Renderiza o formulário
    try:
        emps = requests.get(f"{SUPABASE_URL}/rest/v1/empresas?select=id,nome_empresa&order=nome_empresa.asc", headers=headers).json() or []
    except:
        emps = []
    
    opts = "".join([f'<option value="{e["id"]}">{e["nome_empresa"]}</option>' for e in emps])
    
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <title>Novo Orçamento</title>
    <style>
        body {{ font-family: Arial, sans-serif; background: #f5f7fa; padding: 20px; }}
        .container {{ max-width: 1100px; margin: 0 auto; background: white; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        .header {{ background: #2c3e50; color: white; padding: 25px; text-align: center; border-radius: 10px 10px 0 0; }}
        .content {{ padding: 30px; }}
        .grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
        .form-group {{ margin-bottom: 15px; }}
        label {{ display: block; margin-bottom: 5px; font-weight: bold; color: #2c3e50; }}
        input, select, textarea {{ width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 5px; box-sizing: border-box; }}
        .item-row {{ background: #f8f9fa; padding: 15px; margin: 10px 0; border-radius: 8px; border-left: 4px solid #3498db; }}
        .btn {{ padding: 12px 20px; background: #27ae60; color: white; border: none; border-radius: 5px; cursor: pointer; font-weight: bold; font-size: 16px; }}
        .btn-blue {{ background: #3498db; }}
        .back-link {{ color: #3498db; text-decoration: none; display: inline-block; margin-bottom: 15px; }}
    </style>
    </head>
    <body>
    <div class="container">
        <div class="header"><h1>➕ Novo Orçamento</h1></div>
        <div class="content">
            <a href="/orcamentos" class="back-link">← Voltar</a>
            <form method="post" id="formOrcamento">
                <div class="grid">
                    <div class="form-group">
                        <label>Cliente *</label>
                        <select name="empresa_id" required>
                            <option value="">Selecione uma empresa</option>
                            {opts}
                        </select>
                    </div>
                    <div class="form-group">
                        <label>Data de Abertura *</label>
                        <input type="date" name="data_abertura" value="{datetime.now().strftime('%Y-%m-%d')}" required>
                    </div>
                </div>
                <div class="form-group">
                    <label>Prazo de Entrega (dias úteis após aprovação da arte)</label>
                    <input type="number" name="prazo_dias" value="7" min="1">
                    <small style="color: #7f8c8d;">* A contagem inicia após aprovação da arte</small>
                </div>
                <div class="form-group">
                    <label>Observações Gerais</label>
                    <textarea name="observacoes_gerais" rows="2"></textarea>
                </div>
                <h3 style="color: #2c3e50; margin: 25px 0 15px 0;">Itens do Orçamento</h3>
                <div id="itens-container">
                    <div class="item-row">
                        <div class="grid" style="grid-template-columns: 2fr 1fr 1fr 1fr 1fr auto; gap: 10px;">
                            <div>
                                <label>Material/Descrição *</label>
                                <input type="text" name="item_titulo[]" placeholder="Ex: Banner Lona" required>
                            </div>
                            <div>
                                <label>Quantidade *</label>
                                <input type="number" name="item_quantidade[]" step="any" value="1" required>
                            </div>
                            <div>
                                <label>Valor Unit. (R$) *</label>
                                <input type="text" name="item_valor_unit[]" class="valor-input" placeholder="0,00" required>
                            </div>
                            <div>
                                <label>Dimensão (opcional)</label>
                                <input type="text" name="item_dimensao[]" placeholder="Ex: 100x50">
                            </div>
                            <div>
                                <label>Cores</label>
                                <input type="number" name="item_cores[]" step="1" value="4">
                            </div>
                            <div>
                                <button type="button" onclick="this.closest('.item-row').remove()" style="background:#e74c3c;color:white;border:none;padding:8px;border-radius:5px;margin-top:28px;cursor:pointer;">🗑️</button>
                            </div>
                        </div>
                    </div>
                </div>
                <button type="button" onclick="addRow()" class="btn btn-blue" style="margin: 15px 0; width: 100%;">+ Adicionar Item</button>
                <button type="submit" class="btn" style="width: 100%; margin-top: 15px;">💾 Gerar Orçamento</button>
            </form>
        </div>
    </div>
    
    <script>
    // Formatação CORRETA de valor monetário - SEM ZEROS À ESQUERDA
    document.addEventListener('input', function(e) {{
        if (e.target.classList.contains('valor-input')) {{
            let value = e.target.value.replace(/\\D/g, '');
            value = (parseInt(value || '0') / 100).toFixed(2).replace('.', ',');
            e.target.value = value;
        }}
    }});
    
    function addRow() {{
        const div = document.createElement('div');
        div.className = 'item-row';
        div.innerHTML = `
            <div class="grid" style="grid-template-columns: 2fr 1fr 1fr 1fr 1fr auto; gap: 10px;">
                <div><label>Material/Descrição *</label><input type="text" name="item_titulo[]" placeholder="Ex: Banner Lona" required></div>
                <div><label>Quantidade *</label><input type="number" name="item_quantidade[]" step="any" value="1" required></div>
                <div><label>Valor Unit. (R$) *</label><input type="text" name="item_valor_unit[]" class="valor-input" placeholder="0,00" required></div>
                <div><label>Dimensão (opcional)</label><input type="text" name="item_dimensao[]" placeholder="Ex: 100x50"></div>
                <div><label>Cores</label><input type="number" name="item_cores[]" step="1" value="4"></div>
                <div><button type="button" onclick="this.closest('.item-row').remove()" style="background:#e74c3c;color:white;border:none;padding:8px;border-radius:5px;margin-top:28px;cursor:pointer;">🗑️</button></div>
            </div>
        `;
        document.getElementById('itens-container').appendChild(div);
    }}
    </script>
    </body>
    </html>
    '''

@app.route('/editar_orcamento/<int:id>', methods=['GET', 'POST'])
def editar_orcamento(id):
    if 'usuario' not in session:
        return redirect(url_for('login'))
    
    try:
        url = f"{SUPABASE_URL}/rest/v1/servicos?id=eq.{id}&select=*,empresas(nome_empresa),itens_orcamento(*)"
        resp = requests.get(url, headers=headers)
        orc = resp.json()[0] if resp.json() else None
        
        if not orc:
            flash("Orçamento não encontrado!")
            return redirect(url_for('listar_orcamentos'))
        
        empresas = requests.get(f"{SUPABASE_URL}/rest/v1/empresas?select=id,nome_empresa&order=nome_empresa.asc", headers=headers).json() or []
        
        if request.method == 'POST':
            # Atualiza dados do orçamento
            dados = {
                "empresa_id": int(request.form['empresa_id']),
                "data_abertura": request.form['data_abertura'],
                "observacoes": request.form.get('observacoes_gerais', '')
            }
            requests.patch(f"{SUPABASE_URL}/rest/v1/servicos?id=eq.{id}", json=dados, headers=headers)
            
            # Remove itens antigos
            requests.delete(f"{SUPABASE_URL}/rest/v1/itens_orcamento?orcamento_id=eq.{id}", headers=headers)
            
            # Processa novos itens
            vt = 0.0
            titulos = request.form.getlist('item_titulo[]')
            quantidades = request.form.getlist('item_quantidade[]')
            valores_unit = request.form.getlist('item_valor_unit[]')
            dimensoes = request.form.getlist('item_dimensao[]')
            cores = request.form.getlist('item_cores[]')
            
            for i in range(len(titulos)):
                t = titulos[i].strip()
                if not t:
                    continue
                
                # Parsing ROBUSTO de quantidade e valor (aceita vírgula OU ponto)
                try:
                    q_str = quantidades[i].strip().replace(',', '.') if i < len(quantidades) else '1'
                    vu_str = valores_unit[i].strip().replace(',', '.') if i < len(valores_unit) else '0'
                    
                    q = float(q_str) if q_str else 1.0
                    vu = float(vu_str) if vu_str else 0.0
                except Exception as e:
                    print(f"Erro ao converter item {i+1}: {e}")
                    continue
                
                dim = dimensoes[i].strip() if i < len(dimensoes) else ''
                cor = cores[i].strip() if i < len(cores) else ''
                total_item = q * vu
                vt += total_item
                
                requests.post(f"{SUPABASE_URL}/rest/v1/itens_orcamento", 
                    json={
                        "orcamento_id": id,
                        "titulo": t,
                        "quantidade": q,
                        "dimensao": dim,
                        "numero_cores": cor,
                        "valor_unitario": vu,
                        "valor_total": total_item
                    }, headers=headers)
            
            # Atualiza valor total
            requests.patch(f"{SUPABASE_URL}/rest/v1/servicos?id=eq.{id}", json={"valor_cobrado": vt}, headers=headers)
            flash("✅ Orçamento atualizado!")
            return redirect(url_for('listar_orcamentos'))
        
        # GET - Renderiza formulário com dados existentes
        opts_empresas = "".join([f'<option value="{e["id"]}" {"selected" if e["id"]==orc["empresa_id"] else ""}>{e["nome_empresa"]}</option>' for e in empresas])
        
        itens_html = ""
        for item in orc.get('itens_orcamento', []):
            # Formata valor para exibição: 25.0 → "25,00"
            vu_display = f"{item.get('valor_unitario', 0):.2f}".replace('.', ',')
            itens_html += f'''
            <div class="item-row">
                <div class="grid" style="grid-template-columns: 2fr 1fr 1fr 1fr 1fr auto; gap: 10px;">
                    <div><label>Material/Descrição *</label><input type="text" name="item_titulo[]" value="{item.get('titulo','')}" required></div>
                    <div><label>Quantidade *</label><input type="number" name="item_quantidade[]" value="{item.get('quantidade',1)}" required></div>
                    <div><label>Valor Unit. (R$) *</label><input type="text" name="item_valor_unit[]" class="valor-input" value="{vu_display}" required></div>
                    <div><label>Dimensão (opcional)</label><input type="text" name="item_dimensao[]" value="{item.get('dimensao','')}"></div>
                    <div><label>Cores</label><input type="number" name="item_cores[]" value="{item.get('numero_cores','')}"></div>
                    <div><button type="button" onclick="this.closest('.item-row').remove()" style="background:#e74c3c;color:white;border:none;padding:8px;border-radius:5px;margin-top:28px;cursor:pointer;">🗑️</button></div>
                </div>
            </div>'''
        
        return f'''
        <!DOCTYPE html>
        <html lang="pt-BR">
        <head>
        <meta charset="UTF-8">
        <title>Editar Orçamento</title>
        <style>
        body {{ font-family: Arial, sans-serif; background: #f5f7fa; padding: 20px; }}
        .container {{ max-width: 1100px; margin: 0 auto; background: white; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        .header {{ background: #2c3e50; color: white; padding: 25px; text-align: center; border-radius: 10px 10px 0 0; }}
        .content {{ padding: 30px; }}
        .grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
        .form-group {{ margin-bottom: 15px; }}
        label {{ display: block; margin-bottom: 5px; font-weight: bold; color: #2c3e50; }}
        input, select, textarea {{ width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 5px; box-sizing: border-box; }}
        .item-row {{ background: #f8f9fa; padding: 15px; margin: 10px 0; border-radius: 8px; border-left: 4px solid #f39c12; }}
        .btn {{ padding: 12px 20px; background: #27ae60; color: white; border: none; border-radius: 5px; cursor: pointer; font-weight: bold; font-size: 16px; }}
        .btn-blue {{ background: #3498db; }}
        .back-link {{ color: #3498db; text-decoration: none; display: inline-block; margin-bottom: 15px; }}
        </style>
        </head>
        <body>
        <div class="container">
            <div class="header"><h1>✏️ Editar Orçamento {orc.get('codigo_servico','')}</h1></div>
            <div class="content">
                <a href="/orcamentos" class="back-link">← Voltar</a>
                <form method="post">
                    <div class="grid">
                        <div class="form-group">
                            <label>Cliente</label>
                            <select name="empresa_id" required>
                                {opts_empresas}
                            </select>
                        </div>
                        <div class="form-group">
                            <label>Data de Abertura</label>
                            <input type="date" name="data_abertura" value="{orc.get('data_abertura','')[:10] if orc.get('data_abertura') else ''}" required>
                        </div>
                    </div>
                    <div class="form-group">
                        <label>Observações</label>
                        <textarea name="observacoes_gerais" rows="2">{orc.get('observacoes','')}</textarea>
                    </div>
                    <h3 style="color: #2c3e50; margin: 25px 0 15px 0;">Itens do Orçamento</h3>
                    <div id="itens-container">{itens_html}</div>
                    <button type="button" onclick="addRow()" class="btn btn-blue" style="margin: 15px 0; width: 100%;">+ Adicionar Item</button>
                    <button type="submit" class="btn" style="width: 100%; margin-top: 15px;">💾 Salvar Alterações</button>
                </form>
            </div>
        </div>
        <script>
        // Formatação de valor - funciona tanto para digitar quanto para valor pré-preenchido
        document.addEventListener('input', function(e) {{
            if (e.target.classList.contains('valor-input')) {{
                let value = e.target.value.replace(/\\D/g, '');
                value = (parseInt(value || '0') / 100).toFixed(2).replace('.', ',');
                e.target.value = value;
            }}
        }});
        
        function addRow() {{
            const div = document.createElement('div');
            div.className = 'item-row';
            div.innerHTML = `
                <div class="grid" style="grid-template-columns: 2fr 1fr 1fr 1fr 1fr auto; gap: 10px;">
                    <div><label>Material/Descrição *</label><input type="text" name="item_titulo[]" placeholder="Ex: Banner Lona" required></div>
                    <div><label>Quantidade *</label><input type="number" name="item_quantidade[]" step="any" value="1" required></div>
                    <div><label>Valor Unit. (R$) *</label><input type="text" name="item_valor_unit[]" class="valor-input" placeholder="0,00" required></div>
                    <div><label>Dimensão (opcional)</label><input type="text" name="item_dimensao[]" placeholder="Ex: 100x50"></div>
                    <div><label>Cores</label><input type="number" name="item_cores[]" step="1" value="4"></div>
                    <div><button type="button" onclick="this.closest('.item-row').remove()" style="background:#e74c3c;color:white;border:none;padding:8px;border-radius:5px;margin-top:28px;cursor:pointer;">🗑️</button></div>
                </div>
            `;
            document.getElementById('itens-container').appendChild(div);
        }}
        </script>
        </body>
        </html>
        '''
    except Exception as e:
        print("ERRO:", e)
        flash("Erro ao editar orçamento")
        return redirect(url_for('listar_orcamentos'))


@app.route('/calcular_data_entrega', methods=['POST'])
def calcular_data_entrega_api():
    from datetime import datetime
    dados = request.get_json()
    data_abertura = dados.get('data_abertura')
    dias = int(dados.get('dias', 7))
    
    data_inicio = datetime.strptime(data_abertura, "%Y-%m-%d")
    data_entrega = adicionar_dias_uteis(data_inicio, dias)
    
    return jsonify({{
        'data_entrega': data_entrega.strftime('%Y-%m-%d')
    }})

# MÓDULO DE RASTREAMENTO DE ENVIOS

def buscar_envios():
    try:
        url = f"{SUPABASE_URL}/rest/v1/envios?select=*,empresas(nome_empresa)&order=data_envio.desc"
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            return response.json()
        return []
    except Exception as e:
        print("Erro ao buscar envios:", e)
        return []

def criar_envio(tipo_envio, empresa_id, descricao, codigo_rastreio):
    try:
        url = f"{SUPABASE_URL}/rest/v1/envios"
        dados = {
            "tipo_envio": tipo_envio,
            "empresa_id": int(empresa_id),
            "descricao": descricao,
            "codigo_rastreio": codigo_rastreio,
            "data_envio": datetime.now().isoformat(),
            "status": "Enviado"
        }
        response = requests.post(url, json=dados, headers=headers)
        return response.status_code == 201
    except Exception as e:
        print("Erro ao criar envio:", e)
        return False

def marcar_entregue(id):
    try:
        url = f"{SUPABASE_URL}/rest/v1/envios?id=eq.{id}"
        dados = {
            "status": "Entregue",
            "data_entrega": datetime.now().isoformat()
        }
        response = requests.patch(url, json=dados, headers=headers)
        return response.status_code == 204
    except Exception as e:
        print("Erro ao marcar entrega:", e)
        return False

@app.route('/registrar_envio')
def registrar_envio():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    empresas = buscar_empresas()
    try:
        url_serv = f"{SUPABASE_URL}/rest/v1/servicos?select=id,codigo_servico,titulo&tipo=neq.Orçamento&order=codigo_servico.desc"
        response = requests.get(url_serv, headers=headers)
        servicos = response.json() if response.status_code == 200 else []
    except:
        servicos = []
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Registrar Envio - Rastreamento</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f5f7fa; color: #333; min-height: 100vh; padding: 0; margin: 0; }}
    .container {{ max-width: 900px; margin: 30px auto; background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0,0,0,0.1); overflow: hidden; }}
    .header {{ background: #2c3e50; color: white; text-align: center; padding: 30px; }}
    h1 {{ font-size: 28px; margin: 0; font-weight: 600; }}
    .user-info {{ background: #34495e; color: white; padding: 15px 20px; font-size: 15px; display: flex; justify-content: space-between; align-items: center; }}
    .form-container {{ padding: 30px; }}
    .form-container label {{ display: block; margin: 10px 0 5px 0; font-weight: 600; color: #2c3e50; }}
    .form-container input, .form-container select, .form-container textarea {{ width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 6px; font-size: 14px; }}
    .btn {{ padding: 12px 20px; background: #27ae60; color: white; border: none; border-radius: 8px; font-size: 16px; font-weight: 600; cursor: pointer; }}
    .back-link {{ display: inline-block; margin: 20px 30px; color: #3498db; text-decoration: none; font-weight: 500; }}
    .footer {{ text-align: center; padding: 20px; background: #ecf0f1; color: #7f8c8d; font-size: 13px; border-top: 1px solid #bdc3c7; }}
    </style>
    </head>
    <body>
    <div class="container">
    <div class="header"><h1>📦 Registrar Envio para Rastreamento</h1></div>
    <div class="user-info"><span>👤 {session['usuario']} ({session['nivel'].upper()})</span><a href="/logout">🚪 Sair</a></div>
    <a href="/envios" class="back-link">← Voltar à Lista de Envios</a>
    <form method="post" action="/salvar_envio" class="form-container">
    <div><label>Tipo de Envio *</label><select name="tipo_envio" id="tipo_envio" onchange="toggleServico()" required><option value="">Selecione</option><option value="Serviço">Serviço (vinculado a OS)</option><option value="Amostra">Amostra Grátis</option></select></div>
    <div id="servico-campo" style="display: none;"><label>Serviço *</label><select name="servico_id"><option value="">Selecione um serviço</option>{''.join(f'<option value="{s["id"]}">{s["codigo_servico"]} - {s["titulo"]}</option>' for s in servicos)}</select></div>
    <div><label>Cliente *</label><select name="empresa_id" required><option value="">Selecione uma empresa</option>{''.join(f'<option value="{e["id"]}">{e["nome_empresa"]}</option>' for e in empresas)}</select></div>
    <div><label>O que foi enviado? *</label><textarea name="descricao" rows="3" placeholder="Ex: Amostra de papel couché 300g" required></textarea></div>
    <div><label>Código de Rastreio *</label><input type="text" name="codigo_rastreio" placeholder="Ex: PQ123456789BR" required></div>
    <button type="submit" class="btn">💾 Registrar Envio</button>
    </form>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </div>
    <script>
    function toggleServico() {{ const tipo = document.getElementById('tipo_envio').value; const campo = document.getElementById('servico-campo'); if (tipo === 'Serviço') {{ campo.style.display = 'block'; }} else {{ campo.style.display = 'none'; }} }}
    </script>
    </body>
    </html>
    '''

@app.route('/salvar_envio', methods=['POST'])
def salvar_envio():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    tipo_envio = request.form.get('tipo_envio')
    empresa_id = request.form.get('empresa_id')
    descricao = request.form.get('descricao')
    codigo_rastreio = request.form.get('codigo_rastreio')
    if not tipo_envio or not empresa_id or not descricao or not codigo_rastreio:
        flash("Preencha todos os campos obrigatórios!")
        return redirect(url_for('registrar_envio'))
    try:
        dados = {
            "tipo_envio": tipo_envio,
            "empresa_id": int(empresa_id),
            "descricao": descricao,
            "codigo_rastreio": codigo_rastreio,
            "data_envio": datetime.now().isoformat()
        }
        if tipo_envio == "Serviço":
            servico_id = request.form.get('servico_id')
            if servico_id:
                dados["servico_id"] = int(servico_id)
                try:
                    url_serv = f"{SUPABASE_URL}/rest/v1/servicos?id=eq.{servico_id}"
                    requests.patch(url_serv, json={"status": "Enviado"}, headers=headers)
                except:
                    pass
        response = requests.post(f"{SUPABASE_URL}/rest/v1/envios", json=dados, headers=headers)
        if response.status_code == 201:
            flash("✅ Envio registrado com sucesso!")
        else:
            flash(f"❌ Erro ao registrar envio. Código: {response.status_code}")
    except Exception as e:
        print("Erro:", str(e))
        flash("❌ Erro de conexão.")
    return redirect(url_for('envios'))

@app.route('/envios')
def envios():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    lista_envios = buscar_envios()
    envios_enviados = []
    envios_entregues = []
    for e in lista_envios:
        if e.get('status') == "Entregue":
            envios_entregues.append(e)
        else:
            envios_enviados.append(e)
    html_enviados = ""
    for e in envios_enviados:
        data_entrega = format_data(e.get('data_entrega')) if e.get('data_entrega') else "—"
        html_enviados += f'''
        <tr>
        <td>{format_data(e.get('data_envio'))}</td>
        <td>{e['empresas']['nome_empresa'] if e.get('empresas') else '—'}</td>
        <td>{e['tipo_envio']}</td>
        <td>{e['descricao']}</td>
        <td>{e['codigo_rastreio']}</td>
        <td><span style="color: #e67e22; font-weight: bold;">{e['status']}</span></td>
        <td>{data_entrega}</td>
        <td>
        <div style="display: flex; gap: 5px; align-items: center;">
        <a href="https://www.linkcorreios.com.br/{e['codigo_rastreio']}" target="_blank" class="btn btn-blue">🔍 Rastrear</a>
        <a href="/editar_envio/{e['id']}" class="btn btn-yellow">✏️ Editar</a>
        <a href="/excluir_envio/{e['id']}" class="btn btn-red" onclick="return confirm('Tem certeza que deseja excluir?')">🗑️ Excluir</a>
        <a href="/marcar_entregue/{e['id']}" class="btn btn-green">✅ Entregue</a>
        </div>
        </td>
        </tr>
        '''
    html_entregues = ""
    for e in envios_entregues:
        data_entrega = format_data(e.get('data_entrega')) if e.get('data_entrega') else "—"
        html_entregues += f'''
        <tr>
        <td>{format_data(e.get('data_envio'))}</td>
        <td>{e['empresas']['nome_empresa'] if e.get('empresas') else '—'}</td>
        <td>{e['tipo_envio']}</td>
        <td>{e['descricao']}</td>
        <td>{e['codigo_rastreio']}</td>
        <td><span style="color: #27ae60; font-weight: bold;">{e['status']}</span></td>
        <td>{data_entrega}</td>
        <td>
        <div style="display: flex; gap: 5px; align-items: center;">
        <a href="https://www.linkcorreios.com.br/{e['codigo_rastreio']}" target="_blank" class="btn btn-blue">🔍 Rastrear</a>
        <a href="/editar_envio/{e['id']}" class="btn btn-yellow">✏️ Editar</a>
        <a href="/excluir_envio/{e['id']}" class="btn btn-red" onclick="return confirm('Tem certeza que deseja excluir?')">🗑️ Excluir</a>
        <span style="color: #95a5a6;">Já entregue</span>
        </div>
        </td>
        </tr>
        '''
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Rastreamento de Envios</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f5f7fa; color: #333; min-height: 100vh; padding: 0; margin: 0; }}
    .container {{ max-width: 1400px; margin: 30px auto; background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0,0,0,0.1); overflow: hidden; }}
    .header {{ background: #2c3e50; color: white; text-align: center; padding: 30px; }}
    h1 {{ font-size: 28px; margin: 0; font-weight: 600; }}
    .user-info {{ background: #34495e; color: white; padding: 15px 20px; font-size: 15px; display: flex; justify-content: space-between; align-items: center; }}
    .btn {{ padding: 8px 12px; border: none; border-radius: 6px; font-size: 14px; cursor: pointer; text-decoration: none; margin-right: 5px; }}
    .btn-blue {{ background: #3498db; color: white; }} .btn-green {{ background: #27ae60; color: white; }} .btn-yellow {{ background: #f39c12; color: white; }} .btn-red {{ background: #e74c3c; color: white; }}
    table {{ width: 100%; border-collapse: collapse; }}
    th, td {{ padding: 12px 15px; text-align: left; }}
    th {{ background: #ecf0f1; color: #2c3e50; font-weight: 600; }}
    .back-link {{ display: inline-block; margin: 20px 30px; color: #3498db; text-decoration: none; font-weight: 500; }}
    .footer {{ text-align: center; padding: 20px; background: #ecf0f1; color: #7f8c8d; font-size: 13px; border-top: 1px solid #bdc3c7; }}
    .section {{ padding: 20px 30px; }}
    .section-title {{ font-size: 20px; margin: 0 0 15px 0; color: #2c3e50; border-bottom: 1px solid #ddd; padding-bottom: 10px; }}
    </style>
    </head>
    <body>
    <div class="container">
    <div class="header"><h1>📦 Rastreamento de Envios</h1></div>
    <div class="user-info"><span>👤 {session['usuario']} ({session['nivel'].upper()})</span><a href="/logout">🚪 Sair</a></div>
    {MENU_FLUTUANTE}
    <a href="/registrar_envio" class="btn btn-green" style="display: inline-block; margin: 0 30px;">➕ Novo Envio</a>
    <div class="section">
    <h2 class="section-title">📬 Envios Enviados (Aguardando Confirmação)</h2>
    <table><thead><tr><th>Data Envio</th><th>Cliente</th><th>Tipo</th><th>O que foi enviado</th><th>Código Rastreio</th><th>Status</th><th>Data Entrega</th><th>Ações</th></tr></thead><tbody>{html_enviados if html_enviados else '<tr><td colspan="8" style="text-align: center;">Nenhum envio aguardando entrega</td></tr>'}</tbody></table>
    </div>
    <div class="section">
    <h2 class="section-title">✅ Envios Entregues</h2>
    <table><thead><tr><th>Data Envio</th><th>Cliente</th><th>Tipo</th><th>O que foi enviado</th><th>Código Rastreio</th><th>Status</th><th>Data Entrega</th><th>Ações</th></tr></thead><tbody>{html_entregues if html_entregues else '<tr><td colspan="8" style="text-align: center;">Nenhum envio entregue ainda</td></tr>'}</tbody></table>
    </div>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </div>
    </body>
    </html>
    '''

@app.route('/marcar_entregue/<int:id>')
def marcar_entregue_view(id):
    if 'usuario' not in session:
        return redirect(url_for('login'))
    if marcar_entregue(id):
        flash("✅ Entrega registrada com sucesso!")
    else:
        flash("❌ Erro ao registrar entrega.")
    return redirect(url_for('envios'))

@app.route('/editar_envio/<int:id>', methods=['GET', 'POST'])
def editar_envio(id):
    if 'usuario' not in session:
        return redirect(url_for('login'))
    try:
        url = f"{SUPABASE_URL}/rest/v1/envios?id=eq.{id}"
        response = requests.get(url, headers=headers)
        if response.status_code != 200 or not response.json():
            flash("Envio não encontrado.")
            return redirect(url_for('envios'))
        envio = response.json()[0]
        empresas = buscar_empresas()
        try:
            url_serv = f"{SUPABASE_URL}/rest/v1/servicos?select=id,codigo_servico,titulo&tipo=neq.Orçamento&order=codigo_servico.desc"
            response_serv = requests.get(url_serv, headers=headers)
            servicos = response_serv.json() if response_serv.status_code == 200 else []
        except:
            servicos = []
    except Exception as e:
        flash("Erro ao carregar envio.")
        return redirect(url_for('envios'))
    if request.method == 'POST':
        tipo_envio = request.form.get('tipo_envio')
        empresa_id = request.form.get('empresa_id')
        descricao = request.form.get('descricao')
        codigo_rastreio = request.form.get('codigo_rastreio')
        if not tipo_envio or not empresa_id or not descricao or not codigo_rastreio:
            flash("Preencha todos os campos obrigatórios!")
            return redirect(request.url)
        try:
            dados = {
                "tipo_envio": tipo_envio,
                "empresa_id": int(empresa_id),
                "descricao": descricao,
                "codigo_rastreio": codigo_rastreio
            }
            if tipo_envio == "Serviço":
                servico_id = request.form.get('servico_id')
                if servico_id:
                    dados["servico_id"] = int(servico_id)
            response = requests.patch(url, json=dados, headers=headers)
            if response.status_code == 204:
                flash("✅ Envio atualizado com sucesso!")
                return redirect(url_for('envios'))
            else:
                flash("❌ Erro ao atualizar envio.")
        except Exception as e:
            flash("❌ Erro de conexão.")
        return redirect(request.url)
    return f'''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Editar Envio</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Segoe+UI:wght@400;600&display=swap');
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f5f7fa; color: #333; min-height: 100vh; padding: 0; margin: 0; }}
    .container {{ max-width: 900px; margin: 30px auto; background: white; border-radius: 16px; box-shadow: 0 15px 35px rgba(0,0,0,0.1); overflow: hidden; }}
    .header {{ background: #2c3e50; color: white; text-align: center; padding: 30px; }}
    h1 {{ font-size: 28px; margin: 0; font-weight: 600; }}
    .user-info {{ background: #34495e; color: white; padding: 15px 20px; font-size: 15px; display: flex; justify-content: space-between; align-items: center; }}
    .form-container {{ padding: 30px; }}
    .form-container label {{ display: block; margin: 10px 0 5px 0; font-weight: 600; color: #2c3e50; }}
    .form-container input, .form-container select, .form-container textarea {{ width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 6px; font-size: 14px; }}
    .btn {{ padding: 12px 20px; background: #f39c12; color: white; border: none; border-radius: 8px; font-size: 16px; font-weight: 600; cursor: pointer; }}
    .back-link {{ display: inline-block; margin: 20px 30px; color: #3498db; text-decoration: none; font-weight: 500; }}
    .footer {{ text-align: center; padding: 20px; background: #ecf0f1; color: #7f8c8d; font-size: 13px; border-top: 1px solid #bdc3c7; }}
    </style>
    </head>
    <body>
    <div class="container">
    <div class="header"><h1>✏️ Editar Envio</h1></div>
    <div class="user-info"><span>👤 {session['usuario']} ({session['nivel'].upper()})</span><a href="/logout">🚪 Sair</a></div>
    <a href="/envios" class="back-link">← Voltar à lista</a>
    <form method="post" class="form-container">
    <div><label>Tipo de Envio *</label><select name="tipo_envio" id="tipo_envio" onchange="toggleServico()" required><option value="">Selecione</option><option value="Serviço" {"selected" if envio['tipo_envio'] == 'Serviço' else ""}>Serviço (vinculado a OS)</option><option value="Amostra" {"selected" if envio['tipo_envio'] == 'Amostra' else ""}>Amostra Grátis</option></select></div>
    <div id="servico-campo" style="display: {'block' if envio['tipo_envio'] == 'Serviço' else 'none'};"><label>Serviço *</label><select name="servico_id"><option value="">Selecione um serviço</option>{''.join(f'<option value="{s["id"]}" {"selected" if s.get("id") == envio.get("servico_id") else ""}>{s["codigo_servico"]} - {s["titulo"]}</option>' for s in servicos)}</select></div>
    <div><label>Cliente *</label><select name="empresa_id" required><option value="">Selecione uma empresa</option>{''.join(f'<option value="{e["id"]}" {"selected" if e["id"] == envio["empresa_id"] else ""}>{e["nome_empresa"]}</option>' for e in empresas)}</select></div>
    <div><label>O que foi enviado? *</label><textarea name="descricao" rows="3" required>{envio['descricao']}</textarea></div>
    <div><label>Código de Rastreio *</label><input type="text" name="codigo_rastreio" value="{envio['codigo_rastreio']}" required></div>
    <button type="submit" class="btn">💾 Salvar Alterações</button>
    </form>
    <div class="footer">Sistema de Gestão para Gráfica Rápida | © 2025</div>
    </div>
    <script>
    function toggleServico() {{ const tipo = document.getElementById('tipo_envio').value; const campo = document.getElementById('servico-campo'); if (tipo === 'Serviço') {{ campo.style.display = 'block'; }} else {{ campo.style.display = 'none'; }} }}
    </script>
    </body>
    </html>
    '''

@app.route('/excluir_envio/<int:id>')
def excluir_envio(id):
    if 'usuario' not in session:
        return redirect(url_for('login'))
    try:
        url = f"{SUPABASE_URL}/rest/v1/envios?id=eq.{id}"
        response = requests.delete(url, headers=headers)
        if response.status_code == 204:
            flash("🗑️ Envio excluído com sucesso!")
        else:
            flash("❌ Erro ao excluir envio.")
    except Exception as e:
        flash("❌ Erro de conexão.")
    return redirect(url_for('envios'))

# ========================
# Exportação e Importação Excel
# ========================
@app.route('/exportar_excel')
def exportar_excel():
    if 'usuario' not in session or session['nivel'] != 'administrador':
        flash("Acesso negado!")
        return redirect(url_for('clientes'))
    output = BytesIO()
    wb = Workbook()
    # Empresas
    ws_empresas = wb.active
    ws_empresas.title = "Empresas"
    empresas = buscar_empresas()
    ws_empresas.append(["ID", "Nome da Empresa", "CNPJ", "Responsável", "WhatsApp", "Email", "Endereço", "Bairro", "Cidade", "Estado", "CEP", "Número", "Entrega Endereço", "Entrega Número", "Entrega Bairro", "Entrega Cidade", "Entrega Estado", "Entrega CEP"])
    for e in empresas:
        ws_empresas.append([e.get("id"), e.get("nome_empresa", ""), e.get("cnpj", ""), e.get("responsavel", ""), e.get("whatsapp", ""), e.get("email", ""), e.get("endereco", ""), e.get("bairro", ""), e.get("cidade", ""), e.get("estado", ""), e.get("cep", ""), e.get("numero", ""), e.get("entrega_endereco", ""), e.get("entrega_numero", ""), e.get("entrega_bairro", ""), e.get("entrega_cidade", ""), e.get("entrega_estado", ""), e.get("entrega_cep", "")])
    for cell in ws_empresas[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D0E2FF", end_color="D0E2FF", fill_type="solid")
    # Materiais
    ws_materiais = wb.create_sheet("Materiais")
    materiais = buscar_materiais()
    ws_materiais.append(["ID", "Denominação", "Marca", "Grupo", "Unidade", "Valor Unitário", "Fornecedor", "Especificação"])
    for m in materiais:
        ws_materiais.append([m.get("id"), m.get("denominacao", ""), m.get("marca", ""), m.get("grupo_material", ""), m.get("unidade_medida", ""), m.get("valor_unitario", 0), m.get("fornecedor", ""), m.get("especificacao", "")])
    for cell in ws_materiais[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D0E2FF", end_color="D0E2FF", fill_type="solid")
    # Estoque (movimentações)
    ws_estoque = wb.create_sheet("Estoque")
    movimentacoes = buscar_movimentacoes_com_materiais()
    ws_estoque.append(["ID", "Material", "Tipo", "Quantidade", "Valor Unit.", "Valor Total", "Data", "Motivo", "Tamanho"])
    for m in movimentacoes:
        material_nome = m["materiais"]["denominacao"] if m.get("materiais") else "Excluído"
        ws_estoque.append([m.get("id"), material_nome, m.get("tipo", "").upper(), m.get("quantidade", 0), m.get("valor_unitario", 0), m.get("valor_total", 0), m.get("data_movimentacao", "")[:16].replace("T", " "), m.get("motivo", ""), m.get("tamanho", "")])
    for cell in ws_estoque[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D0E2FF", end_color="D0E2FF", fill_type="solid")
    # Usuários
    ws_usuarios = wb.create_sheet("Usuários")
    usuarios = buscar_usuarios()
    ws_usuarios.append(["ID", "Nome de Usuário", "Nível", "Telefone"])
    for u in usuarios:
        ws_usuarios.append([u.get("id"), u.get("nome de usuario", ""), u.get("NÍVEL", "").upper(), u.get("telefone", "")])
    for cell in ws_usuarios[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D0E2FF", end_color="D0E2FF", fill_type="solid")
    # Fornecedores
    ws_fornecedores = wb.create_sheet("Fornecedores")
    fornecedores = buscar_fornecedores()
    ws_fornecedores.append(["ID", "Nome", "CNPJ", "Contato", "Telefone", "Email", "Endereço"])
    for f in fornecedores:
        ws_fornecedores.append([f.get("id"), f.get("nome", ""), f.get("cnpj", ""), f.get("contato", ""), f.get("telefone", ""), f.get("email", ""), f.get("endereco", "")])
    for cell in ws_fornecedores[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D0E2FF", end_color="D0E2FF", fill_type="solid")
    # Serviços e Orçamentos
    ws_servicos = wb.create_sheet("Serviços")
    try:
        url_serv = f"{SUPABASE_URL}/rest/v1/servicos?select=*,empresas(nome_empresa)&order=codigo_servico.desc"
        response = requests.get(url_serv, headers=headers)
        servicos = response.json() if response.status_code == 200 else []
    except:
        servicos = []
    ws_servicos.append(["ID", "Código", "Título", "Cliente", "Tipo", "Status", "Quantidade", "Dimensão", "Nº Cores", "Valor Cobrado", "Data Abertura", "Previsão Entrega", "Observações"])
    for s in servicos:
        ws_servicos.append([s.get("id"), s.get("codigo_servico", ""), s.get("titulo", ""), s.get("empresas", {}).get("nome_empresa", "") if s.get("empresas") else "", s.get("tipo", ""), s.get("status", ""), s.get("quantidade", ""), s.get("dimensao", ""), s.get("numero_cores", ""), s.get("valor_cobrado", 0), s.get("data_abertura", "")[:10] if s.get("data_abertura") else "", s.get("previsao_entrega", "")[:10] if s.get("previsao_entrega") else "", s.get("observacoes", "")])
    for cell in ws_servicos[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D0E2FF", end_color="D0E2FF", fill_type="solid")
    # Envios (Rastreamento)
    ws_envios = wb.create_sheet("Envios")
    try:
        url_envios = f"{SUPABASE_URL}/rest/v1/envios?select=*,empresas(nome_empresa)&order=data_envio.desc"
        response = requests.get(url_envios, headers=headers)
        envios = response.json() if response.status_code == 200 else []
    except:
        envios = []
    ws_envios.append(["ID", "Tipo", "Cliente", "Descrição", "Código Rastreio", "Data Envio", "Data Entrega", "Status"])
    for e in envios:
        ws_envios.append([e.get("id"), e.get("tipo_envio", ""), e.get("empresas", {}).get("nome_empresa", "") if e.get("empresas") else "", e.get("descricao", ""), e.get("codigo_rastreio", ""), e.get("data_envio", "")[:16].replace("T", " ") if e.get("data_envio") else "", e.get("data_entrega", "")[:16].replace("T", " ") if e.get("data_entrega") else "", e.get("status", "")])
    for cell in ws_envios[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D0E2FF", end_color="D0E2FF", fill_type="solid")
    # Ajustar largura das colunas
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="backup_sistema_grafica.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/importar_excel', methods=['GET', 'POST'])
def importar_excel():
    if 'usuario' not in session or session['nivel'] != 'administrador':
        flash("Acesso negado!")
        return redirect(url_for('clientes'))
    if request.method == 'POST':
        if 'arquivo' not in request.files:
            flash("Nenhum arquivo enviado.")
            return redirect(request.url)
        arquivo = request.files['arquivo']
        if arquivo.filename == '':
            flash("Nenhum arquivo selecionado.")
            return redirect(request.url)
        if not arquivo.filename.endswith(('.xlsx', '.xls')):
            flash("Apenas arquivos Excel (.xlsx) são permitidos.")
            return redirect(request.url)
        try:
            df = pd.read_excel(arquivo, sheet_name=None)
            log = []
            if 'Empresas' in df:
                for _, row in df['Empresas'].iterrows():
                    try:
                        criar_empresa(nome=row['Nome da Empresa'], cnpj=row.get('CNPJ', ''), responsavel=row.get('Responsável', ''), telefone=row.get('Telefone', ''), whatsapp=row.get('WhatsApp', ''), email=row.get('Email', ''), endereco=row.get('Endereço', ''), bairro=row.get('Bairro', ''), cidade=row.get('Cidade', ''), estado=row.get('Estado', ''), cep=row.get('CEP', ''), numero=row.get('Número', ''), entrega_endereco=row.get('Entrega Endereço', ''), entrega_numero=row.get('Entrega Número', ''), entrega_bairro=row.get('Entrega Bairro', ''), entrega_cidade=row.get('Entrega Cidade', ''), entrega_estado=row.get('Entrega Estado', ''), entrega_cep=row.get('Entrega CEP', ''))
                        log.append(f"✅ Empresa '{row['Nome da Empresa']}' importada.")
                    except Exception as e:
                        log.append(f"❌ Erro ao importar empresa: {str(e)}")
            if 'Materiais' in df:
                for _, row in df['Materiais'].iterrows():
                    try:
                        url = f"{SUPABASE_URL}/rest/v1/materiais"
                        dados = {"denominacao": row['Denominação'], "marca": row.get('Marca', ''), "grupo_material": row.get('Grupo', ''), "unidade_medida": row.get('Unidade', 'unidade'), "valor_unitario": float(row.get('Valor Unitário', 0)), "fornecedor": row.get('Fornecedor', ''), "especificacao": row.get('Especificação', '')}
                        response = requests.post(url, json=dados, headers=headers)
                        if response.status_code == 201:
                            log.append(f"✅ Material '{row['Denominação']}' cadastrado.")
                        else:
                            log.append(f"❌ Erro ao cadastrar material: {response.text}")
                    except Exception as e:
                        log.append(f"❌ Erro ao processar material: {str(e)}")
            if 'Fornecedores' in df:
                for _, row in df['Fornecedores'].iterrows():
                    try:
                        criar_fornecedor(nome=row['Nome'], cnpj=row.get('CNPJ', ''), contato=row.get('Contato', ''), telefone=row.get('Telefone', ''), email=row.get('Email', ''), endereco=row.get('Endereço', ''))
                        log.append(f"✅ Fornecedor '{row['Nome']}' cadastrado.")
                    except Exception as e:
                        log.append(f"❌ Erro ao cadastrar fornecedor: {str(e)}")
            if 'Serviços' in df:
                for _, row in df['Serviços'].iterrows():
                    try:
                        empresa_id = None
                        if row.get('Cliente'):
                            empresas = buscar_empresas()
                            for emp in empresas:
                                if emp.get('nome_empresa') == row['Cliente']:
                                    empresa_id = emp['id']
                                    break
                        url = f"{SUPABASE_URL}/rest/v1/servicos"
                        dados = {"codigo_servico": row['Código'], "titulo": row['Título'], "empresa_id": empresa_id, "tipo": row['Tipo'], "status": row['Status'], "quantidade": row.get('Quantidade', ''), "dimensao": row.get('Dimensão', ''), "numero_cores": row.get('Nº Cores'), "valor_cobrado": float(row.get('Valor Cobrado', 0)), "data_abertura": row.get('Data Abertura'), "previsao_entrega": row.get('Previsão Entrega'), "observacoes": row.get('Observações', '')}
                        response = requests.post(url, json=dados, headers=headers)
                        if response.status_code == 201:
                            log.append(f"✅ Serviço '{row['Código']}' cadastrado.")
                        else:
                            log.append(f"❌ Erro ao cadastrar serviço: {response.text}")
                    except Exception as e:
                        log.append(f"❌ Erro ao processar serviço: {str(e)}")
            if 'Envios' in df:
                for _, row in df['Envios'].iterrows():
                    try:
                        empresa_id = None
                        if row.get('Cliente'):
                            empresas = buscar_empresas()
                            for emp in empresas:
                                if emp.get('nome_empresa') == row['Cliente']:
                                    empresa_id = emp['id']
                                    break
                        if not empresa_id:
                            log.append(f"⚠️ Empresa '{row['Cliente']}' não encontrada para envio. Pulando...")
                            continue
                        url = f"{SUPABASE_URL}/rest/v1/envios"
                        dados = {"tipo_envio": row['Tipo'], "empresa_id": empresa_id, "descricao": row['Descrição'], "codigo_rastreio": row['Código Rastreio'], "data_envio": row['Data Envio'], "data_entrega": row.get('Data Entrega'), "status": row['Status']}
                        response = requests.post(url, json=dados, headers=headers)
                        if response.status_code == 201:
                            log.append(f"✅ Envio '{row['Código Rastreio']}' cadastrado.")
                        else:
                            log.append(f"❌ Erro ao cadastrar envio: {response.text}")
                    except Exception as e:
                        log.append(f"❌ Erro ao processar envio: {str(e)}")
            return render_template('importar_excel.html', log=log)
        except Exception as e:
            flash(f"❌ Erro ao ler o arquivo: {str(e)}")
            return redirect(request.url)
    return render_template('importar_excel.html', log=None)

# ========================
# PDF DO ORÇAMENTO - Layout Profissional
# ========================
@app.route('/pdf_orcamento/<int:id>')
def pdf_orcamento(id):
    if 'usuario' not in session:
        return redirect(url_for('login'))
    
    try:
        url = SUPABASE_URL + "/rest/v1/servicos?id=eq." + str(id) + "&select=*,empresas(nome_empresa,cnpj,telefone,email),itens_orcamento(*)"
        resp = requests.get(url, headers=headers)
        orc = resp.json()[0]
        
        emp = orc.get('empresas', {})
        cliente = emp.get('nome_empresa', 'Cliente')
        logo_url = "https://i.postimg.cc/RVqcJzzQ/logo.png"
        
        def fmt(data):
            return "/".join(str(data)[:10].split("-")[::-1]) if data and len(str(data)) >= 10 else "—"
        
        data_abr = fmt(orc.get('data_abertura'))
        total = sum(float(i.get('valor_total', 0) or 0) for i in orc.get('itens_orcamento', []))
        
        # HTML Profissional - Layout Moderno A4
        html = f'''<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<title>Orçamento {orc.get('codigo_servico','')}</title>
<style>
@page {{ 
    size: A4; 
    margin: 2cm; 
}}
* {{ 
    margin: 0; 
    padding: 0; 
    box-sizing: border-box; 
}}
body {{ 
    font-family: Arial, sans-serif; 
    font-size: 12px; 
    color: #333; 
    line-height: 1.6;
}}
.container {{ 
    max-width: 794px; 
    margin: 0 auto; 
}}
/* Header */
.header {{ 
    text-align: center;
    margin-bottom: 30px;
    padding-bottom: 20px;
    border-bottom: 4px solid #2c3e50;
}}
.logo {{ 
    max-width: 200px; 
    margin-bottom: 15px; 
}}
.titulo {{ 
    font-size: 36px; 
    font-weight: bold; 
    color: #2c3e50; 
    letter-spacing: 5px;
    text-transform: uppercase;
    margin: 10px 0;
}}
.codigo {{
    font-size: 20px;
    font-weight: bold;
    color: #7f8c8d;
    background: #ecf0f1;
    padding: 8px 25px;
    border-radius: 5px;
    display: inline-block;
}}
/* Info Section */
.info-section {{ 
    display: flex;
    gap: 25px;
    margin-bottom: 30px;
}}
.info-box {{ 
    flex: 1;
    background: #f8f9fa;
    padding: 20px;
    border-radius: 8px;
}}
.info-label {{ 
    font-size: 11px; 
    text-transform: uppercase; 
    color: #7f8c8d; 
    font-weight: bold; 
    margin-bottom: 12px;
    letter-spacing: 1px;
}}
.cliente-nome {{ 
    font-size: 20px; 
    color: #2c3e50; 
    font-weight: bold; 
    margin-bottom: 15px;
}}
.info-detail {{
    font-size: 12px;
    color: #555;
    margin: 8px 0;
}}
/* Delivery Box */
.entrega-box {{ 
    background: #e8f5e9; 
    border-left: 5px solid #27ae60; 
    padding: 15px 20px; 
    margin: 25px 0;
    border-radius: 5px;
}}
.entrega-label {{
    font-size: 12px; 
    color: #27ae60; 
    font-weight: bold; 
    margin-bottom: 8px;
    text-transform: uppercase;
}}
.entrega-text {{ 
    font-size: 13px; 
    color: #27ae60; 
    line-height: 1.6;
}}
/* Items Table */
.items-section {{ 
    margin: 30px 0;
}}
.section-title {{
    font-size: 16px; 
    font-weight: bold; 
    color: #2c3e50; 
    margin-bottom: 15px;
    text-transform: uppercase;
    border-bottom: 2px solid #ecf0f1;
    padding-bottom: 8px;
}}
.items-table {{ 
    width: 100%; 
    border-collapse: collapse;
    margin: 20px 0;
}}
.items-table thead {{
    background: #2c3e50; 
    color: white;
}}
.items-table th {{ 
    padding: 14px 12px; 
    text-align: left; 
    font-size: 11px;
    font-weight: bold;
    text-transform: uppercase;
}}
.items-table td {{ 
    padding: 16px 12px; 
    border-bottom: 2px solid #ecf0f1; 
    font-size: 13px;
}}
.items-table tbody tr:nth-child(even) {{
    background: #f8f9fa;
}}
.text-center {{ text-align: center; }}
.text-right {{ text-align: right; }}
/* Total Section */
.total-section {{
    margin-top: 40px;
    text-align: right;
}}
.total-box {{ 
    background: #2c3e50; 
    color: white; 
    padding: 25px 35px; 
    border-radius: 8px; 
    display: inline-block;
    min-width: 280px;
}}
.total-label {{ 
    font-size: 13px; 
    text-transform: uppercase; 
    margin-bottom: 12px;
    opacity: 0.9;
}}
.total-value {{ 
    font-size: 36px; 
    font-weight: bold;
    letter-spacing: 2px;
}}
/* Footer */
.footer {{
    margin-top: 50px;
    padding-top: 20px;
    border-top: 2px solid #ecf0f1;
    text-align: center;
    font-size: 10px;
    color: #95a5a6;
}}
.footer strong {{
    color: #2c3e50;
    font-size: 11px;
}}
</style>
</head>
<body>
<div class="container">
    <!-- Header com Logo Centralizado -->
    <div class="header">
        <img src="{logo_url}" class="logo" alt="Liraprint" onerror="this.style.display='none'">
        <div class="titulo">ORÇAMENTO</div>
        <div class="codigo">{orc.get('codigo_servico','—')}</div>
    </div>
    
    <!-- Informações do Cliente e Dados -->
    <div class="info-section">
        <div class="info-box">
            <div class="info-label">Cliente</div>
            <div class="cliente-nome">{cliente}</div>
            <div class="info-detail"><strong>CNPJ:</strong> {emp.get('cnpj','—')}</div>
            <div class="info-detail"><strong>Tel:</strong> {emp.get('telefone','—')}</div>
            <div class="info-detail"><strong>Email:</strong> {emp.get('email','—')}</div>
        </div>
        <div class="info-box">
            <div class="info-label">Dados do Orçamento</div>
            <div class="info-detail"><strong>Emissão:</strong> {data_abr}</div>
            <div class="info-detail"><strong>Status:</strong> {orc.get('status','Pendente')}</div>
            <div class="info-detail"><strong>Validade:</strong> 30 dias</div>
        </div>
    </div>
    
    <!-- Prazo de Entrega -->
    <div class="entrega-box">
        <div class="entrega-label">📅 Prazo de Entrega</div>
        <div class="entrega-text">
            Dias úteis após aprovação da arte.<br>
            <em style="font-size: 11px;">* A contagem inicia após aprovação.</em>
        </div>
    </div>
    
    <!-- Itens Orçados -->
    <div class="items-section">
        <div class="section-title">Itens Orçados</div>
        <table class="items-table">
            <thead>
                <tr>
                    <th width="5%">#</th>
                    <th width="45%">Descrição</th>
                    <th width="15%" class="text-center">Qtd</th>
                    <th width="20%">Dimensão</th>
                    <th width="15%" class="text-right">Valor</th>
                </tr>
            </thead>
            <tbody>'''
        
        itens = orc.get('itens_orcamento', [])
        if itens and len(itens) > 0:
            for i, item in enumerate(itens):
                html += f'''
                <tr>
                    <td class="text-center">{i+1}</td>
                    <td><strong>{item.get('titulo','—')}</strong></td>
                    <td class="text-center">{item.get('quantidade','1')}</td>
                    <td>{item.get('dimensao','—')}</td>
                    <td class="text-right" style="font-size: 14px; font-weight: 600;">R$ {float(item.get('valor_total',0) or 0):.2f}</td>
                </tr>'''
        else:
            html += '''
                <tr>
                    <td colspan="5" style="text-align: center; color: #95a5a6; padding: 40px; font-size: 14px;">
                        Nenhum item adicionado
                    </td>
                </tr>'''
        
        html += f'''
            </tbody>
        </table>
    </div>
    
    <!-- Total -->
    <div class="total-section">
        <div class="total-box">
            <div class="total-label">Total do Orçamento</div>
            <div class="total-value">R$ {total:.2f}</div>
        </div>
    </div>
    
    <!-- Footer -->
    <div class="footer">
        <div style="margin-bottom: 8px;">
            <strong>LIRAPRINT</strong> - Gráfica Rápida<br>
            R. Dr. Roberto Fernandes, 81 - Jardim Palmira - Guarulhos/SP - CEP: 07076-070
        </div>
        <div>
            Documento gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}<br>
            Este orçamento é válido por 30 dias a partir da data de emissão
        </div>
    </div>
</div>
</body>
</html>'''
        
        pdf = pdfkit.from_string(html, False, options={"quiet": "", "encoding": "UTF-8", "page-size": "A4", "dpi": 300})
        return send_file(BytesIO(pdf), as_attachment=True, download_name=f"Orcamento_{orc.get('codigo_servico','')}.pdf", mimetype="application/pdf")
        
    except Exception as e:
        print("ERRO PDF:", str(e))
        import traceback
        traceback.print_exc()
        flash("❌ Erro ao gerar PDF: " + str(e))
        return redirect(url_for('listar_orcamentos'))


# ========================
# INICIAR O APP (FORA DE QUALQUER FUNÇÃO!)
# ========================
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)